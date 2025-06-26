#!/usr/bin/env python3
"""
Excel Explorer - Quick Integration Test
Validates that core Phase 1 components work together.
"""

import sys
import tempfile
from pathlib import Path

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / "src"))

def create_test_excel_file():
    """Create a simple test Excel file for validation"""
    try:
        import openpyxl
        
        # Create a simple workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Data"
        
        # Add some test data
        headers = ["Name", "Age", "Department", "Salary"]
        test_data = [
            ["Alice", 30, "Engineering", 75000],
            ["Bob", 25, "Marketing", 55000],
            ["Charlie", 35, "Engineering", 85000],
            ["Diana", 28, "Sales", 60000],
            ["Eve", 32, "Engineering", 80000]
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Write data
        for row, data_row in enumerate(test_data, 2):
            for col, value in enumerate(data_row, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Add a formula
        ws.cell(row=7, column=4, value="=AVERAGE(D2:D6)")
        
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(temp_file.name)
        wb.close()
        
        return temp_file.name
        
    except Exception as e:
        print(f"Error creating test file: {e}")
        return None


def test_core_components():
    """Test core Phase 1 components"""
    print("🧪 Testing Phase 1 Core Components")
    print("=" * 40)
    
    try:
        # Test 1: Import core modules
        print("1. Testing imports...")
        
        from src.utils.memory_manager import MemoryManager, get_memory_manager
        from src.utils.error_handler import initialize_error_handler, ExcelAnalysisError
        from src.core.analysis_context import AnalysisContext, AnalysisConfig
        from src.core.module_result import ModuleResult, ResultStatus
        from src.core.base_analyzer import BaseAnalyzer
        from src.modules.health_checker import HealthChecker
        from src.modules.structure_mapper import StructureMapper
        print("   ✅ All imports successful")
        
        # Test 2: Memory Manager
        print("2. Testing Memory Manager...")
        memory_manager = get_memory_manager()
        usage = memory_manager.get_current_usage()
        assert 'current_mb' in usage
        assert usage['current_mb'] > 0
        print("   ✅ Memory manager working")
        
        # Test 3: Error Handler
        print("3. Testing Error Handler...")
        error_handler = initialize_error_handler()
        test_error = ExcelAnalysisError("Test error", module_name="test")
        error_handler.handle_error(test_error)
        print("   ✅ Error handler working")
        
        # Test 4: Create test file
        print("4. Creating test Excel file...")
        test_file = create_test_excel_file()
        if not test_file:
            print("   ❌ Failed to create test file")
            return False
        print(f"   ✅ Test file created: {Path(test_file).name}")
        
        # Test 5: Analysis Context
        print("5. Testing Analysis Context...")
        config = AnalysisConfig(max_memory_mb=1024)
        with AnalysisContext(test_file, config) as context:
            assert context.file_path.exists()
            assert context.file_metadata.file_size_mb > 0
            print("   ✅ Analysis context working")
        
        # Test 6: Health Checker
        print("6. Testing Health Checker...")
        health_checker = HealthChecker()
        with AnalysisContext(test_file, config) as context:
            result = health_checker.analyze(context)
            assert result.is_successful
            assert result.data.file_accessible
            print("   ✅ Health checker working")
        
        # Test 7: Structure Mapper
        print("7. Testing Structure Mapper...")
        structure_mapper = StructureMapper()
        with AnalysisContext(test_file, config) as context:
            # Register health check result first (dependency)
            context.register_module_result("health_checker", result, True)
            
            result = structure_mapper.analyze(context)
            assert result.is_successful
            assert result.data.worksheet_count > 0
            print("   ✅ Structure mapper working")
        
        # Cleanup
        Path(test_file).unlink()
        print("\n🎉 All Phase 1 core components working correctly!")
        return True
        
    except Exception as e:
        print(f"   ❌ Test failed: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_orchestrator():
    """Test the orchestrator integration"""
    print("\n🔄 Testing Orchestrator Integration")
    print("=" * 40)
    
    try:
        # Create test file
        test_file = create_test_excel_file()
        if not test_file:
            print("❌ Failed to create test file")
            return False
        
        # Test orchestrator
        from src.core.orchestrator import ExcelExplorer
        
        print("1. Initializing Excel Explorer...")
        explorer = ExcelExplorer()
        print("   ✅ Explorer initialized")
        
        print("2. Running analysis...")
        results = explorer.analyze_file(test_file)
        
        # Validate results
        assert 'analysis_metadata' in results
        assert 'module_results' in results
        print("   ✅ Analysis completed")
        
        # Check specific results
        metadata = results['analysis_metadata']
        print(f"   Success Rate: {metadata.get('success_rate', 0):.1%}")
        print(f"   Modules Executed: {len(metadata.get('modules_executed', []))}")
        
        # Cleanup
        Path(test_file).unlink()
        print("\n🎉 Orchestrator integration working correctly!")
        return True
        
    except Exception as e:
        print(f"❌ Orchestrator test failed: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Run all tests"""
    print("Excel Explorer Phase 1 Integration Test")
    print("=" * 50)
    
    # Test core components
    core_success = test_core_components()
    
    # Test orchestrator
    orchestrator_success = test_orchestrator()
    
    # Final result
    print("\n" + "=" * 50)
    if core_success and orchestrator_success:
        print("🎉 ALL TESTS PASSED - Phase 1 Ready for Use!")
        return 0
    else:
        print("❌ Some tests failed - Review implementation")
        return 1


if __name__ == "__main__":
    sys.exit(main())
