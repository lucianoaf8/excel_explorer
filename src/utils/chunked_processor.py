"""
Chunked Sheet Processor - Memory-efficient Excel data processing utility
Provides streaming analysis with configurable batch sizes and adaptive memory management.
"""

import time
import logging
from typing import Iterator, List, Dict, Any, Optional, Callable, Tuple, Union
from dataclasses import dataclass
from enum import Enum
import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .memory_manager import get_memory_manager, MemoryManager
from .error_handler import ExcelAnalysisError, ErrorSeverity, ErrorCategory


class ChunkingStrategy(Enum):
    """Different chunking strategies for processing"""
    ROW_BASED = "row_based"
    COLUMN_BASED = "column_based"
    CELL_BASED = "cell_based"
    ADAPTIVE = "adaptive"


@dataclass
class ChunkConfig:
    """Configuration for chunked processing"""
    chunk_size_rows: int = 10000
    chunk_size_columns: int = 100
    max_memory_mb: float = 512.0
    strategy: ChunkingStrategy = ChunkingStrategy.ROW_BASED
    enable_progress_tracking: bool = True
    intermediate_save: bool = False
    intermediate_save_path: Optional[Path] = None


@dataclass
class ProcessingStats:
    """Statistics for chunk processing operation"""
    total_chunks: int = 0
    processed_chunks: int = 0
    total_rows: int = 0
    total_columns: int = 0
    processing_time: float = 0.0
    memory_peak_mb: float = 0.0
    errors_encountered: int = 0
    warnings_generated: int = 0


class ChunkProcessor:
    """Base class for chunk processing operations"""
    
    def process_chunk(self, chunk_data: pd.DataFrame, chunk_index: int, 
                     chunk_metadata: Dict[str, Any]) -> Any:
        """Process a single chunk of data
        
        Args:
            chunk_data: DataFrame with chunk data
            chunk_index: Zero-based chunk index
            chunk_metadata: Additional metadata about the chunk
            
        Returns:
            Processing result for this chunk
        """
        raise NotImplementedError


class ChunkedSheetProcessor:
    """Memory-efficient worksheet processor with configurable chunking strategies"""
    
    def __init__(self, config: Optional[ChunkConfig] = None):
        self.config = config or ChunkConfig()
        self.memory_manager = get_memory_manager()
        self.logger = logging.getLogger(__name__)
        self.stats = ProcessingStats()
        self._cancelled = False
        
    def process_worksheet(self, worksheet: Worksheet, processor: ChunkProcessor,
                         data_region: Optional[Dict[str, int]] = None) -> List[Any]:
        """Process worksheet using chunked strategy
        
        Args:
            worksheet: openpyxl worksheet to process
            processor: ChunkProcessor instance to handle chunks
            data_region: Optional data region bounds
            
        Returns:
            List of chunk processing results
        """
        try:
            self._reset_stats()
            start_time = time.time()
            
            # Detect data region if not provided
            if not data_region:
                data_region = self._detect_data_region(worksheet)
                if not data_region:
                    self.logger.warning("No data region detected in worksheet")
                    return []
            
            # Calculate chunking parameters
            chunk_params = self._calculate_chunk_parameters(data_region)
            self.stats.total_chunks = chunk_params['total_chunks']
            self.stats.total_rows = data_region['max_row'] - data_region['min_row'] + 1
            self.stats.total_columns = data_region['max_col'] - data_region['min_col'] + 1
            
            self.logger.info(
                f"Processing {self.stats.total_rows} rows × {self.stats.total_columns} columns "
                f"in {self.stats.total_chunks} chunks using {self.config.strategy.value} strategy"
            )
            
            # Process chunks
            results = []
            for chunk_data, chunk_index, chunk_metadata in self._generate_chunks(worksheet, data_region, chunk_params):
                if self._cancelled:
                    self.logger.warning("Processing cancelled")
                    break
                
                try:
                    # Memory pressure check
                    if self.memory_manager.check_memory_pressure():
                        self._adapt_chunk_size()
                    
                    # Process chunk
                    result = processor.process_chunk(chunk_data, chunk_index, chunk_metadata)
                    results.append(result)
                    
                    self.stats.processed_chunks += 1
                    
                    # Progress tracking
                    if self.config.enable_progress_tracking:
                        progress = (self.stats.processed_chunks / self.stats.total_chunks) * 100
                        self.logger.debug(f"Chunk {chunk_index + 1}/{self.stats.total_chunks} processed ({progress:.1f}%)")
                    
                except Exception as e:
                    self.stats.errors_encountered += 1
                    self.logger.error(f"Error processing chunk {chunk_index}: {e}")
                    raise ExcelAnalysisError(
                        f"Chunk processing failed: {e}",
                        severity=ErrorSeverity.MEDIUM,
                        category=ErrorCategory.DATA_CORRUPTION,
                        module_name="ChunkedSheetProcessor"
                    )
            
            # Update final stats
            self.stats.processing_time = time.time() - start_time
            self.stats.memory_peak_mb = self.memory_manager.get_current_usage()['peak_mb']
            
            self.logger.info(
                f"Chunked processing complete: {self.stats.processed_chunks}/{self.stats.total_chunks} chunks "
                f"in {self.stats.processing_time:.1f}s"
            )
            
            return results
            
        except Exception as e:
            self.logger.error(f"Worksheet processing failed: {e}")
            raise
    
    def process_multiple_sheets(self, workbook_path: Path, processor: ChunkProcessor,
                               sheet_names: Optional[List[str]] = None) -> Dict[str, List[Any]]:
        """Process multiple worksheets with memory management
        
        Args:
            workbook_path: Path to Excel workbook
            processor: ChunkProcessor instance
            sheet_names: Optional list of sheet names to process
            
        Returns:
            Dict mapping sheet names to processing results
        """
        results = {}
        
        try:
            with openpyxl.load_workbook(workbook_path, read_only=True, data_only=True) as wb:
                target_sheets = sheet_names if sheet_names else wb.sheetnames
                
                for sheet_name in target_sheets:
                    if self._cancelled:
                        break
                    
                    try:
                        self.logger.info(f"Processing sheet: {sheet_name}")
                        worksheet = wb[sheet_name]
                        sheet_results = self.process_worksheet(worksheet, processor)
                        results[sheet_name] = sheet_results
                        
                        # Memory cleanup between sheets
                        if self.memory_manager.check_memory_pressure():
                            self.memory_manager.clear_cache()
                        
                    except Exception as e:
                        self.logger.error(f"Failed to process sheet {sheet_name}: {e}")
                        results[sheet_name] = []
        
        except Exception as e:
            self.logger.error(f"Failed to process workbook {workbook_path}: {e}")
            raise
        
        return results
    
    def cancel_processing(self) -> None:
        """Cancel current processing operation"""
        self._cancelled = True
        self.logger.info("Processing cancellation requested")
    
    def get_processing_stats(self) -> ProcessingStats:
        """Get current processing statistics"""
        return self.stats
    
    def _reset_stats(self) -> None:
        """Reset processing statistics"""
        self.stats = ProcessingStats()
        self._cancelled = False
    
    def _detect_data_region(self, worksheet: Worksheet) -> Optional[Dict[str, int]]:
        """Detect the primary data region in worksheet"""
        if not worksheet.max_row or not worksheet.max_column:
            return None
        
        # Find actual data boundaries (not just used range)
        min_row, max_row = None, None
        min_col, max_col = None, None
        
        # Sample-based detection for performance
        sample_step = max(1, worksheet.max_row // 100)
        
        for row_idx in range(1, worksheet.max_row + 1, sample_step):
            row_values = [
                worksheet.cell(row=row_idx, column=col).value
                for col in range(1, min(worksheet.max_column + 1, 50))  # Limit column sampling
            ]
            
            if any(val is not None and str(val).strip() for val in row_values):
                if min_row is None:
                    min_row = row_idx
                max_row = row_idx
        
        # Find column boundaries
        if min_row and max_row:
            for col_idx in range(1, worksheet.max_column + 1):
                sample_rows = list(range(min_row, min(min_row + 20, max_row + 1)))
                col_values = [
                    worksheet.cell(row=row, column=col_idx).value
                    for row in sample_rows
                ]
                
                if any(val is not None and str(val).strip() for val in col_values):
                    if min_col is None:
                        min_col = col_idx
                    max_col = col_idx
        
        if all(v is not None for v in [min_row, max_row, min_col, max_col]):
            return {
                'min_row': min_row,
                'max_row': max_row,
                'min_col': min_col,
                'max_col': max_col
            }
        
        return None
    
    def _calculate_chunk_parameters(self, data_region: Dict[str, int]) -> Dict[str, Any]:
        """Calculate optimal chunking parameters"""
        total_rows = data_region['max_row'] - data_region['min_row'] + 1
        total_cols = data_region['max_col'] - data_region['min_col'] + 1
        
        if self.config.strategy == ChunkingStrategy.ROW_BASED:
            chunk_size = self.config.chunk_size_rows
            total_chunks = (total_rows + chunk_size - 1) // chunk_size
            
        elif self.config.strategy == ChunkingStrategy.COLUMN_BASED:
            chunk_size = self.config.chunk_size_columns
            total_chunks = (total_cols + chunk_size - 1) // chunk_size
            
        elif self.config.strategy == ChunkingStrategy.ADAPTIVE:
            # Adaptive strategy based on memory usage
            estimated_memory_per_cell = 50  # bytes
            total_cells = total_rows * total_cols
            estimated_memory_mb = (total_cells * estimated_memory_per_cell) / 1_048_576
            
            if estimated_memory_mb <= self.config.max_memory_mb:
                # Process all at once
                chunk_size = total_rows
                total_chunks = 1
            else:
                # Calculate optimal chunk size
                target_cells_per_chunk = (self.config.max_memory_mb * 1_048_576) // estimated_memory_per_cell
                chunk_size = max(1, int(target_cells_per_chunk / total_cols))
                total_chunks = (total_rows + chunk_size - 1) // chunk_size
        
        else:  # CELL_BASED
            # Very conservative approach for problematic files
            chunk_size = min(100, total_rows)
            total_chunks = (total_rows + chunk_size - 1) // chunk_size
        
        return {
            'chunk_size': chunk_size,
            'total_chunks': total_chunks,
            'strategy': self.config.strategy
        }
    
    def _generate_chunks(self, worksheet: Worksheet, data_region: Dict[str, int], 
                        chunk_params: Dict[str, Any]) -> Iterator[Tuple[pd.DataFrame, int, Dict[str, Any]]]:
        """Generate data chunks based on strategy"""
        chunk_size = chunk_params['chunk_size']
        strategy = chunk_params['strategy']
        
        if strategy in [ChunkingStrategy.ROW_BASED, ChunkingStrategy.ADAPTIVE, ChunkingStrategy.CELL_BASED]:
            yield from self._generate_row_chunks(worksheet, data_region, chunk_size)
        elif strategy == ChunkingStrategy.COLUMN_BASED:
            yield from self._generate_column_chunks(worksheet, data_region, chunk_size)
    
    def _generate_row_chunks(self, worksheet: Worksheet, data_region: Dict[str, int], 
                           chunk_size: int) -> Iterator[Tuple[pd.DataFrame, int, Dict[str, Any]]]:
        """Generate row-based chunks"""
        min_row = data_region['min_row']
        max_row = data_region['max_row']
        min_col = data_region['min_col']
        max_col = data_region['max_col']
        
        chunk_index = 0
        
        for start_row in range(min_row, max_row + 1, chunk_size):
            end_row = min(start_row + chunk_size - 1, max_row)
            
            # Extract chunk data
            chunk_data = []
            for row in worksheet.iter_rows(
                min_row=start_row, max_row=end_row,
                min_col=min_col, max_col=max_col,
                values_only=True
            ):
                chunk_data.append(list(row))
            
            if chunk_data:
                # Create DataFrame
                num_cols = len(chunk_data[0]) if chunk_data else 0
                df = pd.DataFrame(
                    chunk_data, 
                    columns=[f'col_{i}' for i in range(num_cols)]
                )
                
                chunk_metadata = {
                    'start_row': start_row,
                    'end_row': end_row,
                    'start_col': min_col,
                    'end_col': max_col,
                    'chunk_type': 'row_based'
                }
                
                yield df, chunk_index, chunk_metadata
                chunk_index += 1
    
    def _generate_column_chunks(self, worksheet: Worksheet, data_region: Dict[str, int], 
                              chunk_size: int) -> Iterator[Tuple[pd.DataFrame, int, Dict[str, Any]]]:
        """Generate column-based chunks"""
        min_row = data_region['min_row']
        max_row = data_region['max_row']
        min_col = data_region['min_col']
        max_col = data_region['max_col']
        
        chunk_index = 0
        
        for start_col in range(min_col, max_col + 1, chunk_size):
            end_col = min(start_col + chunk_size - 1, max_col)
            
            # Extract chunk data
            chunk_data = []
            for row in worksheet.iter_rows(
                min_row=min_row, max_row=max_row,
                min_col=start_col, max_col=end_col,
                values_only=True
            ):
                chunk_data.append(list(row))
            
            if chunk_data:
                # Create DataFrame
                num_cols = len(chunk_data[0]) if chunk_data else 0
                df = pd.DataFrame(
                    chunk_data,
                    columns=[f'col_{i}' for i in range(num_cols)]
                )
                
                chunk_metadata = {
                    'start_row': min_row,
                    'end_row': max_row,
                    'start_col': start_col,
                    'end_col': end_col,
                    'chunk_type': 'column_based'
                }
                
                yield df, chunk_index, chunk_metadata
                chunk_index += 1
    
    def _adapt_chunk_size(self) -> None:
        """Adapt chunk size based on memory pressure"""
        if self.config.chunk_size_rows > 1000:
            self.config.chunk_size_rows = max(500, self.config.chunk_size_rows // 2)
            self.logger.info(f"Adapted chunk size to {self.config.chunk_size_rows} rows due to memory pressure")


class DataProfilingProcessor(ChunkProcessor):
    """Specialized chunk processor for data profiling operations"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.aggregated_stats = {
            'row_count': 0,
            'null_count': 0,
            'unique_values': set(),
            'data_types': {},
            'outliers': [],
            'patterns': []
        }
    
    def process_chunk(self, chunk_data: pd.DataFrame, chunk_index: int, 
                     chunk_metadata: Dict[str, Any]) -> Dict[str, Any]:
        """Process chunk for data profiling"""
        try:
            chunk_stats = {
                'chunk_index': chunk_index,
                'row_count': len(chunk_data),
                'column_count': len(chunk_data.columns),
                'null_percentages': {},
                'data_types': {},
                'outliers_detected': 0,
                'duplicate_rows': 0
            }
            
            # Calculate null percentages
            for col in chunk_data.columns:
                null_count = chunk_data[col].isna().sum()
                chunk_stats['null_percentages'][col] = null_count / len(chunk_data) if len(chunk_data) > 0 else 0
            
            # Infer data types
            for col in chunk_data.columns:
                series = chunk_data[col].dropna()
                if not series.empty:
                    chunk_stats['data_types'][col] = self._infer_data_type(series)
            
            # Detect outliers in numeric columns
            outliers_count = 0
            for col in chunk_data.columns:
                if pd.api.types.is_numeric_dtype(chunk_data[col]):
                    outliers_count += self._detect_outliers_in_series(chunk_data[col])
            chunk_stats['outliers_detected'] = outliers_count
            
            # Count duplicates
            chunk_stats['duplicate_rows'] = chunk_data.duplicated().sum()
            
            # Update aggregated stats
            self.aggregated_stats['row_count'] += chunk_stats['row_count']
            self.aggregated_stats['null_count'] += sum(chunk_data.isna().sum())
            
            return chunk_stats
            
        except Exception as e:
            self.logger.error(f"Error processing chunk {chunk_index}: {e}")
            return {'error': str(e), 'chunk_index': chunk_index}
    
    def _infer_data_type(self, series: pd.Series) -> str:
        """Infer data type for a series"""
        if pd.api.types.is_numeric_dtype(series):
            return 'numeric'
        elif pd.api.types.is_datetime64_any_dtype(series):
            return 'datetime'
        elif pd.api.types.is_bool_dtype(series):
            return 'boolean'
        else:
            return 'text'
    
    def _detect_outliers_in_series(self, series: pd.Series) -> int:
        """Detect outliers using IQR method"""
        try:
            numeric_series = pd.to_numeric(series, errors='coerce').dropna()
            if len(numeric_series) < 4:
                return 0
            
            Q1 = numeric_series.quantile(0.25)
            Q3 = numeric_series.quantile(0.75)
            IQR = Q3 - Q1
            
            lower_bound = Q1 - 1.5 * IQR
            upper_bound = Q3 + 1.5 * IQR
            
            outliers = numeric_series[(numeric_series < lower_bound) | (numeric_series > upper_bound)]
            return len(outliers)
            
        except Exception:
            return 0
    
    def get_aggregated_results(self) -> Dict[str, Any]:
        """Get final aggregated profiling results"""
        return self.aggregated_stats.copy()
