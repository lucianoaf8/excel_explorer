"""
Streaming Data Processor - Advanced streaming support for Excel analysis
Provides row-based, column-based, and cell-level streaming with progressive analysis.
"""

import time
import logging
from typing import Iterator, List, Dict, Any, Optional, Callable, Tuple, Union, Generator
from dataclasses import dataclass
from enum import Enum
import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .memory_manager import get_memory_manager, MemoryManager
from .error_handler import ExcelAnalysisError, ErrorSeverity, ErrorCategory


class StreamingStrategy(Enum):
    """Different streaming strategies for data processing"""
    ROW_STREAM = "row_stream"
    COLUMN_STREAM = "column_stream"
    CELL_STREAM = "cell_stream"
    PROGRESSIVE_ANALYSIS = "progressive_analysis"


@dataclass
class StreamConfig:
    """Configuration for streaming operations"""
    strategy: StreamingStrategy = StreamingStrategy.ROW_STREAM
    buffer_size: int = 1000
    max_memory_mb: float = 256.0
    enable_persistence: bool = False
    persistence_path: Optional[Path] = None
    progressive_depth: int = 3
    yield_frequency: int = 100


class StreamingDataProcessor:
    """Advanced streaming processor for Excel data with memory-efficient operations"""
    
    def __init__(self, config: Optional[StreamConfig] = None):
        self.config = config or StreamConfig()
        self.memory_manager = get_memory_manager()
        self.logger = logging.getLogger(__name__)
        self._processed_count = 0
        self._buffer = []
        self._persistence_enabled = self.config.enable_persistence
        
    def stream_worksheet_rows(self, worksheet: Worksheet, 
                             data_region: Optional[Dict[str, int]] = None) -> Iterator[Tuple[int, List[Any]]]:
        """Stream worksheet data row by row with memory management
        
        Args:
            worksheet: openpyxl worksheet object
            data_region: Optional data region bounds
            
        Yields:
            Tuple of (row_number, row_data)
        """
        if not data_region:
            data_region = self._detect_streaming_region(worksheet)
        
        min_row = data_region['min_row']
        max_row = data_region['max_row']
        min_col = data_region['min_col']
        max_col = data_region['max_col']
        
        self.logger.info(f"Streaming {max_row - min_row + 1} rows from worksheet")
        
        for row_idx in range(min_row, max_row + 1):
            try:
                # Check memory pressure
                if self.memory_manager.check_memory_pressure():
                    self._handle_memory_pressure()
                
                # Extract row data
                row_data = []
                for col_idx in range(min_col, max_col + 1):
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    row_data.append(cell_value)
                
                self._processed_count += 1
                
                # Yield row with index
                yield row_idx, row_data
                
                # Periodic memory check
                if self._processed_count % self.config.yield_frequency == 0:
                    self._periodic_maintenance()
                
            except Exception as e:
                self.logger.warning(f"Error streaming row {row_idx}: {e}")
                continue
    
    def stream_worksheet_columns(self, worksheet: Worksheet,
                                data_region: Optional[Dict[str, int]] = None) -> Iterator[Tuple[int, List[Any]]]:
        """Stream worksheet data column by column
        
        Args:
            worksheet: openpyxl worksheet object
            data_region: Optional data region bounds
            
        Yields:
            Tuple of (column_number, column_data)
        """
        if not data_region:
            data_region = self._detect_streaming_region(worksheet)
        
        min_row = data_region['min_row']
        max_row = data_region['max_row']
        min_col = data_region['min_col']
        max_col = data_region['max_col']
        
        self.logger.info(f"Streaming {max_col - min_col + 1} columns from worksheet")
        
        for col_idx in range(min_col, max_col + 1):
            try:
                # Check memory pressure
                if self.memory_manager.check_memory_pressure():
                    self._handle_memory_pressure()
                
                # Extract column data
                column_data = []
                for row_idx in range(min_row, max_row + 1):
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    column_data.append(cell_value)
                
                self._processed_count += 1
                
                # Yield column with index
                yield col_idx, column_data
                
                # Periodic memory check
                if self._processed_count % self.config.yield_frequency == 0:
                    self._periodic_maintenance()
                
            except Exception as e:
                self.logger.warning(f"Error streaming column {col_idx}: {e}")
                continue
    
    def stream_worksheet_cells(self, worksheet: Worksheet,
                              data_region: Optional[Dict[str, int]] = None) -> Iterator[Tuple[int, int, Any]]:
        """Stream worksheet data cell by cell for formula-heavy workbooks
        
        Args:
            worksheet: openpyxl worksheet object
            data_region: Optional data region bounds
            
        Yields:
            Tuple of (row, column, cell_value)
        """
        if not data_region:
            data_region = self._detect_streaming_region(worksheet)
        
        min_row = data_region['min_row']
        max_row = data_region['max_row']
        min_col = data_region['min_col']
        max_col = data_region['max_col']
        
        total_cells = (max_row - min_row + 1) * (max_col - min_col + 1)
        self.logger.info(f"Streaming {total_cells} cells from worksheet")
        
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                try:
                    # Check memory pressure more frequently for cell streaming
                    if self._processed_count % 50 == 0 and self.memory_manager.check_memory_pressure():
                        self._handle_memory_pressure()
                    
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    self._processed_count += 1
                    
                    # Yield cell coordinates and value
                    yield row_idx, col_idx, cell_value
                    
                    # More frequent maintenance for cell streaming
                    if self._processed_count % (self.config.yield_frequency // 2) == 0:
                        self._periodic_maintenance()
                
                except Exception as e:
                    self.logger.warning(f"Error streaming cell ({row_idx}, {col_idx}): {e}")
                    continue
    
    def progressive_analysis_stream(self, worksheet: Worksheet,
                                   analyzer_func: Callable[[pd.DataFrame], Dict[str, Any]],
                                   data_region: Optional[Dict[str, int]] = None) -> Iterator[Dict[str, Any]]:
        """Stream data with progressive analysis at configurable depth levels
        
        Args:
            worksheet: openpyxl worksheet object
            analyzer_func: Function to analyze data chunks
            data_region: Optional data region bounds
            
        Yields:
            Analysis results for each progressive level
        """
        if not data_region:
            data_region = self._detect_streaming_region(worksheet)
        
        total_rows = data_region['max_row'] - data_region['min_row'] + 1
        depth_levels = self._calculate_progressive_levels(total_rows)
        
        self.logger.info(f"Progressive analysis with {len(depth_levels)} depth levels")
        
        for level, sample_size in enumerate(depth_levels):
            try:
                # Extract sample data for this level
                sample_data = self._extract_progressive_sample(
                    worksheet, data_region, sample_size
                )
                
                if sample_data.empty:
                    continue
                
                # Run analysis on this sample
                analysis_result = analyzer_func(sample_data)
                
                # Add metadata about the progressive level
                analysis_result.update({
                    'progressive_level': level + 1,
                    'sample_size': len(sample_data),
                    'total_rows': total_rows,
                    'coverage_ratio': len(sample_data) / total_rows,
                    'analysis_depth': self.config.progressive_depth
                })
                
                yield analysis_result
                
                # Memory management between levels
                del sample_data
                self._periodic_maintenance()
                
            except Exception as e:
                self.logger.warning(f"Error in progressive analysis level {level}: {e}")
                continue
    
    def stream_with_intermediate_persistence(self, worksheet: Worksheet,
                                           processor_func: Callable[[List[Any]], Dict[str, Any]],
                                           data_region: Optional[Dict[str, int]] = None) -> Iterator[Dict[str, Any]]:
        """Stream with intermediate result persistence for very large files
        
        Args:
            worksheet: openpyxl worksheet object
            processor_func: Function to process data batches
            data_region: Optional data region bounds
            
        Yields:
            Processing results with persistence metadata
        """
        if not self._persistence_enabled:
            raise ValueError("Persistence not enabled in StreamConfig")
        
        persistence_path = self.config.persistence_path or Path("temp_streaming")
        persistence_path.mkdir(exist_ok=True)
        
        batch_count = 0
        
        try:
            for row_idx, row_data in self.stream_worksheet_rows(worksheet, data_region):
                self._buffer.append(row_data)
                
                # Process buffer when full
                if len(self._buffer) >= self.config.buffer_size:
                    try:
                        # Process current buffer
                        result = processor_func(self._buffer.copy())
                        
                        # Add persistence metadata
                        result.update({
                            'batch_number': batch_count,
                            'buffer_size': len(self._buffer),
                            'persistence_enabled': True
                        })
                        
                        # Persist intermediate result
                        self._persist_intermediate_result(result, batch_count, persistence_path)
                        
                        yield result
                        
                        # Clear buffer
                        self._buffer.clear()
                        batch_count += 1
                        
                    except Exception as e:
                        self.logger.error(f"Error processing batch {batch_count}: {e}")
                        self._buffer.clear()  # Clear corrupted buffer
                        continue
            
            # Process remaining buffer
            if self._buffer:
                try:
                    result = processor_func(self._buffer.copy())
                    result.update({
                        'batch_number': batch_count,
                        'buffer_size': len(self._buffer),
                        'is_final_batch': True
                    })
                    
                    self._persist_intermediate_result(result, batch_count, persistence_path)
                    yield result
                    
                except Exception as e:
                    self.logger.error(f"Error processing final batch: {e}")
        
        finally:
            self._buffer.clear()
    
    def _detect_streaming_region(self, worksheet: Worksheet) -> Dict[str, int]:
        """Detect data region for streaming operations"""
        # Simple boundary detection optimized for streaming
        min_row, max_row = 1, worksheet.max_row or 1
        min_col, max_col = 1, worksheet.max_column or 1
        
        # Quick validation with sampling
        sample_rows = min(10, max_row)
        sample_cols = min(10, max_col)
        
        # Find actual data boundaries by sampling
        data_found = False
        for r in range(1, sample_rows + 1):
            for c in range(1, sample_cols + 1):
                try:
                    if worksheet.cell(row=r, column=c).value is not None:
                        data_found = True
                        break
                except:
                    continue
            if data_found:
                break
        
        if not data_found:
            # Fallback to minimal region
            max_row = min(100, max_row)
            max_col = min(26, max_col)
        
        return {
            'min_row': min_row,
            'max_row': max_row,
            'min_col': min_col,
            'max_col': max_col
        }
    
    def _calculate_progressive_levels(self, total_rows: int) -> List[int]:
        """Calculate sample sizes for progressive analysis levels"""
        base_samples = [100, 500, 1000, 5000]
        levels = []
        
        for sample_size in base_samples:
            if sample_size <= total_rows:
                levels.append(sample_size)
            else:
                break
        
        # Always include full dataset as final level if reasonable
        if total_rows <= 50000:  # Reasonable limit for full analysis
            levels.append(total_rows)
        elif levels and levels[-1] < total_rows * 0.5:
            # Add a larger sample if current max is less than 50% of data
            levels.append(min(25000, total_rows))
        
        return levels[:self.config.progressive_depth] if levels else [min(1000, total_rows)]
    
    def _extract_progressive_sample(self, worksheet: Worksheet, 
                                   data_region: Dict[str, int], 
                                   sample_size: int) -> pd.DataFrame:
        """Extract a representative sample for progressive analysis"""
        total_rows = data_region['max_row'] - data_region['min_row'] + 1
        
        if sample_size >= total_rows:
            # Extract all data
            sample_rows = range(data_region['min_row'], data_region['max_row'] + 1)
        else:
            # Strategic sampling: include first rows, last rows, and distributed middle
            first_rows = min(sample_size // 4, 50)
            last_rows = min(sample_size // 4, 50)
            middle_rows = sample_size - first_rows - last_rows
            
            sample_rows = []
            
            # First rows
            sample_rows.extend(range(
                data_region['min_row'], 
                data_region['min_row'] + first_rows
            ))
            
            # Distributed middle rows
            if middle_rows > 0:
                middle_start = data_region['min_row'] + first_rows
                middle_end = data_region['max_row'] - last_rows
                if middle_end > middle_start:
                    step = max(1, (middle_end - middle_start) // middle_rows)
                    sample_rows.extend(range(middle_start, middle_end, step))
            
            # Last rows
            sample_rows.extend(range(
                max(data_region['max_row'] - last_rows, data_region['min_row']),
                data_region['max_row'] + 1
            ))
            
            # Remove duplicates and sort
            sample_rows = sorted(set(sample_rows))
        
        # Extract data for sampled rows
        data = []
        for row_idx in sample_rows:
            row_data = []
            for col_idx in range(data_region['min_col'], data_region['max_col'] + 1):
                try:
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    row_data.append(cell_value)
                except:
                    row_data.append(None)
            data.append(row_data)
        
        # Convert to DataFrame
        if data:
            num_cols = len(data[0]) if data else 0
            return pd.DataFrame(data, columns=[f'col_{i}' for i in range(num_cols)])
        else:
            return pd.DataFrame()
    
    def _handle_memory_pressure(self) -> None:
        """Handle memory pressure during streaming"""
        self.logger.warning("Memory pressure detected during streaming - initiating cleanup")
        
        # Clear buffer if using buffered operations
        if self._buffer:
            self.logger.info(f"Clearing buffer with {len(self._buffer)} items")
            self._buffer.clear()
        
        # Force garbage collection
        import gc
        gc.collect()
        
        # Reduce buffer size for future operations
        self.config.buffer_size = max(100, self.config.buffer_size // 2)
        
        self.logger.info(f"Reduced buffer size to {self.config.buffer_size}")
    
    def _periodic_maintenance(self) -> None:
        """Perform periodic maintenance during streaming"""
        # Light garbage collection
        if self._processed_count % (self.config.yield_frequency * 5) == 0:
            import gc
            gc.collect()
        
        # Log progress
        if self._processed_count % (self.config.yield_frequency * 10) == 0:
            memory_info = self.memory_manager.get_current_usage()
            self.logger.info(
                f"Streaming progress: {self._processed_count} items processed, "
                f"Memory usage: {memory_info['usage_ratio']:.1%}"
            )
    
    def _persist_intermediate_result(self, result: Dict[str, Any], 
                                   batch_number: int, 
                                   persistence_path: Path) -> None:
        """Persist intermediate result to disk"""
        try:
            import json
            result_file = persistence_path / f"batch_{batch_number:06d}.json"
            
            # Convert numpy types to native Python types for JSON serialization
            serializable_result = self._make_json_serializable(result)
            
            with open(result_file, 'w') as f:
                json.dump(serializable_result, f, indent=2)
            
            self.logger.debug(f"Persisted batch {batch_number} to {result_file}")
            
        except Exception as e:
            self.logger.warning(f"Failed to persist batch {batch_number}: {e}")
    
    def _make_json_serializable(self, obj: Any) -> Any:
        """Convert numpy and pandas types to JSON-serializable types"""
        if isinstance(obj, dict):
            return {k: self._make_json_serializable(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [self._make_json_serializable(item) for item in obj]
        elif isinstance(obj, (np.integer, np.floating)):
            return obj.item()
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        elif pd.isna(obj):
            return None
        else:
            return obj
    
    def get_streaming_stats(self) -> Dict[str, Any]:
        """Get current streaming statistics"""
        return {
            'processed_count': self._processed_count,
            'buffer_size': len(self._buffer),
            'config': {
                'strategy': self.config.strategy.value,
                'buffer_size': self.config.buffer_size,
                'max_memory_mb': self.config.max_memory_mb,
                'progressive_depth': self.config.progressive_depth
            },
            'memory_usage': self.memory_manager.get_current_usage()
        }
    
    def reset_stats(self) -> None:
        """Reset streaming statistics"""
        self._processed_count = 0
        self._buffer.clear()


class FormulaStreamProcessor:
    """Specialized streaming processor for formula-heavy workbooks"""
    
    def __init__(self, config: Optional[StreamConfig] = None):
        self.config = config or StreamConfig()
        self.logger = logging.getLogger(__name__)
        self.formula_cache = {}
        
    def stream_formulas(self, worksheet: Worksheet,
                       data_region: Optional[Dict[str, int]] = None) -> Iterator[Tuple[int, int, str, Any]]:
        """Stream formulas with caching and dependency tracking
        
        Args:
            worksheet: openpyxl worksheet object
            data_region: Optional data region bounds
            
        Yields:
            Tuple of (row, column, formula_text, calculated_value)
        """
        if not data_region:
            data_region = self._detect_formula_region(worksheet)
        
        formula_count = 0
        
        for row_idx in range(data_region['min_row'], data_region['max_row'] + 1):
            for col_idx in range(data_region['min_col'], data_region['max_col'] + 1):
                try:
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    
                    # Check if cell contains a formula
                    if cell.data_type == 'f' or (cell.value and str(cell.value).startswith('=')):
                        formula_text = str(cell.value)
                        calculated_value = self._get_calculated_value(cell)
                        
                        formula_count += 1
                        yield row_idx, col_idx, formula_text, calculated_value
                        
                        # Periodic maintenance for large formula sets
                        if formula_count % 1000 == 0:
                            self._maintain_formula_cache()
                
                except Exception as e:
                    self.logger.warning(f"Error processing formula at ({row_idx}, {col_idx}): {e}")
                    continue
    
    def _detect_formula_region(self, worksheet: Worksheet) -> Dict[str, int]:
        """Detect regions likely to contain formulas"""
        # Use full worksheet range for formula detection
        return {
            'min_row': 1,
            'max_row': worksheet.max_row or 1,
            'min_col': 1,
            'max_col': worksheet.max_column or 1
        }
    
    def _get_calculated_value(self, cell) -> Any:
        """Get calculated value from formula cell with caching"""
        try:
            # Try to get calculated value
            if hasattr(cell, 'value') and cell.value is not None:
                return cell.value
            else:
                return None
        except Exception:
            return None
    
    def _maintain_formula_cache(self) -> None:
        """Maintain formula cache to prevent memory bloat"""
        if len(self.formula_cache) > 10000:
            # Keep only the most recently accessed formulas
            self.formula_cache = dict(list(self.formula_cache.items())[-5000:])


def create_streaming_processor(strategy: StreamingStrategy = StreamingStrategy.ROW_STREAM,
                             **kwargs) -> StreamingDataProcessor:
    """Factory function to create streaming processor with specified strategy"""
    config = StreamConfig(strategy=strategy, **kwargs)
    return StreamingDataProcessor(config)
