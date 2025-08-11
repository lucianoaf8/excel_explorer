#!/usr/bin/env python3
"""
Excel Data Anonymizer
Replaces sensitive data in Excel files with fake values while maintaining mappings
"""

import json
import re
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple, Any
from datetime import datetime
import hashlib

try:
    from faker import Faker
except ImportError:
    print("Warning: faker library not installed. Install with: pip install faker")
    Faker = None

import openpyxl
from openpyxl.utils import get_column_letter


class ExcelAnonymizer:
    """Anonymizes sensitive data in Excel files with reversible mappings"""
    
    # Patterns for detecting sensitive columns
    NAME_PATTERNS = [
        r'.*name.*', r'.*contractor.*', r'.*client.*', r'.*customer.*',
        r'.*person.*', r'.*employee.*', r'.*staff.*', r'.*worker.*',
        r'.*owner.*', r'.*manager.*', r'.*contact.*'
    ]
    
    COMPANY_PATTERNS = [
        r'.*company.*', r'.*organization.*', r'.*org\b.*', r'.*vendor.*',
        r'.*supplier.*', r'.*partner.*', r'.*firm.*', r'.*business.*',
        r'.*enterprise.*', r'.*corporation.*', r'.*corp\b.*'
    ]
    
    EMAIL_PATTERNS = [
        r'.*email.*', r'.*e-mail.*', r'.*mail.*'
    ]
    
    PHONE_PATTERNS = [
        r'.*phone.*', r'.*mobile.*', r'.*cell.*', r'.*tel.*', r'.*fax.*'
    ]
    
    ADDRESS_PATTERNS = [
        r'.*address.*', r'.*street.*', r'.*city.*', r'.*state.*',
        r'.*zip.*', r'.*postal.*', r'.*location.*'
    ]
    
    def __init__(self, file_path: str, locale: str = 'en_US'):
        """
        Initialize the anonymizer
        
        Args:
            file_path: Path to Excel file to anonymize
            locale: Locale for Faker (default: en_US)
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        if Faker is None:
            raise ImportError("faker library required. Install with: pip install faker")
        
        self.faker = Faker(locale)
        Faker.seed(42)  # For reproducible fake data during testing
        
        # Mapping dictionaries: {column_name: {original_value: fake_value}}
        self.mappings: Dict[str, Dict[str, str]] = {}
        
        # Cache for consistent fake values
        self.fake_cache: Dict[str, str] = {}
        
        # Load the workbook
        self.workbook = openpyxl.load_workbook(self.file_path)
        
    def find_sensitive_columns(self, sheet_name: Optional[str] = None) -> Dict[str, List[Tuple[str, str]]]:
        """
        Find columns that likely contain sensitive data
        
        Args:
            sheet_name: Specific sheet to analyze (None for all sheets)
            
        Returns:
            Dictionary mapping sheet names to list of (column_index, column_type) tuples
        """
        sensitive_columns = {}
        
        sheets = [sheet_name] if sheet_name else self.workbook.sheetnames
        
        for sheet in sheets:
            if sheet not in self.workbook.sheetnames:
                print(f"Warning: Sheet '{sheet}' not found")
                continue
                
            ws = self.workbook[sheet]
            sheet_columns = []
            
            # Check first row for headers
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers.append((col, str(header).lower()))
            
            # Match patterns
            for col_idx, header in headers:
                col_letter = get_column_letter(col_idx)
                
                # Check for name patterns
                if any(re.match(pattern, header, re.IGNORECASE) for pattern in self.NAME_PATTERNS):
                    sheet_columns.append((col_letter, 'name'))
                    
                # Check for company patterns
                elif any(re.match(pattern, header, re.IGNORECASE) for pattern in self.COMPANY_PATTERNS):
                    sheet_columns.append((col_letter, 'company'))
                    
                # Check for email patterns
                elif any(re.match(pattern, header, re.IGNORECASE) for pattern in self.EMAIL_PATTERNS):
                    sheet_columns.append((col_letter, 'email'))
                    
                # Check for phone patterns
                elif any(re.match(pattern, header, re.IGNORECASE) for pattern in self.PHONE_PATTERNS):
                    sheet_columns.append((col_letter, 'phone'))
                    
                # Check for address patterns
                elif any(re.match(pattern, header, re.IGNORECASE) for pattern in self.ADDRESS_PATTERNS):
                    sheet_columns.append((col_letter, 'address'))
            
            if sheet_columns:
                sensitive_columns[sheet] = sheet_columns
                
        return sensitive_columns
    
    def _get_fake_value(self, original: str, data_type: str, column_key: str) -> str:
        """
        Get a fake value for the original, ensuring consistency
        
        Args:
            original: Original value to replace
            data_type: Type of data (name, company, email, etc.)
            column_key: Unique key for the column (sheet:column)
            
        Returns:
            Fake value
        """
        # Create unique cache key
        cache_key = f"{column_key}:{original}"
        
        # Check cache first
        if cache_key in self.fake_cache:
            return self.fake_cache[cache_key]
        
        # Generate new fake value based on type
        if data_type == 'name':
            fake = self.faker.name()
        elif data_type == 'company':
            fake = self.faker.company()
        elif data_type == 'email':
            # Generate email based on fake name if we have one
            if original in self.mappings.get(column_key, {}):
                name = self.mappings[column_key][original]
                fake = f"{name.lower().replace(' ', '.')}@{self.faker.domain_name()}"
            else:
                fake = self.faker.email()
        elif data_type == 'phone':
            fake = self.faker.phone_number()
        elif data_type == 'address':
            fake = self.faker.address().replace('\n', ', ')
        else:
            # Default: use hash-based replacement
            hash_val = hashlib.md5(original.encode()).hexdigest()[:8]
            fake = f"ANON_{hash_val}"
        
        # Cache the value
        self.fake_cache[cache_key] = fake
        
        # Store in mappings
        if column_key not in self.mappings:
            self.mappings[column_key] = {}
        self.mappings[column_key][original] = fake
        
        return fake
    
    def anonymize_columns(self, 
                         columns_to_anonymize: Optional[Dict[str, List[Tuple[str, str]]]] = None,
                         auto_detect: bool = True) -> Dict[str, int]:
        """
        Anonymize specified columns or auto-detect sensitive columns
        
        Args:
            columns_to_anonymize: Dict mapping sheet names to (column, type) tuples
            auto_detect: Whether to auto-detect sensitive columns if none specified
            
        Returns:
            Statistics about anonymization (counts per column)
        """
        stats = {}
        
        # Determine which columns to anonymize
        if columns_to_anonymize is None and auto_detect:
            columns_to_anonymize = self.find_sensitive_columns()
            print(f"Auto-detected {sum(len(cols) for cols in columns_to_anonymize.values())} sensitive columns")
        elif columns_to_anonymize is None:
            print("No columns specified and auto-detect disabled")
            return stats
        
        # Process each sheet
        for sheet_name, columns in columns_to_anonymize.items():
            ws = self.workbook[sheet_name]
            
            for col_letter, data_type in columns:
                column_key = f"{sheet_name}:{col_letter}"
                anonymized_count = 0
                
                # Get column index
                col_idx = openpyxl.utils.column_index_from_string(col_letter)
                
                # Get header name for reporting
                header = ws.cell(row=1, column=col_idx).value
                print(f"Anonymizing column '{header}' ({col_letter}) in sheet '{sheet_name}' as {data_type}")
                
                # Process all rows (skip header)
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    original = cell.value
                    
                    # Skip empty cells
                    if original is None or str(original).strip() == '':
                        continue
                    
                    # Convert to string for processing
                    original_str = str(original)
                    
                    # Get fake value
                    fake_value = self._get_fake_value(original_str, data_type, column_key)
                    
                    # Update cell
                    cell.value = fake_value
                    anonymized_count += 1
                
                stats[column_key] = anonymized_count
                print(f"  → Anonymized {anonymized_count} values")
        
        return stats
    
    def save_anonymized_file(self, output_path: Optional[str] = None) -> str:
        """
        Save the anonymized Excel file
        
        Args:
            output_path: Output file path (default: adds _anonymized to original name)
            
        Returns:
            Path to saved file
        """
        if output_path is None:
            # Generate default output path
            stem = self.file_path.stem
            suffix = self.file_path.suffix
            output_path = self.file_path.parent / f"{stem}_anonymized{suffix}"
        else:
            output_path = Path(output_path)
        
        # Save the workbook
        self.workbook.save(output_path)
        print(f"Anonymized file saved: {output_path}")
        
        return str(output_path)
    
    def save_mappings(self, mapping_path: Optional[str] = None, 
                     format: str = 'json') -> str:
        """
        Save the mapping dictionary for reversal
        
        Args:
            mapping_path: Output path for mappings (default: based on input file)
            format: Output format ('json' or 'excel')
            
        Returns:
            Path to saved mapping file
        """
        if mapping_path is None:
            stem = self.file_path.stem
            ext = '.json' if format == 'json' else '.xlsx'
            mapping_path = self.file_path.parent / f"{stem}_mappings{ext}"
        else:
            mapping_path = Path(mapping_path)
        
        if format == 'json':
            # Save as JSON
            output_data = {
                'metadata': {
                    'created': datetime.now().isoformat(),
                    'source_file': str(self.file_path),
                    'total_mappings': sum(len(m) for m in self.mappings.values()),
                    'columns': list(self.mappings.keys())
                },
                'mappings': self.mappings
            }
            
            with open(mapping_path, 'w', encoding='utf-8') as f:
                json.dump(output_data, f, indent=2, ensure_ascii=False)
            
            print(f"Mappings saved to JSON: {mapping_path}")
            
        elif format == 'excel':
            # Save as Excel with sheet per column
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create summary sheet
            summary = wb.create_sheet('Summary')
            summary.append(['Column', 'Total Mappings'])
            
            for column_key, mappings in self.mappings.items():
                # Add to summary
                summary.append([column_key, len(mappings)])
                
                # Create sheet for this column
                sheet_name = column_key.replace(':', '_')[:31]  # Excel sheet name limit
                ws = wb.create_sheet(sheet_name)
                
                # Add headers
                ws.append(['Original', 'Anonymized', 'Type'])
                
                # Add mappings
                for original, fake in mappings.items():
                    # Detect type from column key
                    data_type = 'unknown'
                    for col_type in ['name', 'company', 'email', 'phone', 'address']:
                        if col_type in column_key.lower():
                            data_type = col_type
                            break
                    
                    ws.append([original, fake, data_type])
            
            wb.save(mapping_path)
            print(f"Mappings saved to Excel: {mapping_path}")
        
        return str(mapping_path)
    
    def reverse_anonymization(self, anonymized_file: str, mapping_file: str,
                            output_path: Optional[str] = None) -> str:
        """
        Reverse the anonymization using a mapping file
        
        Args:
            anonymized_file: Path to anonymized Excel file
            mapping_file: Path to mapping file (JSON or Excel)
            output_path: Output path for restored file
            
        Returns:
            Path to restored file
        """
        # Load mappings
        mapping_path = Path(mapping_file)
        
        if mapping_path.suffix == '.json':
            with open(mapping_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                mappings = data['mappings']
        else:
            # Load from Excel
            wb = openpyxl.load_workbook(mapping_path)
            mappings = {}
            
            for sheet_name in wb.sheetnames:
                if sheet_name == 'Summary':
                    continue
                
                ws = wb[sheet_name]
                column_key = sheet_name.replace('_', ':', 1)
                mappings[column_key] = {}
                
                # Skip header row
                for row in range(2, ws.max_row + 1):
                    original = ws.cell(row=row, column=1).value
                    fake = ws.cell(row=row, column=2).value
                    if original and fake:
                        # Reverse mapping: fake -> original
                        mappings[column_key][str(fake)] = str(original)
        
        # Load anonymized file
        wb_anon = openpyxl.load_workbook(anonymized_file)
        
        # Process each mapping
        restored_count = 0
        for column_key, column_mappings in mappings.items():
            sheet_name, col_letter = column_key.split(':')
            
            if sheet_name not in wb_anon.sheetnames:
                continue
            
            ws = wb_anon[sheet_name]
            col_idx = openpyxl.utils.column_index_from_string(col_letter)
            
            # Reverse the mapping (fake -> original)
            reverse_map = {v: k for k, v in column_mappings.items()}
            
            # Process all rows (skip header)
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                current_value = cell.value
                
                if current_value and str(current_value) in reverse_map:
                    cell.value = reverse_map[str(current_value)]
                    restored_count += 1
        
        # Save restored file
        if output_path is None:
            anon_path = Path(anonymized_file)
            output_path = anon_path.parent / f"{anon_path.stem}_restored{anon_path.suffix}"
        else:
            output_path = Path(output_path)
        
        wb_anon.save(output_path)
        print(f"Restored {restored_count} values")
        print(f"Restored file saved: {output_path}")
        
        return str(output_path)


def anonymize_file(file_path: str, 
                  output_path: Optional[str] = None,
                  mapping_path: Optional[str] = None,
                  columns: Optional[Dict[str, List[Tuple[str, str]]]] = None,
                  auto_detect: bool = True,
                  mapping_format: str = 'json') -> Tuple[str, str]:
    """
    Convenience function to anonymize a file
    
    Args:
        file_path: Path to Excel file
        output_path: Output path for anonymized file
        mapping_path: Output path for mapping file
        columns: Specific columns to anonymize
        auto_detect: Whether to auto-detect sensitive columns
        mapping_format: Format for mapping file ('json' or 'excel')
        
    Returns:
        Tuple of (anonymized_file_path, mapping_file_path)
    """
    anonymizer = ExcelAnonymizer(file_path)
    
    # Anonymize columns
    stats = anonymizer.anonymize_columns(columns, auto_detect)
    
    if not stats:
        print("No columns were anonymized")
        return None, None
    
    # Save files
    anon_path = anonymizer.save_anonymized_file(output_path)
    map_path = anonymizer.save_mappings(mapping_path, mapping_format)
    
    print(f"\nAnonymization complete!")
    print(f"Total values anonymized: {sum(stats.values())}")
    
    return anon_path, map_path