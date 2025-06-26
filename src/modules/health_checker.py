"""
File integrity validation module
Enhanced to work with ModuleResult framework and AnalysisContext.
"""

from typing import List, Optional
from pathlib import Path
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
import time

from ..core.base_analyzer import BaseAnalyzer
from ..core.analysis_context import AnalysisContext
from ..core.module_result import HealthCheckData, ValidationResult, ConfidenceLevel
from ..utils.file_handler import file_exists, is_excel_file, get_file_size_mb, contains_macros
from ..utils.error_handler import ExcelAnalysisError, ErrorSeverity, ErrorCategory


class HealthChecker(BaseAnalyzer):
    """Enhanced health checker with comprehensive file validation"""
    
    def __init__(self, name: str = "health_checker", dependencies: Optional[List[str]] = None):
        super().__init__(name, dependencies or [])
        
    def _perform_analysis(self, context: AnalysisContext) -> HealthCheckData:
        """Perform comprehensive file health validation
        
        Args:
            context: AnalysisContext with file access
            
        Returns:
            HealthCheckData with validation results
        """
        file_path = context.file_path
        
        # Basic file checks
        if not file_exists(file_path):
            raise ExcelAnalysisError(
                f"File not found: {file_path}",
                severity=ErrorSeverity.CRITICAL,
                category=ErrorCategory.FILE_ACCESS,
                module_name=self.name,
                file_path=str(file_path)
            )
        
        if not is_excel_file(file_path):
            raise ExcelAnalysisError(
                "Unsupported file type - expected Excel format",
                severity=ErrorSeverity.CRITICAL,
                category=ErrorCategory.FILE_ACCESS,
                module_name=self.name,
                file_path=str(file_path)
            )
        
        # File size validation
        max_mb = self.config.get("max_file_size_mb", 500)
        size_mb = get_file_size_mb(file_path)
        
        # Security and integrity checks
        security_issues = []
        repair_suggestions = []
        
        # Size check
        if size_mb > max_mb:
            security_issues.append(f"File size {size_mb:.1f}MB exceeds limit {max_mb}MB")
            repair_suggestions.append("Consider splitting large file or increasing memory limits")
        
        # Corruption detection
        corruption_detected = self._check_corruption(file_path)
        if corruption_detected:
            security_issues.append("File corruption detected")
            repair_suggestions.append("Try opening file in Excel and saving as new copy")
        
        # Macro detection
        macro_enabled = contains_macros(file_path)
        if macro_enabled:
            security_issues.append("File contains VBA macros")
            repair_suggestions.append("Review macros for security before enabling")
        
        # Password protection check
        password_protected = self._check_password_protection(file_path)
        if password_protected:
            security_issues.append("File is password protected")
            repair_suggestions.append("Provide password or request unprotected version")
        
        # Excel version detection
        excel_version = self._detect_excel_version(file_path)
        
        return HealthCheckData(
            file_accessible=not corruption_detected and not password_protected,
            file_size_mb=size_mb,
            corruption_detected=corruption_detected,
            security_issues=security_issues,
            excel_version=excel_version,
            macro_enabled=macro_enabled,
            password_protected=password_protected,
            repair_suggestions=repair_suggestions
        )
    
    def _validate_result(self, data: HealthCheckData, context: AnalysisContext) -> ValidationResult:
        """Validate health check results
        
        Args:
            data: HealthCheckData to validate
            context: AnalysisContext for additional validation
            
        Returns:
            ValidationResult with quality metrics
        """
        # Completeness: Did we check all required aspects?
        completeness = 1.0  # Health check is comprehensive by design
        
        # Accuracy: Based on file accessibility and issue detection
        if data.file_accessible and not data.corruption_detected:
            accuracy = 0.95
        elif data.file_accessible:
            accuracy = 0.8  # Accessible but has issues
        else:
            accuracy = 0.3  # Major accessibility problems
        
        # Consistency: Internal logic consistency
        consistency = 0.9
        if data.password_protected and data.file_accessible:
            consistency -= 0.2  # Logical inconsistency
        if data.corruption_detected and data.file_accessible:
            consistency -= 0.3  # Major inconsistency
        
        # Confidence based on validation results
        if accuracy > 0.9 and consistency > 0.8:
            confidence = ConfidenceLevel.HIGH
        elif accuracy > 0.7:
            confidence = ConfidenceLevel.MEDIUM
        else:
            confidence = ConfidenceLevel.LOW
        
        validation_notes = []
        if data.security_issues:
            validation_notes.append(f"Found {len(data.security_issues)} security issues")
        if data.repair_suggestions:
            validation_notes.append(f"Generated {len(data.repair_suggestions)} repair suggestions")
        
        return ValidationResult(
            completeness_score=completeness,
            accuracy_score=max(0.0, accuracy),
            consistency_score=max(0.0, consistency),
            confidence_level=confidence,
            validation_notes=validation_notes
        )
    
    def _check_corruption(self, file_path: Path) -> bool:
        """Check for file corruption using multiple methods
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            bool: True if corruption detected
        """
        try:
            # Primary check: Try to load workbook
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            # Secondary check: Try to access basic properties
            _ = wb.sheetnames
            _ = len(wb.worksheets)
            
            wb.close()
            return False
            
        except InvalidFileException:
            self.logger.warning(f"Invalid file format detected: {file_path}")
            return True
        except PermissionError:
            self.logger.warning(f"Permission denied accessing file: {file_path}")
            return False  # Not corruption, just access issue
        except Exception as e:
            self.logger.warning(f"Potential corruption detected: {e}")
            return True
    
    def _check_password_protection(self, file_path: Path) -> bool:
        """Check if file is password protected
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            bool: True if password protected
        """
        try:
            # openpyxl raises InvalidFileException for password-protected files
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            wb.close()
            return False
        except InvalidFileException as e:
            # Check if error message indicates password protection
            error_msg = str(e).lower()
            if any(keyword in error_msg for keyword in ['password', 'encrypted', 'protected']):
                return True
            return False  # Other type of invalid file
        except Exception:
            return False  # Assume not password protected if other error
    
    def _detect_excel_version(self, file_path: Path) -> Optional[str]:
        """Detect Excel file version/format
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Optional[str]: Excel version or format info
        """
        try:
            suffix = file_path.suffix.lower()
            
            if suffix == '.xlsx':
                return "Excel 2007+ (XLSX)"
            elif suffix == '.xlsm':
                return "Excel 2007+ with Macros (XLSM)"
            elif suffix == '.xls':
                return "Excel 97-2003 (XLS)"
            elif suffix == '.xlsb':
                return "Excel Binary Workbook (XLSB)"
            elif suffix == '.csv':
                return "Comma Separated Values (CSV)"
            else:
                return f"Unknown format ({suffix})"
                
        except Exception:
            return None
    
    def estimate_complexity(self, context: AnalysisContext) -> float:
        """Health check is always low complexity
        
        Args:
            context: AnalysisContext (unused for health check)
            
        Returns:
            float: Complexity factor (always 0.1 for health check)
        """
        return 0.1  # Health check is very fast
    
    def _count_processed_items(self, data: HealthCheckData) -> int:
        """Count validation checks performed
        
        Args:
            data: HealthCheckData result
            
        Returns:
            int: Number of validation checks
        """
        checks = [
            'file_accessible',
            'file_size_mb',
            'corruption_detected', 
            'macro_enabled',
            'password_protected'
        ]
        return len(checks)


# Legacy compatibility for existing code
def create_health_checker(config: dict = None) -> HealthChecker:
    """Factory function for backward compatibility"""
    checker = HealthChecker()
    if config:
        checker.configure(config)
    return checker
