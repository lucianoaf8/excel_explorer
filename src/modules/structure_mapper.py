"""
Workbook architecture analysis module
Enhanced to work with ModuleResult framework and AnalysisContext.
"""

from typing import List, Optional, Dict, Any
from pathlib import Path
import openpyxl
from openpyxl.worksheet.table import Table

from ..core.base_analyzer import BaseAnalyzer
from ..core.analysis_context import AnalysisContext
from ..core.module_result import StructureData, ValidationResult, ConfidenceLevel
from ..utils.error_handler import ExcelAnalysisError, ErrorSeverity, ErrorCategory


class StructureMapper(BaseAnalyzer):
    """Enhanced structure mapper for comprehensive workbook architecture analysis"""
    
    def __init__(self, name: str = "structure_mapper", dependencies: Optional[List[str]] = None):
        super().__init__(name, dependencies or ["health_checker"])
    
    def _perform_analysis(self, context: AnalysisContext) -> StructureData:
        """Perform comprehensive workbook structure analysis
        
        Args:
            context: AnalysisContext with workbook access
            
        Returns:
            StructureData with complete architecture mapping
        """
        try:
            with context.get_workbook_access().get_workbook() as wb:
                # Basic worksheet analysis
                worksheet_names = wb.sheetnames
                worksheet_count = len(worksheet_names)
                
                # Hidden sheets detection
                hidden_sheets = []
                visible_sheets = []
                
                include_hidden = self.config.get("include_hidden_sheets", True)
                
                for ws in wb.worksheets:
                    if ws.sheet_state == "visible":
                        visible_sheets.append(ws.title)
                    else:
                        hidden_sheets.append(ws.title)
                
                # Named ranges analysis
                named_ranges = self._analyze_named_ranges(wb)
                
                # Sheet relationships (references between sheets)
                sheet_relationships = self._analyze_sheet_relationships(wb)
                
                # Chart and pivot table counting
                chart_count = self._count_charts(wb)
                pivot_table_count = self._count_pivot_tables(wb)
                
                # Cell statistics
                total_cells, non_empty_cells = self._calculate_cell_statistics(wb, include_hidden)
                
                return StructureData(
                    worksheet_count=worksheet_count,
                    worksheet_names=worksheet_names,
                    named_ranges=named_ranges,
                    sheet_relationships=sheet_relationships,
                    hidden_sheets=hidden_sheets,
                    visible_sheets=visible_sheets,
                    chart_count=chart_count,
                    pivot_table_count=pivot_table_count,
                    total_cells=total_cells,
                    total_cells_with_data=non_empty_cells
                )
                
        except Exception as e:
            raise ExcelAnalysisError(
                f"Structure mapping failed: {e}",
                severity=ErrorSeverity.HIGH,
                category=ErrorCategory.DATA_CORRUPTION,
                module_name=self.name,
                file_path=str(context.file_path)
            )
    
    def _analyze_named_ranges(self, wb) -> Dict[str, Any]:
        """Analyze named ranges in workbook"""
        try:
            return {"count": len(wb.defined_names), "names": list(wb.defined_names.definedName)}
        except:
            return {"count": 0, "names": []}
    
    def _analyze_sheet_relationships(self, wb) -> Dict[str, Any]:
        """Analyze relationships between sheets"""
        return {"cross_sheet_references": 0, "relationships": []}
    
    def _count_charts(self, wb) -> int:
        """Count charts in workbook"""
        try:
            count = 0
            for ws in wb.worksheets:
                count += len(ws._charts)
            return count
        except:
            return 0
    
    def _count_pivot_tables(self, wb) -> int:
        """Count pivot tables in workbook"""
        try:
            count = 0
            for ws in wb.worksheets:
                count += len(ws._pivots)
            return count
        except:
            return 0
    
    def _calculate_cell_statistics(self, wb, include_hidden: bool = True) -> tuple:
        """Calculate cell statistics"""
        try:
            total_cells = 0
            non_empty_cells = 0
            
            for ws in wb.worksheets:
                if not include_hidden and ws.sheet_state != "visible":
                    continue
                    
                total_cells += ws.max_row * ws.max_column
                
                # Count non-empty cells (sample-based for performance)
                sample_rows = min(100, ws.max_row)
                for row in ws.iter_rows(max_row=sample_rows, values_only=True):
                    non_empty_cells += sum(1 for cell in row if cell is not None)
            
            return total_cells, non_empty_cells
        except:
            return 0, 0
    
    def _validate_result(self, data: StructureData, context: AnalysisContext) -> ValidationResult:
        """Validate structure mapping results
        
        Args:
            data: StructureData to validate
            context: AnalysisContext for additional validation
            
        Returns:
            ValidationResult with quality metrics
        """
        validation_notes = []
        
        # Completeness check
        completeness = 0.0
        if data.worksheet_names:
            completeness += 0.3
        if data.named_ranges is not None:
            completeness += 0.2
        if data.total_cells > 0:
            completeness += 0.3
        if data.chart_count >= 0:
            completeness += 0.1
        if data.pivot_table_count >= 0:
            completeness += 0.1
        
        # Accuracy checks
        accuracy = 1.0
        
        # Basic sanity checks
        if data.worksheet_count != len(data.worksheet_names):
            accuracy -= 0.3
            validation_notes.append("Worksheet count mismatch")
        
        if data.total_cells < data.non_empty_cells:
            accuracy -= 0.4
            validation_notes.append("Cell count inconsistency")
        
        if data.worksheet_count <= 0:
            accuracy -= 0.5
            validation_notes.append("No worksheets detected")
        
        # Consistency checks
        consistency = 0.9
        
        # Check for logical relationships
        if data.hidden_sheets:
            hidden_count = len(data.hidden_sheets)
            visible_count = data.worksheet_count - hidden_count
            if visible_count < 0:
                consistency -= 0.3
                validation_notes.append("Hidden sheet count inconsistency")
        
        # Confidence assessment
        if accuracy > 0.9 and completeness > 0.8:
            confidence = ConfidenceLevel.HIGH
        elif accuracy > 0.7 and completeness > 0.6:
            confidence = ConfidenceLevel.MEDIUM
        else:
            confidence = ConfidenceLevel.LOW
        
        if data.worksheet_count > 20:
            validation_notes.append("Large workbook detected")
        if len(data.named_ranges) > 50:
            validation_notes.append("Many named ranges detected")
        
        return ValidationResult(
            completeness_score=completeness,
            accuracy_score=max(0.0, accuracy),
            consistency_score=max(0.0, consistency),
            confidence_level=confidence,
            validation_notes=validation_notes
        )
    
    def _analyze_named_ranges(self, wb) -> Dict[str, str]:
        """Analyze workbook named ranges
        
        Args:
            wb: openpyxl Workbook object
            
        Returns:
            Dict mapping name to reference string
        """
        named_ranges = {}
        
        try:
            # Handle different openpyxl versions
            defined_names = getattr(wb.defined_names, "definedName", wb.defined_names)
            
            for defn in defined_names:
                # Skip auto-generated names (filters, print areas, etc.)
                if defn.name.startswith("_xlnm"):
                    continue
                
                # Skip names with local scope for now
                if defn.localSheetId is not None:
                    continue
                
                named_ranges[defn.name] = defn.attr_text or ""
                
        except Exception as e:
            self.logger.warning(f"Error analyzing named ranges: {e}")
        
        return named_ranges
    
    def _analyze_sheet_relationships(self, wb) -> Dict[str, List[str]]:
        """Analyze references between worksheets
        
        Args:
            wb: openpyxl Workbook object
            
        Returns:
            Dict mapping sheet names to referenced sheets
        """
        relationships = {}
        
        try:
            for ws in wb.worksheets:
                referenced_sheets = set()
                
                # Sample cells to find cross-sheet references
                # For performance, only check a subset of cells
                max_checks = self.config.get("max_cell_checks", 1000)
                checked = 0
                
                for row in ws.iter_rows(values_only=False):
                    if checked >= max_checks:
                        break
                    
                    for cell in row:
                        if checked >= max_checks:
                            break
                        
                        if cell.value and isinstance(cell.value, str):
                            # Look for sheet references in formulas
                            if cell.value.startswith('=') and '!' in cell.value:
                                # Simple parsing - look for SheetName! patterns
                                formula = cell.value
                                for sheet_name in wb.sheetnames:
                                    if f"'{sheet_name}'!" in formula or f"{sheet_name}!" in formula:
                                        referenced_sheets.add(sheet_name)
                        
                        checked += 1
                
                if referenced_sheets:
                    relationships[ws.title] = list(referenced_sheets)
                    
        except Exception as e:
            self.logger.warning(f"Error analyzing sheet relationships: {e}")
        
        return relationships
    
    def _count_charts(self, wb) -> int:
        """Count charts across all worksheets
        
        Args:
            wb: openpyxl Workbook object
            
        Returns:
            int: Total chart count
        """
        chart_count = 0
        
        try:
            for ws in wb.worksheets:
                chart_count += len(ws._charts)
        except Exception as e:
            self.logger.warning(f"Error counting charts: {e}")
        
        return chart_count
    
    def _count_pivot_tables(self, wb) -> int:
        """Count pivot tables across all worksheets
        
        Args:
            wb: openpyxl Workbook object
            
        Returns:
            int: Total pivot table count
        """
        pivot_count = 0
        
        try:
            for ws in wb.worksheets:
                # Count pivot table objects
                if hasattr(ws, '_pivots'):
                    pivot_count += len(ws._pivots)
        except Exception as e:
            self.logger.warning(f"Error counting pivot tables: {e}")
        
        return pivot_count
    
    def _calculate_cell_statistics(self, wb, include_hidden: bool = True) -> tuple[int, int]:
        """Calculate total and non-empty cell counts
        
        Args:
            wb: openpyxl Workbook object
            include_hidden: Whether to include hidden sheets
            
        Returns:
            Tuple of (total_cells, non_empty_cells)
        """
        total_cells = 0
        non_empty_cells = 0
        
        try:
            for ws in wb.worksheets:
                # Skip hidden sheets if not including them
                if not include_hidden and ws.sheet_state != "visible":
                    continue
                
                # Get worksheet dimensions
                if ws.max_row and ws.max_column:
                    sheet_total = ws.max_row * ws.max_column
                    total_cells += sheet_total
                    
                    # Count non-empty cells (sample for performance)
                    max_sample = self.config.get("max_cell_sample", 10000)
                    sampled = 0
                    sheet_non_empty = 0
                    
                    for row in ws.iter_rows(values_only=True):
                        if sampled >= max_sample:
                            # Extrapolate from sample
                            if sampled > 0:
                                ratio = sheet_non_empty / sampled
                                sheet_non_empty = int(sheet_total * ratio)
                            break
                        
                        for cell_value in row:
                            if cell_value is not None and cell_value != "":
                                sheet_non_empty += 1
                            sampled += 1
                            
                            if sampled >= max_sample:
                                break
                    
                    non_empty_cells += sheet_non_empty
                    
        except Exception as e:
            self.logger.warning(f"Error calculating cell statistics: {e}")
        
        return total_cells, non_empty_cells
    
    def estimate_complexity(self, context: AnalysisContext) -> float:
        """Estimate complexity based on workbook size and sheet count
        
        Args:
            context: AnalysisContext with file metadata
            
        Returns:
            float: Complexity multiplier
        """
        base_complexity = super().estimate_complexity(context)
        
        try:
            # Try to get sheet count for better estimation
            sheet_names = context.get_workbook_access().get_sheet_names()
            sheet_count = len(sheet_names)
            
            if sheet_count > 20:
                return base_complexity * 2.0
            elif sheet_count > 10:
                return base_complexity * 1.5
            else:
                return base_complexity
                
        except Exception:
            return base_complexity
    
    def _count_processed_items(self, data: StructureData) -> int:
        """Count items processed during analysis
        
        Args:
            data: StructureData result
            
        Returns:
            int: Number of processed items
        """
        return (
            data.worksheet_count +
            len(data.named_ranges) +
            data.chart_count +
            data.pivot_table_count
        )


# Legacy compatibility
def create_structure_mapper(config: dict = None) -> StructureMapper:
    """Factory function for backward compatibility"""
    mapper = StructureMapper()
    if config:
        mapper.configure(config)
    return mapper
