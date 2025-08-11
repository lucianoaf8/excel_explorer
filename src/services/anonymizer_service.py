#!/usr/bin/env python3
"""
Shared Anonymizer Service
Provides unified interface for anonymization operations used by both CLI and GUI
"""

from pathlib import Path
from typing import Dict, List, Optional, Tuple, Callable, Any
import json
from datetime import datetime

from utils.anonymizer import ExcelAnonymizer, anonymize_file


class AnonymizerService:
    """
    Unified service for anonymization operations
    Eliminates code duplication between CLI and GUI
    """
    
    def __init__(self, progress_callback: Optional[Callable[[str, float], None]] = None):
        """
        Initialize the anonymizer service
        
        Args:
            progress_callback: Optional callback for progress updates (message, progress)
        """
        self.progress_callback = progress_callback
        self._update_progress = self._create_progress_updater()
    
    def _create_progress_updater(self):
        """Create a progress updater function"""
        def update_progress(message: str, progress: float = 0.0):
            if self.progress_callback:
                self.progress_callback(message, progress)
        return update_progress
    
    def detect_sensitive_columns(self, file_path: str) -> Dict[str, List[Tuple[str, str]]]:
        """
        Detect sensitive columns in an Excel file
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary mapping sheet names to list of (column, type) tuples
        """
        try:
            self._update_progress("Detecting sensitive columns...", 0.1)
            anonymizer = ExcelAnonymizer(file_path)
            sensitive = anonymizer.find_sensitive_columns()
            self._update_progress("Column detection complete", 1.0)
            return sensitive
        except Exception as e:
            self._update_progress(f"Detection failed: {e}", 0.0)
            raise
    
    def anonymize_excel_file(self,
                           file_path: str,
                           output_path: Optional[str] = None,
                           mapping_path: Optional[str] = None,
                           columns: Optional[Dict[str, List[Tuple[str, str]]]] = None,
                           auto_detect: bool = True,
                           mapping_format: str = 'json') -> Tuple[str, str, Dict[str, int]]:
        """
        Anonymize an Excel file with progress tracking
        
        Args:
            file_path: Path to input Excel file
            output_path: Output path for anonymized file
            mapping_path: Output path for mapping file
            columns: Specific columns to anonymize
            auto_detect: Whether to auto-detect sensitive columns
            mapping_format: Format for mapping file ('json' or 'excel')
            
        Returns:
            Tuple of (anonymized_file_path, mapping_file_path, stats)
        """
        try:
            self._update_progress("Initializing anonymizer...", 0.1)
            anonymizer = ExcelAnonymizer(file_path)
            
            # Detect columns if needed
            if columns is None and auto_detect:
                self._update_progress("Detecting sensitive columns...", 0.2)
                columns = anonymizer.find_sensitive_columns()
                if columns:
                    total_cols = sum(len(cols) for cols in columns.values())
                    self._update_progress(f"Found {total_cols} sensitive columns", 0.3)
                else:
                    self._update_progress("No sensitive columns detected", 0.3)
                    return None, None, {}
            
            # Anonymize columns
            self._update_progress("Anonymizing data...", 0.4)
            stats = anonymizer.anonymize_columns(columns, auto_detect=False)
            
            if not stats:
                self._update_progress("No columns were anonymized", 0.4)
                return None, None, {}
            
            total_values = sum(stats.values())
            self._update_progress(f"Anonymized {total_values} values", 0.7)
            
            # Save files
            self._update_progress("Saving anonymized file...", 0.8)
            anon_path = anonymizer.save_anonymized_file(output_path)
            
            self._update_progress("Saving mapping file...", 0.9)
            map_path = anonymizer.save_mappings(mapping_path, mapping_format)
            
            self._update_progress("Anonymization complete!", 1.0)
            return anon_path, map_path, stats
            
        except Exception as e:
            self._update_progress(f"Anonymization failed: {e}", 0.0)
            raise
    
    def reverse_anonymization(self,
                            anonymized_file: str,
                            mapping_file: str,
                            output_path: Optional[str] = None) -> str:
        """
        Reverse anonymization using a mapping file
        
        Args:
            anonymized_file: Path to anonymized Excel file
            mapping_file: Path to mapping file
            output_path: Output path for restored file
            
        Returns:
            Path to restored file
        """
        try:
            self._update_progress("Loading mapping file...", 0.2)
            
            # Create anonymizer instance for reversal
            anonymizer = ExcelAnonymizer(anonymized_file)
            
            self._update_progress("Reversing anonymization...", 0.5)
            restored_path = anonymizer.reverse_anonymization(
                anonymized_file, mapping_file, output_path
            )
            
            self._update_progress("Reversal complete!", 1.0)
            return restored_path
            
        except Exception as e:
            self._update_progress(f"Reversal failed: {e}", 0.0)
            raise
    
    def get_mapping_summary(self, mapping_file: str) -> Dict[str, Any]:
        """
        Get summary information from a mapping file
        
        Args:
            mapping_file: Path to mapping file
            
        Returns:
            Dictionary with mapping summary information
        """
        try:
            mapping_path = Path(mapping_file)
            
            if mapping_path.suffix == '.json':
                with open(mapping_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    
                return {
                    'format': 'json',
                    'created': data.get('metadata', {}).get('created'),
                    'source_file': data.get('metadata', {}).get('source_file'),
                    'total_mappings': data.get('metadata', {}).get('total_mappings', 0),
                    'columns': data.get('metadata', {}).get('columns', []),
                    'column_counts': {
                        col: len(mappings) 
                        for col, mappings in data.get('mappings', {}).items()
                    }
                }
            else:
                # Excel format
                import openpyxl
                wb = openpyxl.load_workbook(mapping_path)
                
                column_counts = {}
                for sheet_name in wb.sheetnames:
                    if sheet_name != 'Summary':
                        ws = wb[sheet_name]
                        column_counts[sheet_name] = ws.max_row - 1  # Exclude header
                
                return {
                    'format': 'excel',
                    'created': None,  # Not stored in Excel format
                    'source_file': None,
                    'total_mappings': sum(column_counts.values()),
                    'columns': list(column_counts.keys()),
                    'column_counts': column_counts
                }
                
        except Exception as e:
            raise Exception(f"Failed to read mapping file: {e}")
    
    def validate_files(self, file_path: str) -> bool:
        """
        Validate that file exists and is a supported Excel format
        
        Args:
            file_path: Path to file to validate
            
        Returns:
            True if valid, False otherwise
        """
        try:
            path = Path(file_path)
            return (path.exists() and 
                   path.suffix.lower() in ['.xlsx', '.xls', '.xlsm'])
        except Exception:
            return False


class AnonymizerResults:
    """Container for anonymization results"""
    
    def __init__(self, 
                 success: bool,
                 anonymized_file: Optional[str] = None,
                 mapping_file: Optional[str] = None,
                 stats: Optional[Dict[str, int]] = None,
                 error: Optional[str] = None):
        self.success = success
        self.anonymized_file = anonymized_file
        self.mapping_file = mapping_file
        self.stats = stats or {}
        self.error = error
        self.total_values = sum(stats.values()) if stats else 0
    
    def has_results(self) -> bool:
        """Check if anonymization produced results"""
        return self.success and bool(self.stats)
    
    def get_summary(self) -> str:
        """Get a summary string of the results"""
        if not self.success:
            return f"Failed: {self.error}"
        
        if not self.has_results():
            return "No columns were anonymized"
        
        return f"Anonymized {self.total_values} values across {len(self.stats)} columns"