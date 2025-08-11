"""
Structure Analyzer - Analyzes Excel workbook structure and organization
"""

from typing import Dict, Any
import openpyxl
from .base import BaseAnalyzer


class StructureAnalyzer(BaseAnalyzer):
    """Analyzes workbook structure, sheets, and organization features"""
    
    def analyze(self, workbook: openpyxl.Workbook) -> Dict[str, Any]:
        """
        Perform comprehensive workbook structure analysis
        
        Args:
            workbook: Loaded openpyxl workbook
            
        Returns:
            Dictionary containing structure analysis results
        """
        self.start_timing()
        self.log_progress("Starting structure analysis")
        
        visible_sheets = []
        hidden_sheets = []
        sheet_details = []
        
        # Analyze each sheet
        for ws in workbook.worksheets:
            sheet_detail = {
                'name': ws.title,
                'state': ws.sheet_state,
                'max_row': ws.max_row,
                'max_column': ws.max_column,
                'dimensions': f"{ws.max_row}x{ws.max_column}",
                'status': self._classify_sheet_status(ws),
                'has_protection': self._check_sheet_protection(ws),
                'tab_color': self._get_tab_color(ws)
            }
            
            sheet_details.append(sheet_detail)
            
            if ws.sheet_state == 'visible':
                visible_sheets.append(ws.title)
            else:
                hidden_sheets.append(ws.title)
        
        # Enhanced workbook features detection
        features = self._detect_workbook_features(workbook)
        
        # Named ranges analysis
        named_ranges_info = self._analyze_named_ranges(workbook)
        
        # Table structures
        table_info = self._analyze_table_structures(workbook)
        
        # Workbook protection
        protection_info = self._analyze_workbook_protection(workbook)
        
        self.log_progress(f"Structure analysis completed in {self.get_duration():.2f}s")
        
        return {
            'total_sheets': len(workbook.sheetnames),
            'visible_sheets': [ws.title for ws in workbook.worksheets if getattr(ws, 'sheet_state', 'visible') == 'visible'],
            'hidden_sheets': [ws.title for ws in workbook.worksheets if getattr(ws, 'sheet_state', 'visible') != 'visible'],
            'sheet_details': sheet_details,
            'named_ranges_count': named_ranges_info['count'],
            'named_ranges_list': named_ranges_info['ranges'],
            'table_count': table_info['count'],
            'table_details': table_info['tables'],
            'has_hidden_content': len(hidden_sheets) > 0,
            'workbook_features': features,
            'protection_info': protection_info,
            'analysis_duration': self.get_duration()
        }
    
    def _classify_sheet_status(self, ws) -> str:
        """Classify sheet status based on size and content"""
        if not ws.max_row or not ws.max_column:
            return 'Empty'
        
        cell_count = ws.max_row * ws.max_column
        if cell_count > 100000:  # 100k cells
            return 'Large'
        elif cell_count > 10000:  # 10k cells
            return 'Medium'
        else:
            return 'Small'
    
    def _check_sheet_protection(self, ws) -> bool:
        """Check if sheet has protection enabled"""
        try:
            return getattr(ws, 'protection', type('obj', (object,), {'sheet': False})).sheet if hasattr(ws, 'protection') else False
        except:
            return False
    
    def _get_tab_color(self, ws) -> str:
        """Get sheet tab color if set"""
        try:
            return getattr(getattr(ws, 'sheet_properties', None), 'tabColor', None) if hasattr(ws, 'sheet_properties') else None
        except:
            return None
    
    def _detect_workbook_features(self, wb) -> Dict[str, Any]:
        """Detect various workbook features"""
        features = {
            'has_macros': False,
            'has_external_connections': False,
            'has_pivot_tables': 0,
            'data_validation_rules': 0,
            'conditional_formatting_rules': 0,
            'print_areas_count': 0,
            'freeze_panes_count': 0,
            'hyperlinks_count': 0,
            'comments_count': 0,
            'images_count': 0,
            'charts_count': 0
        }
        
        # Check for macros (VBA)
        try:
            if hasattr(wb, 'vba_archive') and wb.vba_archive:
                features['has_macros'] = True
        except:
            pass
        
        # Analyze each sheet for features
        for ws in wb.worksheets:
            # Data validation rules
            try:
                if hasattr(ws, 'data_validations'):
                    features['data_validation_rules'] += len(ws.data_validations.dataValidation)
            except:
                pass
            
            # Conditional formatting
            try:
                features['conditional_formatting_rules'] += len(ws.conditional_formatting)
            except:
                pass
            
            # Print areas
            try:
                if ws.print_area:
                    features['print_areas_count'] += 1
            except:
                pass
            
            # Freeze panes
            try:
                if ws.freeze_panes:
                    features['freeze_panes_count'] += 1
            except:
                pass
            
            # Count hyperlinks, comments, images, charts
            try:
                features['hyperlinks_count'] += len(ws._hyperlinks)
            except:
                pass
            
            try:
                features['comments_count'] += len(ws._comments)
            except:
                pass
            
            try:
                features['images_count'] += len(ws._images)
            except:
                pass
            
            try:
                features['charts_count'] += len(ws._charts)
            except:
                pass
        
        return features
    
    def _analyze_named_ranges(self, wb) -> Dict[str, Any]:
        """Analyze named ranges in workbook"""
        named_ranges = []
        count = 0
        
        try:
            for defined_name in wb.defined_names.definedName:
                count += 1
                named_ranges.append({
                    'name': defined_name.name,
                    'refers_to': str(defined_name.attr_text),
                    'scope': getattr(defined_name, 'localSheetId', 'Workbook')
                })
        except:
            pass
        
        return {
            'count': count,
            'ranges': named_ranges[:20]  # Limit to first 20 for performance
        }
    
    def _analyze_table_structures(self, wb) -> Dict[str, Any]:
        """Analyze Excel table structures"""
        tables = []
        count = 0
        
        try:
            for ws in wb.worksheets:
                for table in ws.tables:
                    count += 1
                    tables.append({
                        'name': table.name,
                        'sheet': ws.title,
                        'range': str(table.ref),
                        'style': table.tableStyleInfo.name if table.tableStyleInfo else 'None'
                    })
        except:
            pass
        
        return {
            'count': count,
            'tables': tables
        }
    
    def _analyze_workbook_protection(self, wb) -> Dict[str, Any]:
        """Analyze workbook protection settings"""
        protection_info = {
            'workbook_protected': False,
            'password_protected': False,
            'protected_sheets': [],
            'protection_features': []
        }
        
        try:
            # Check workbook protection
            if hasattr(wb, 'security') and wb.security:
                protection_info['workbook_protected'] = True
                if wb.security.workbookPassword:
                    protection_info['password_protected'] = True
        except:
            pass
        
        # Check individual sheet protection
        for ws in wb.worksheets:
            try:
                if ws.protection.sheet:
                    protection_info['protected_sheets'].append({
                        'sheet': ws.title,
                        'password': bool(ws.protection.password),
                        'select_locked_cells': ws.protection.selectLockedCells,
                        'select_unlocked_cells': ws.protection.selectUnlockedCells,
                        'format_cells': ws.protection.formatCells,
                        'format_columns': ws.protection.formatColumns,
                        'format_rows': ws.protection.formatRows,
                        'insert_columns': ws.protection.insertColumns,
                        'insert_rows': ws.protection.insertRows,
                        'delete_columns': ws.protection.deleteColumns,
                        'delete_rows': ws.protection.deleteRows
                    })
            except:
                pass
        
        return protection_info