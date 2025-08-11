"""
Data Analyzer - Analyzes Excel data content, types, and quality
"""

from typing import Dict, Any
from collections import Counter, defaultdict
import openpyxl
from openpyxl.utils import get_column_letter
from .base import BaseAnalyzer


class DataAnalyzer(BaseAnalyzer):
    """Analyzes data content, types, quality, and patterns"""
    
    def analyze(self, workbook: openpyxl.Workbook) -> Dict[str, Any]:
        """
        Perform comprehensive data analysis
        
        Args:
            workbook: Loaded openpyxl workbook
            
        Returns:
            Dictionary containing data analysis results
        """
        self.start_timing()
        self.log_progress("Starting data analysis")
        
        sheet_data = {}
        total_cells = 0
        total_data_cells = 0
        overall_data_types = Counter()
        
        for ws in workbook.worksheets:
            if not ws.max_row or not ws.max_column:
                sheet_data[ws.title] = self._get_empty_sheet_data()
                continue
            
            sheet_cells = ws.max_row * ws.max_column
            total_cells += sheet_cells
            
            # Get sample rows based on sheet size
            sample_rows = self._get_sample_rows(ws)
            
            # Analyze data quality and types
            data_cells, type_distribution = self._analyze_sheet_data(ws, sample_rows)
            total_data_cells += data_cells
            
            # Update overall data type distribution
            for data_type, count in type_distribution.items():
                overall_data_types[data_type] += count
            
            # Calculate sheet metrics
            sheet_data[ws.title] = {
                'dimensions': f"{ws.max_row}x{ws.max_column}",
                'used_range': self._get_used_range(ws),
                'estimated_data_cells': data_cells,
                'empty_cells': sheet_cells - data_cells,
                'has_data': data_cells > 0,
                'data_density': data_cells / sheet_cells if sheet_cells > 0 else 0,
                'data_type_distribution': dict(type_distribution),
                'column_analysis': self._analyze_columns(ws, sample_rows),
                'data_quality_score': self._calculate_data_quality_score(ws, sample_rows)
            }
        
        # Calculate overall metrics
        overall_quality = self._calculate_overall_quality(total_cells, total_data_cells, overall_data_types)
        
        self.log_progress(f"Data analysis completed in {self.get_duration():.2f}s")
        
        return {
            'sheet_analysis': sheet_data,
            'total_cells': total_cells,
            'total_data_cells': total_data_cells,
            'overall_data_density': total_data_cells / max(1, total_cells),
            'data_type_distribution': dict(overall_data_types),
            'data_quality_score': overall_quality,
            'analysis_duration': self.get_duration()
        }
    
    def _get_empty_sheet_data(self) -> Dict[str, Any]:
        """Return data structure for empty sheets"""
        return {
            'dimensions': '0x0',
            'used_range': 'A1:A1',
            'estimated_data_cells': 0,
            'empty_cells': 0,
            'has_data': False,
            'data_density': 0.0,
            'data_type_distribution': {},
            'column_analysis': [],
            'data_quality_score': 0.0
        }
    
    def _get_sample_rows(self, ws) -> int:
        """Get appropriate sample size based on sheet size"""
        sample_rows = self.get_sample_limit()
        
        # Adjust sample size based on sheet size
        if ws.max_row > 100000 or ws.max_column > 100:
            sample_rows = min(50, sample_rows)
        elif ws.max_row > 10000 or ws.max_column > 50:
            sample_rows = min(75, sample_rows)
        
        return min(sample_rows, ws.max_row)
    
    def _get_used_range(self, ws) -> str:
        """Get the used range of the worksheet"""
        try:
            return getattr(ws, 'dimensions', f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        except:
            return f"A1:{get_column_letter(ws.max_column)}{ws.max_row}" if ws.max_row and ws.max_column else 'A1:A1'
    
    def _analyze_sheet_data(self, ws, sample_rows: int) -> tuple[int, Counter]:
        """Analyze data content in a sheet"""
        data_cells = 0
        type_distribution = Counter()
        
        try:
            # Sample data from the sheet
            for row in ws.iter_rows(max_row=sample_rows, values_only=True):
                for cell_value in row:
                    if cell_value is not None and str(cell_value).strip():
                        data_cells += 1
                        cell_type = self._classify_cell_type(cell_value)
                        type_distribution[cell_type] += 1
        except Exception as e:
            self.log_progress(f"Error analyzing sheet data: {e}", 'warning')
        
        # Estimate total data cells if we sampled
        if ws.max_row > sample_rows:
            data_cells = int(data_cells * (ws.max_row / sample_rows))
        
        return data_cells, type_distribution
    
    def _classify_cell_type(self, value) -> str:
        """Classify the type of a cell value"""
        if value is None:
            return 'empty'
        
        if isinstance(value, (int, float)):
            return 'numeric'
        elif isinstance(value, bool):
            return 'boolean'
        elif hasattr(value, 'date'):  # datetime objects
            return 'date'
        else:
            str_value = str(value).strip()
            if not str_value:
                return 'empty'
            elif str_value.lower() in ['true', 'false', 'yes', 'no']:
                return 'boolean_text'
            elif self._is_numeric_string(str_value):
                return 'numeric_text'
            elif self._is_date_string(str_value):
                return 'date_text'
            else:
                return 'text'
    
    def _is_numeric_string(self, value: str) -> bool:
        """Check if string represents a number"""
        try:
            float(value.replace(',', ''))
            return True
        except ValueError:
            return False
    
    def _is_date_string(self, value: str) -> bool:
        """Check if string represents a date"""
        import re
        # Simple date pattern matching
        date_patterns = [
            r'\\d{1,2}[/-]\\d{1,2}[/-]\\d{2,4}',  # MM/DD/YYYY or similar
            r'\\d{4}[/-]\\d{1,2}[/-]\\d{1,2}',    # YYYY-MM-DD
        ]
        
        for pattern in date_patterns:
            if re.match(pattern, value):
                return True
        return False
    
    def _analyze_columns(self, ws, sample_rows: int) -> list:
        """Analyze individual columns"""
        columns = []
        
        try:
            for col_num in range(1, min(ws.max_column + 1, 51)):  # Limit to 50 columns for performance
                col_letter = get_column_letter(col_num)
                col_data = []
                
                # Sample column data
                for row_num in range(1, min(sample_rows + 1, ws.max_row + 1)):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    if cell_value is not None:
                        col_data.append(cell_value)
                
                if col_data:
                    col_types = Counter(self._classify_cell_type(val) for val in col_data)
                    dominant_type = col_types.most_common(1)[0][0]
                    
                    columns.append({
                        'letter': col_letter,
                        'number': col_num,
                        'data_type': dominant_type,
                        'fill_rate': len(col_data) / sample_rows,
                        'unique_values': len(set(str(val) for val in col_data)),
                        'sample_count': len(col_data),
                        'type_distribution': dict(col_types)
                    })
        
        except Exception as e:
            self.log_progress(f"Error analyzing columns: {e}", 'warning')
        
        return columns
    
    def _calculate_data_quality_score(self, ws, sample_rows: int) -> float:
        """Calculate overall data quality score for the sheet"""
        try:
            total_cells = 0
            filled_cells = 0
            
            for row in ws.iter_rows(max_row=sample_rows, values_only=True):
                for cell_value in row:
                    total_cells += 1
                    if cell_value is not None and str(cell_value).strip():
                        filled_cells += 1
            
            if total_cells == 0:
                return 0.0
            
            # Quality score based on data density
            fill_rate = filled_cells / total_cells
            
            # Adjust for sheet size (larger sheets with good fill rates get higher scores)
            size_factor = min(1.0, ws.max_row * ws.max_column / 10000)  # Normalize to 10k cells
            
            return min(1.0, fill_rate * (0.7 + 0.3 * size_factor))
        
        except Exception as e:
            self.log_progress(f"Error calculating quality score: {e}", 'warning')
            return 0.0
    
    def _calculate_overall_quality(self, total_cells: int, total_data_cells: int, type_distribution: Counter) -> float:
        """Calculate overall data quality across all sheets"""
        if total_cells == 0:
            return 0.0
        
        # Base quality on data density
        data_density = total_data_cells / total_cells
        
        # Bonus for type diversity (indicates structured data)
        type_diversity = len(type_distribution) / 6.0  # Max 6 basic types
        
        # Combined quality score
        quality_score = (data_density * 0.8) + (type_diversity * 0.2)
        
        return min(1.0, quality_score)