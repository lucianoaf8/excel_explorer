"""
Formula Analyzer - Analyzes Excel formulas, dependencies, and complexity
"""

from typing import Dict, Any, List
import openpyxl
from .base import BaseAnalyzer


class FormulaAnalyzer(BaseAnalyzer):
    """Analyzes formulas, dependencies, and computational complexity"""
    
    def analyze(self, workbook: openpyxl.Workbook) -> Dict[str, Any]:
        """
        Perform comprehensive formula analysis
        
        Args:
            workbook: Loaded openpyxl workbook
            
        Returns:
            Dictionary containing formula analysis results
        """
        self.start_timing()
        self.log_progress("Starting formula analysis")
        
        total_formulas = 0
        complex_formulas = []
        external_refs = False
        formula_functions = {}
        circular_refs = []
        sheet_formulas = {}
        
        max_check = self.config.get('analysis', {}).get('max_formula_check', 1000)
        
        for ws in workbook.worksheets:
            sheet_formula_count = 0
            sheet_complexity_score = 0.0
            checked_cells = 0
            
            for row in ws.iter_rows():
                if checked_cells >= max_check:
                    break
                    
                for cell in row:
                    if checked_cells >= max_check:
                        break
                    
                    checked_cells += 1
                    
                    if self._is_formula_cell(cell):
                        total_formulas += 1
                        sheet_formula_count += 1
                        formula = str(cell.value)
                        
                        # Analyze formula complexity
                        complexity = self._analyze_formula_complexity(formula)
                        sheet_complexity_score += complexity['score']
                        
                        # Track complex formulas
                        if complexity['score'] > 0.7:  # High complexity threshold
                            complex_formulas.append({
                                'sheet': ws.title,
                                'cell': cell.coordinate,
                                'formula': formula[:100],  # Truncate for display
                                'complexity_score': complexity['score'],
                                'issues': complexity['issues']
                            })
                        
                        # Check for external references
                        if self._has_external_references(formula):
                            external_refs = True
                        
                        # Extract function usage
                        functions = self._extract_functions(formula)
                        for func in functions:
                            formula_functions[func] = formula_functions.get(func, 0) + 1
                        
                        # Check for potential circular references
                        if self._might_be_circular_ref(formula, cell.coordinate):
                            circular_refs.append({
                                'sheet': ws.title,
                                'cell': cell.coordinate,
                                'formula': formula[:50]
                            })
            
            # Store sheet-level formula statistics
            if sheet_formula_count > 0:
                sheet_formulas[ws.title] = {
                    'formula_count': sheet_formula_count,
                    'average_complexity': sheet_complexity_score / sheet_formula_count,
                    'formula_density': sheet_formula_count / (ws.max_row * ws.max_column) if ws.max_row and ws.max_column else 0
                }
        
        # Calculate overall metrics
        formula_complexity_score = self._calculate_overall_complexity(complex_formulas, total_formulas)
        function_diversity = len(formula_functions)
        most_used_functions = sorted(formula_functions.items(), key=lambda x: x[1], reverse=True)[:10]
        
        self.log_progress(f"Formula analysis completed in {self.get_duration():.2f}s")
        
        return {
            'total_formulas': total_formulas,
            'complex_formulas': complex_formulas[:10],  # Top 10 most complex
            'has_external_refs': external_refs,
            'circular_references': circular_refs,
            'formula_complexity_score': formula_complexity_score,
            'function_usage': dict(formula_functions),
            'function_diversity': function_diversity,
            'most_used_functions': most_used_functions,
            'sheet_statistics': sheet_formulas,
            'analysis_duration': self.get_duration(),
            'performance_impact': self._assess_performance_impact(total_formulas, len(complex_formulas), external_refs)
        }
    
    def _is_formula_cell(self, cell) -> bool:
        """Check if cell contains a formula"""
        return (cell.value is not None and 
                isinstance(cell.value, str) and 
                cell.value.startswith('='))
    
    def _analyze_formula_complexity(self, formula: str) -> Dict[str, Any]:
        """Analyze individual formula complexity"""
        issues = []
        score = 0.0
        
        # Length-based complexity
        if len(formula) > 100:
            score += 0.3
            issues.append("Long formula")
        elif len(formula) > 50:
            score += 0.1
        
        # Nesting complexity
        nesting_level = formula.count('(')
        if nesting_level > 5:
            score += 0.4
            issues.append("Deep nesting")
        elif nesting_level > 3:
            score += 0.2
        
        # Function count
        function_count = len(self._extract_functions(formula))
        if function_count > 5:
            score += 0.3
            issues.append("Multiple functions")
        elif function_count > 3:
            score += 0.1
        
        # Array formulas or complex references
        if '{' in formula or '}' in formula:
            score += 0.4
            issues.append("Array formula")
        
        # Multiple sheet references
        if formula.count('!') > 2:
            score += 0.2
            issues.append("Multiple sheet references")
        
        # Complex functions that impact performance
        heavy_functions = ['SUMPRODUCT', 'INDIRECT', 'OFFSET', 'EVALUATE', 'VLOOKUP', 'INDEX']
        if any(func in formula.upper() for func in heavy_functions):
            score += 0.2
            issues.append("Performance-heavy functions")
        
        return {
            'score': min(1.0, score),  # Cap at 1.0
            'issues': issues,
            'nesting_level': nesting_level,
            'function_count': function_count,
            'length': len(formula)
        }
    
    def _has_external_references(self, formula: str) -> bool:
        """Check if formula has external file references"""
        return '[' in formula and ']' in formula
    
    def _extract_functions(self, formula: str) -> List[str]:
        """Extract function names from formula"""
        import re
        # Pattern to match function names (letters followed by opening parenthesis)
        function_pattern = r'([A-Z]+(?:[A-Z0-9_]*[A-Z0-9])?)\s*\('
        matches = re.findall(function_pattern, formula.upper())
        
        # Filter out common non-functions that might match the pattern
        exclude = {'IF', 'AND', 'OR', 'NOT'}  # These are actually functions, but very basic
        functions = [func for func in matches if len(func) > 1]
        return list(set(functions))  # Remove duplicates
    
    def _might_be_circular_ref(self, formula: str, cell_address: str) -> bool:
        """Simple heuristic to detect potential circular references"""
        # This is a simplified check - real circular reference detection is complex
        # Check if the cell references itself directly
        return cell_address in formula.upper()
    
    def _calculate_overall_complexity(self, complex_formulas: List[Dict], total_formulas: int) -> float:
        """Calculate overall formula complexity score"""
        if total_formulas == 0:
            return 0.0
        
        # Base score on proportion of complex formulas
        complexity_ratio = len(complex_formulas) / total_formulas
        
        # Weight by average complexity of complex formulas
        if complex_formulas:
            avg_complexity = sum(f['complexity_score'] for f in complex_formulas) / len(complex_formulas)
            return min(1.0, complexity_ratio * avg_complexity)
        
        return complexity_ratio
    
    def _assess_performance_impact(self, total_formulas: int, complex_count: int, has_external_refs: bool) -> Dict[str, Any]:
        """Assess potential performance impact of formulas"""
        impact_score = 0.0
        recommendations = []
        
        # High formula count
        if total_formulas > 1000:
            impact_score += 0.3
            recommendations.append("High formula count may slow recalculation")
        
        # Complex formulas
        if complex_count > total_formulas * 0.1:  # More than 10% complex
            impact_score += 0.4
            recommendations.append("Many complex formulas detected - consider simplification")
        
        # External references
        if has_external_refs:
            impact_score += 0.3
            recommendations.append("External references may cause performance issues if files are unavailable")
        
        # Determine impact level
        if impact_score > 0.7:
            impact_level = "High"
        elif impact_score > 0.4:
            impact_level = "Medium"
        else:
            impact_level = "Low"
        
        return {
            'impact_score': min(1.0, impact_score),
            'impact_level': impact_level,
            'recommendations': recommendations or ["Formula performance appears optimized"]
        }