"""
Analysis Service - High-level service for Excel file analysis
Provides a unified interface for both CLI and GUI applications
"""

import os
import time
from pathlib import Path
from typing import Dict, Any, Optional, List, Callable
import openpyxl
from .config import load_config
from .analyzers.orchestrator import AnalyzerOrchestrator


class AnalysisService:
    """High-level service for Excel file analysis"""
    
    def __init__(self, config_path: Optional[str] = None):
        """
        Initialize the analysis service
        
        Args:
            config_path: Optional path to configuration file
        """
        self.config = load_config(config_path)
        self.orchestrator = AnalyzerOrchestrator(self.config)
        self._last_results = None
        self._last_file_path = None
        
    def analyze_file(self, 
                    file_path: str,
                    modules: Optional[List[str]] = None,
                    progress_callback: Optional[Callable[[str, float], None]] = None) -> Dict[str, Any]:
        """
        Analyze an Excel file and return comprehensive results
        
        Args:
            file_path: Path to Excel file
            modules: List of analysis modules to run (runs all if None)
            progress_callback: Optional callback for progress updates (message, progress_0_to_1)
            
        Returns:
            Dictionary containing analysis results
            
        Raises:
            FileNotFoundError: If file doesn't exist
            Exception: If file cannot be opened or analyzed
        """
        # Validate file
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        file_path = str(Path(file_path).resolve())  # Normalize path
        
        if progress_callback:
            progress_callback("Opening Excel file...", 0.1)
        
        # Load workbook
        try:
            # Use read_only mode for better performance and memory usage
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
        except Exception as e:
            raise Exception(f"Failed to open Excel file: {str(e)}")
        
        if progress_callback:
            progress_callback("Initializing analysis...", 0.2)
        
        try:
            # Perform analysis using orchestrator
            def orchestrator_progress(message: str, progress: float):
                if progress_callback:
                    # Scale progress from 0.2 to 0.9 (leaving 0.1 for finalization)
                    scaled_progress = 0.2 + (progress * 0.7)
                    progress_callback(message, scaled_progress)
            
            results = self.orchestrator.analyze_workbook(
                workbook=workbook,
                modules=modules,
                progress_callback=orchestrator_progress
            )
            
            if progress_callback:
                progress_callback("Finalizing results...", 0.9)
            
            # Add file-specific information
            file_stats = os.stat(file_path)
            results['file_info'].update({
                'file_path': file_path,
                'file_name': Path(file_path).name,
                'file_size_bytes': file_stats.st_size,
                'size_mb': round(file_stats.st_size / (1024 * 1024), 2),
                'modified_time': file_stats.st_mtime,
                'analysis_timestamp': time.time()
            })
            
            # Cache results
            self._last_results = results
            self._last_file_path = file_path
            
            if progress_callback:
                progress_callback("Analysis complete!", 1.0)
            
            return results
            
        finally:
            # Clean up workbook
            if hasattr(workbook, 'close'):
                workbook.close()
    
    def get_quick_summary(self, file_path: str) -> Dict[str, Any]:
        """
        Get a quick summary without full analysis
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary with basic file information
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            
            file_stats = os.stat(file_path)
            
            # Quick metrics
            total_sheets = len(workbook.sheetnames)
            visible_sheets = []
            hidden_sheets = []
            
            for ws in workbook.worksheets:
                if getattr(ws, 'sheet_state', 'visible') == 'visible':
                    visible_sheets.append(ws.title)
                else:
                    hidden_sheets.append(ws.title)
            
            # Quick data estimation for first sheet
            first_sheet = workbook.active
            estimated_data_cells = 0
            if first_sheet and first_sheet.max_row and first_sheet.max_column:
                # Sample first 100 rows to estimate data density
                sample_size = min(100, first_sheet.max_row)
                data_count = 0
                total_sampled = 0
                
                for row in first_sheet.iter_rows(max_row=sample_size, values_only=True):
                    for cell_value in row:
                        total_sampled += 1
                        if cell_value is not None and str(cell_value).strip():
                            data_count += 1
                
                if total_sampled > 0:
                    density = data_count / total_sampled
                    estimated_data_cells = int(density * first_sheet.max_row * first_sheet.max_column)
            
            workbook.close()
            
            return {
                'file_name': Path(file_path).name,
                'file_size_mb': round(file_stats.st_size / (1024 * 1024), 2),
                'total_sheets': total_sheets,
                'visible_sheets': len(visible_sheets),
                'hidden_sheets': len(hidden_sheets),
                'estimated_data_cells': estimated_data_cells,
                'largest_sheet_dimensions': f"{first_sheet.max_row}x{first_sheet.max_column}" if first_sheet and first_sheet.max_row else "0x0",
                'analysis_type': 'quick_summary'
            }
            
        except Exception as e:
            return {
                'error': f"Failed to analyze file: {str(e)}",
                'file_name': Path(file_path).name,
                'analysis_type': 'quick_summary'
            }
    
    def validate_file(self, file_path: str) -> Dict[str, Any]:
        """
        Validate if file can be analyzed
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary with validation results
        """
        result = {
            'is_valid': False,
            'file_exists': False,
            'is_excel_file': False,
            'can_open': False,
            'errors': [],
            'warnings': []
        }
        
        # Check file existence
        if not os.path.exists(file_path):
            result['errors'].append(f"File not found: {file_path}")
            return result
        
        result['file_exists'] = True
        
        # Check file extension
        valid_extensions = ['.xlsx', '.xlsm', '.xlsb', '.xls']
        file_ext = Path(file_path).suffix.lower()
        
        if file_ext not in valid_extensions:
            result['warnings'].append(f"Unusual file extension: {file_ext}")
        else:
            result['is_excel_file'] = True
        
        # Try to open file
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            result['can_open'] = True
            
            # Additional checks
            if len(workbook.sheetnames) == 0:
                result['warnings'].append("No worksheets found in file")
            
            # Check file size
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            if file_size_mb > 100:
                result['warnings'].append(f"Large file size: {file_size_mb:.1f}MB may take longer to analyze")
            
            workbook.close()
            
        except Exception as e:
            result['errors'].append(f"Cannot open file: {str(e)}")
            return result
        
        # Set overall validity
        result['is_valid'] = result['file_exists'] and result['can_open'] and len(result['errors']) == 0
        
        return result
    
    def get_available_modules(self) -> List[str]:
        """Get list of available analysis modules"""
        return self.orchestrator.get_available_modules()
    
    def get_configuration(self) -> Dict[str, Any]:
        """Get current configuration"""
        return self.config.copy()
    
    def update_configuration(self, config_updates: Dict[str, Any]) -> None:
        """
        Update configuration (for this session only)
        
        Args:
            config_updates: Dictionary with configuration updates
        """
        # Deep merge configuration updates
        self._deep_update(self.config, config_updates)
        
        # Recreate orchestrator with new config
        self.orchestrator = AnalyzerOrchestrator(self.config)
    
    def _deep_update(self, target: dict, source: dict) -> None:
        """Recursively update nested dictionaries"""
        for key, value in source.items():
            if key in target and isinstance(target[key], dict) and isinstance(value, dict):
                self._deep_update(target[key], value)
            else:
                target[key] = value
    
    def get_last_results(self) -> Optional[Dict[str, Any]]:
        """Get results from the last analysis"""
        return self._last_results
    
    def get_analysis_history(self) -> List[str]:
        """Get list of recently analyzed files"""
        # This could be extended to maintain a persistent history
        return [self._last_file_path] if self._last_file_path else []