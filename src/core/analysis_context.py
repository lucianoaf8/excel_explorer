"""
Analysis Context - Central state management for Excel analysis pipeline
Manages workbook access, shared state, and inter-module coordination.
"""

import threading
import time
import logging
from typing import Dict, Any, Optional, List, Set, Union
from pathlib import Path
from contextlib import contextmanager
from dataclasses import dataclass, field
from enum import Enum

import openpyxl
import pandas as pd
from openpyxl.workbook.workbook import Workbook

from ..utils.memory_manager import MemoryManager, get_memory_manager
from ..utils.error_handler import ExcelAnalysisError, ErrorSeverity
from .unified_config import UnifiedConfig


class AnalysisPhase(Enum):
    """Analysis pipeline phases"""
    INITIALIZATION = "initialization"
    HEALTH_CHECK = "health_check"
    STRUCTURE_MAPPING = "structure_mapping"
    DATA_PROFILING = "data_profiling"
    FORMULA_ANALYSIS = "formula_analysis"
    VISUAL_ANALYSIS = "visual_analysis"
    CONNECTION_ANALYSIS = "connection_analysis"
    PIVOT_ANALYSIS = "pivot_analysis"
    SYNTHESIS = "synthesis"
    COMPLETE = "complete"


@dataclass
class FileMetadata:
    """File metadata and characteristics"""
    file_path: Path
    file_size_bytes: int
    file_size_mb: float = field(init=False)
    created_time: float
    modified_time: float
    excel_version: Optional[str] = None
    has_macros: bool = False
    is_protected: bool = False
    sheet_count: int = 0
    
    def __post_init__(self):
        self.file_size_mb = self.file_size_bytes / 1_048_576


@dataclass
class AnalysisConfig:
    """Analysis configuration and settings"""
    max_memory_mb: int = 4096
    warning_threshold: float = 0.8
    enable_caching: bool = True
    chunk_size_rows: int = 10000
    max_formulas_analyze: int = 50000
    include_charts: bool = True
    include_pivots: bool = True
    include_connections: bool = True
    deep_analysis: bool = False
    parallel_processing: bool = False


class SafeWorkbookAccess:
    """Thread-safe workbook access manager with resource cleanup"""
    
    def __init__(self, file_path: Path):
        self.file_path = file_path
        self._workbook: Optional[Workbook] = None
        self._lock = threading.RLock()
        self._ref_count = 0
        self._last_access = time.time()
        self._is_closed = False
        
    def _ensure_workbook(self) -> Workbook:
        """Lazy load workbook with error handling"""
        if self._workbook is None:
            try:
                self._workbook = openpyxl.load_workbook(
                    self.file_path, 
                    read_only=True, 
                    keep_vba=False,
                    data_only=True
                )
                logging.info(f"Loaded workbook: {self.file_path}")
            except Exception as e:
                raise ExcelAnalysisError(
                    f"Failed to load workbook: {e}",
                    ErrorSeverity.CRITICAL,
                    module_name="SafeWorkbookAccess",
                    file_path=str(self.file_path)
                )
        return self._workbook

    @contextmanager
    def get_workbook(self):
        """Context manager for safe workbook access"""
        with self._lock:
            if self._is_closed:
                raise ExcelAnalysisError(
                    "Workbook access after closure",
                    ErrorSeverity.CRITICAL,
                    module_name="SafeWorkbookAccess"
                )
            
            self._ref_count += 1
            self._last_access = time.time()
            
            try:
                workbook = self._ensure_workbook()
                yield workbook
            finally:
                self._ref_count -= 1

    def get_sheet_names(self) -> List[str]:
        """Get worksheet names safely"""
        with self.get_workbook() as wb:
            return wb.sheetnames

    def sheet_exists(self, sheet_name: str) -> bool:
        """Check if sheet exists"""
        return sheet_name in self.get_sheet_names()

    def get_sheet_dimensions(self, sheet_name: str) -> Dict[str, int]:
        """Get sheet dimensions (max row/column)"""
        with self.get_workbook() as wb:
            sheet = wb[sheet_name]
            return {
                'max_row': sheet.max_row,
                'max_column': sheet.max_column,
                'min_row': sheet.min_row,
                'min_column': sheet.min_column
            }

    def close(self):
        """Close workbook and cleanup resources"""
        with self._lock:
            if self._workbook is not None and not self._is_closed:
                try:
                    self._workbook.close()
                    logging.info(f"Closed workbook: {self.file_path}")
                except Exception as e:
                    logging.warning(f"Error closing workbook: {e}")
                finally:
                    self._workbook = None
                    self._is_closed = True

    @property
    def is_active(self) -> bool:
        """Check if workbook is actively being used"""
        return self._ref_count > 0

    @property
    def last_access_seconds_ago(self) -> float:
        """Seconds since last access"""
        return time.time() - self._last_access


class AnalysisContext:
    """Central coordination hub for Excel analysis pipeline"""
    
    def __init__(self, file_path: Union[str, Path], config: Optional[AnalysisConfig] = None):
        self.file_path = Path(file_path)
        self.config = config or AnalysisConfig()
        
        # Add unified config
        self.unified_config = UnifiedConfig.from_yaml()
        
        # Core components
        self.memory_manager = get_memory_manager()
        self.workbook_access = SafeWorkbookAccess(self.file_path)
        
        # Analysis state
        self.current_phase = AnalysisPhase.INITIALIZATION
        self.completed_modules: Set[str] = set()
        self.failed_modules: Set[str] = set()
        self.module_results: Dict[str, Any] = {}
        self.shared_cache: Dict[str, Any] = {}
        
        # Metadata
        self.file_metadata = self._initialize_file_metadata()
        self.analysis_start_time = time.time()
        self.phase_start_times: Dict[AnalysisPhase, float] = {}
        
        # Threading
        self._lock = threading.RLock()
        self._shutdown = False
        
        logging.info(f"AnalysisContext initialized for {self.file_path.name}")

    def _initialize_file_metadata(self) -> FileMetadata:
        """Initialize file metadata from filesystem"""
        stat = self.file_path.stat()
        return FileMetadata(
            file_path=self.file_path,
            file_size_bytes=stat.st_size,
            created_time=stat.st_ctime,
            modified_time=stat.st_mtime
        )

    def set_phase(self, phase: AnalysisPhase) -> None:
        """Update current analysis phase"""
        with self._lock:
            if self._shutdown:
                return
                
            self.current_phase = phase
            self.phase_start_times[phase] = time.time()
            
            logging.info(f"Analysis phase: {phase.value}")

    def register_module_result(self, module_name: str, result: Any, success: bool = True) -> None:
        """Register module execution result"""
        with self._lock:
            if success:
                self.completed_modules.add(module_name)
                self.module_results[module_name] = result
                self.failed_modules.discard(module_name)
                logging.info(f"Module completed: {module_name}")
            else:
                self.failed_modules.add(module_name)
                self.completed_modules.discard(module_name)
                logging.error(f"Module failed: {module_name}")

    def get_module_result(self, module_name: str) -> Optional[Any]:
        """Get result from completed module"""
        return self.module_results.get(module_name)

    def is_module_completed(self, module_name: str) -> bool:
        """Check if module completed successfully"""
        return module_name in self.completed_modules

    def has_dependency_failed(self, dependencies: List[str]) -> bool:
        """Check if any required dependencies failed"""
        return any(dep in self.failed_modules for dep in dependencies)

    def cache_set(self, key: str, value: Any, size_hint: int = 0) -> None:
        """Set value in shared cache"""
        with self._lock:
            if self.config.enable_caching:
                self.shared_cache[key] = value
                if size_hint > 0:
                    self.memory_manager.cache_result(key, value, size_hint)

    def cache_get(self, key: str) -> Optional[Any]:
        """Get value from shared cache"""
        if not self.config.enable_caching:
            return None
            
        # Try memory manager cache first
        cached = self.memory_manager.get_cached_result(key)
        if cached is not None:
            return cached
            
        # Fall back to shared cache
        return self.shared_cache.get(key)

    def get_workbook_access(self) -> SafeWorkbookAccess:
        """Get workbook access manager"""
        return self.workbook_access

    def should_reduce_complexity(self) -> bool:
        """Check if analysis complexity should be reduced due to resource constraints"""
        return (
            self.memory_manager.should_reduce_processing_depth() or
            self.file_metadata.file_size_mb > 100 or  # Large file
            len(self.failed_modules) > 2  # Multiple failures
        )

    def estimate_remaining_time(self, completed_modules: int, total_modules: int) -> float:
        """Estimate remaining processing time"""
        if completed_modules == 0:
            return 0.0
            
        elapsed = time.time() - self.analysis_start_time
        rate = completed_modules / elapsed
        remaining_modules = total_modules - completed_modules
        
        return remaining_modules / rate if rate > 0 else 0.0

    def get_analysis_summary(self) -> Dict[str, Any]:
        """Get current analysis state summary"""
        current_time = time.time()
        
        return {
            'file_info': {
                'name': self.file_path.name,
                'size_mb': self.file_metadata.file_size_mb,
                'path': str(self.file_path)
            },
            'progress': {
                'current_phase': self.current_phase.value,
                'completed_modules': len(self.completed_modules),
                'failed_modules': len(self.failed_modules),
                'module_names': {
                    'completed': list(self.completed_modules),
                    'failed': list(self.failed_modules)
                }
            },
            'timing': {
                'elapsed_seconds': current_time - self.analysis_start_time,
                'analysis_start': self.analysis_start_time,
                'phase_durations': {
                    phase.value: current_time - start_time 
                    for phase, start_time in self.phase_start_times.items()
                }
            },
            'resources': self.memory_manager.get_current_usage(),
            'cache_stats': {
                'items': len(self.shared_cache),
                'enabled': self.config.enable_caching
            }
        }

    def cleanup(self) -> None:
        """Cleanup resources and prepare for shutdown"""
        with self._lock:
            if self._shutdown:
                return
                
            self._shutdown = True
            
            # Close workbook
            self.workbook_access.close()
            
            # Clear caches
            self.shared_cache.clear()
            self.memory_manager.clear_cache()
            
            # Log final summary
            summary = self.get_analysis_summary()
            logging.info(f"Analysis cleanup completed. Duration: {summary['timing']['elapsed_seconds']:.1f}s")

    def __enter__(self):
        """Context manager entry"""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit with cleanup"""
        self.cleanup()
        if exc_type is not None:
            logging.error(f"Analysis context exited with error: {exc_val}")
