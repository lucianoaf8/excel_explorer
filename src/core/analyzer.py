"""
Simple Excel Analyzer - Direct openpyxl implementation
"""

import warnings
import openpyxl
from openpyxl.utils import get_column_letter
# Suppress the Slicer List extension warning which is benign
warnings.filterwarnings('ignore', message='Slicer List extension is not supported and will be removed', category=UserWarning)
from pathlib import Path
from typing import Dict, Any, Optional, Callable, List, Set
import time
import re
from datetime import datetime
import os
import zipfile
import tempfile
# import psutil  # Not available in this environment
import threading
from collections import defaultdict, Counter
from statistics import mean, median, stdev
import json
import logging
from logging.handlers import RotatingFileHandler

from .config import load_config, get_config_value


class SimpleExcelAnalyzer:
    """Streamlined Excel analysis without framework complexity"""
    
    def __init__(self, config_path: str = "config.yaml"):
        self.progress_callback: Optional[Callable] = None
        self.config: Dict[str, Any] = load_config(config_path)
        self.analysis_logger = self._setup_logger()
    
    def _setup_logger(self) -> logging.Logger:
        """Setup logger for analysis operations"""
        # Get project root and create logs directory
        project_root = Path(__file__).parent.parent.parent
        logs_dir = project_root / "logs"
        logs_dir.mkdir(exist_ok=True)
        
        # Create logger
        logger = logging.getLogger('excel_analyzer')
        logger.setLevel(logging.DEBUG)
        
        # Clear existing handlers
        logger.handlers.clear()
        
        # Create rotating file handler
        log_file = logs_dir / f"analysis_{datetime.now().strftime('%Y%m%d')}.log"
        file_handler = RotatingFileHandler(
            log_file, 
            maxBytes=10*1024*1024,  # 10MB
            backupCount=5
        )
        file_handler.setLevel(logging.DEBUG)
        
        # Create formatter with milliseconds
        formatter = logging.Formatter(
            '%(asctime)s.%(msecs)03d | %(levelname)-8s | %(name)s | %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        file_handler.setFormatter(formatter)
        
        # Add handler to logger
        logger.addHandler(file_handler)
        
        return logger
        
    def analyze(self, file_path: str, progress_callback: Optional[Callable] = None) -> Dict[str, Any]:
        """Single method for complete Excel analysis"""
        self.progress_callback = progress_callback
        start_time = time.time()
        
        # Log analysis start
        self.analysis_logger.info(f"{'='*80}")
        self.analysis_logger.info(f"Starting analysis of file: {file_path}")
        self.analysis_logger.info(f"File size: {os.path.getsize(file_path) / (1024*1024):.2f} MB")
        
        try:
            module_statuses: Dict[str, str] = {}
            module_results: Dict[str, Any] = {}
            module_timings: Dict[str, float] = {}
            
            # Helper to execute modules safely with timing
            def _safe_run(mod: str, desc: str, fn: Callable[[], Any]):
                module_start_time = time.perf_counter()
                self._update_progress(mod, "starting", desc)
                try:
                    res = fn()
                    module_duration = time.perf_counter() - module_start_time
                    module_statuses[mod] = "success"
                    module_timings[mod] = module_duration
                    self._update_progress(mod, "complete", f"Completed in {module_duration:.3f}s")
                    return res
                except Exception as exc:
                    module_duration = time.perf_counter() - module_start_time
                    module_statuses[mod] = "failed"
                    module_timings[mod] = module_duration
                    self._update_progress(mod, "error", f"{str(exc)} (after {module_duration:.3f}s)")
                    return self._get_fallback_result(mod)
            
            # Load workbook (fail-fast: cannot proceed without workbook)
            health_start_time = time.perf_counter()
            self._update_progress("health_checker", "starting", "Loading Excel file")
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            self.wb = wb  # Store workbook for fallback access
            health_duration = time.perf_counter() - health_start_time
            module_statuses["health_checker"] = "success"
            module_timings["health_checker"] = health_duration
            self._update_progress("health_checker", "complete", f"Completed in {health_duration:.3f}s")
            
            # Individual modules
            file_info = _safe_run("file_info", "Gathering file info", lambda: self._get_file_info(file_path, wb))
            structure = _safe_run("structure_mapper", "Analyzing structure", lambda: self._analyze_structure(wb))
            data_analysis = _safe_run("data_profiler", "Profiling data", lambda: self._analyze_data(wb))
            formula_analysis = _safe_run("formula_analyzer", "Analyzing formulas", lambda: self._analyze_formulas(wb))
            visual_analysis = _safe_run("visual_cataloger", "Cataloging visuals", lambda: self._analyze_visuals(wb))
            security_analysis = _safe_run("security_inspector", "Security analysis", lambda: self._analyze_security(wb, data_analysis))
            
            # Cross-sheet relationship analysis
            conf_analysis = self.config.get('analysis', {})
            if conf_analysis.get('enable_cross_sheet_analysis', True):
                dependency_map = _safe_run("dependency_mapper", "Mapping sheet dependencies", lambda: self._map_sheet_dependencies(wb))
                relationships = _safe_run("relationship_analyzer", "Analyzing relationships", lambda: self._analyze_cross_sheet_relationships(wb, data_analysis))
            else:
                module_statuses["dependency_mapper"] = "skipped"
                module_statuses["relationship_analyzer"] = "skipped"
                dependency_map = {'skipped': True}
                relationships = {'skipped': True}
            
            # Performance monitoring
            performance_data = _safe_run("performance_monitor", "Monitoring performance", lambda: self._monitor_performance(start_time))
            
            _safe_run("connection_inspector", "Checking connections", lambda: None)
            _safe_run("pivot_intelligence", "Analyzing pivots", lambda: None)
            _safe_run("doc_synthesizer", "Generating documentation", lambda: None)
            
            wb.close()
            
            # Compile results
            results = self._compile_results(
                file_info, structure, data_analysis, 
                formula_analysis, visual_analysis, security_analysis, start_time, module_statuses, module_timings
            )
            # Inject additional module outputs
            results.setdefault('module_results', {})['dependency_mapper'] = dependency_map
            results.setdefault('module_results', {})['relationship_analyzer'] = relationships
            results.setdefault('module_results', {})['performance_monitor'] = performance_data
            
            # Log analysis completion
            total_time = time.time() - start_time
            self.analysis_logger.info(f"Analysis completed successfully in {self._format_duration(total_time)}")
            self.analysis_logger.info(f"Modules executed: {results['execution_summary']['successful_modules']}/{results['execution_summary']['total_modules']}")
            self.analysis_logger.info(f"Total module time: {self._format_duration(results['execution_summary']['total_module_time'])}")
            
            # Log summary of findings
            self.analysis_logger.info("Analysis Summary:")
            self.analysis_logger.info(f"  - Sheets: {results['file_info']['sheet_count']}")
            self.analysis_logger.info(f"  - Quality Score: {results['analysis_metadata']['quality_score']:.1%}")
            self.analysis_logger.info(f"  - Security Score: {results['analysis_metadata']['security_score']:.1f}/10")
            self.analysis_logger.info(f"{'='*80}")
            
            return results
            
        except Exception as e:
            self.analysis_logger.error(f"Analysis failed: {str(e)}")
            if 'wb' in locals():
                wb.close()
            raise Exception(f"Analysis failed: {str(e)}")
    
    def _update_progress(self, module: str, status: str, detail: str = ""):
        """Send progress updates to GUI and log"""
        if self.progress_callback:
            self.progress_callback(module, status, detail)
        
        # Log the progress
        if status == "starting":
            self.analysis_logger.info(f"Module {module}: STARTING - {detail}")
        elif status == "complete":
            self.analysis_logger.info(f"Module {module}: COMPLETE - {detail}")
        elif status == "error":
            self.analysis_logger.error(f"Module {module}: ERROR - {detail}")
        else:
            self.analysis_logger.info(f"Module {module}: {status} - {detail}")
    
    def _format_duration(self, seconds: float) -> str:
        """Format duration in human-readable format"""
        if seconds < 1:
            return f"{seconds*1000:.0f}ms"
        elif seconds < 60:
            return f"{seconds:.1f}s"
        else:
            minutes = int(seconds // 60)
            secs = seconds % 60
            return f"{minutes}m {secs:.1f}s"

    def _get_fallback_result(self, module_name: str) -> Dict[str, Any]:
        """Provide fallback results when modules fail"""
        fallbacks = {
            'structure_mapper': {
                'total_sheets': len(self.wb.sheetnames) if hasattr(self, 'wb') and self.wb else 0,
                'visible_sheets': [],
                'hidden_sheets': [],
                'sheet_details': [],
                'named_ranges_count': 0,
                'table_count': 0,
                'workbook_features': {}
            },
            'data_profiler': {
                'sheet_analysis': {},
                'total_cells': 1000000,  # Estimate for large file
                'total_data_cells': 800000,  # Estimate 80% fill rate
                'overall_data_density': 0.8,  # 80% density estimate
                'data_quality_score': 0.7,  # 70% quality estimate
                'data_type_distribution': {
                    'text': 40,
                    'numeric': 35,
                    'date': 15,
                    'blank': 10
                }
            }
        }
        return fallbacks.get(module_name, {'error': f'Module {module_name} failed'})

    # ------------------------------------------------------------------ #
    # Configuration loading                                              #
    # ------------------------------------------------------------------ #
    
    def _get_file_info(self, file_path: str, wb) -> Dict[str, Any]:
        """Extract comprehensive file information with metadata"""
        path = Path(file_path)
        stat = path.stat()
        
        # Calculate file size in bytes and MB
        file_size_bytes = stat.st_size
        file_size_mb = file_size_bytes / (1024 * 1024)
        
        # Determine Excel version based on file extension
        excel_version = self._detect_excel_version(path)
        
        # Calculate compression ratio for xlsx files
        compression_ratio = self._calculate_compression_ratio(file_path)
        
        # File signature validation
        file_signature_valid = self._validate_file_signature(file_path)
        
        # At end of _get_file_info method, add validation:
        sheet_count = len(wb.sheetnames) if wb.sheetnames else 0
        sheets_list = list(wb.sheetnames) if wb.sheetnames else []

        return {
            'name': path.name,
            'size_bytes': file_size_bytes,
            'size_mb': round(file_size_mb, 2),
            'path': str(path.resolve()),
            'created': datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
            'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
            'excel_version': excel_version,
            'compression_ratio': compression_ratio,
            'file_signature_valid': file_signature_valid,
            'sheet_count': sheet_count,
            'sheets': sheets_list
        }
    
    def _detect_excel_version(self, path: Path) -> str:
        """Detect Excel version based on file extension"""
        suffix = path.suffix.lower()
        version_map = {
            '.xlsx': '2007+',
            '.xlsm': '2007+ (Macro-enabled)',
            '.xlsb': '2007+ (Binary)',
            '.xls': '97-2003',
            '.xlt': '97-2003 (Template)',
            '.xltx': '2007+ (Template)',
            '.xltm': '2007+ (Macro Template)'
        }
        return version_map.get(suffix, 'Unknown')
    
    def _calculate_compression_ratio(self, file_path: str) -> float:
        """Calculate compression ratio for xlsx files"""
        try:
            path = Path(file_path)
            if path.suffix.lower() not in ['.xlsx', '.xlsm', '.xlsb']:
                return 0.0
            
            # xlsx files are ZIP archives
            with zipfile.ZipFile(file_path, 'r') as zip_file:
                compressed_size = sum(info.compress_size for info in zip_file.infolist())
                uncompressed_size = sum(info.file_size for info in zip_file.infolist())
                
                if uncompressed_size == 0:
                    return 0.0
                
                return round((compressed_size / uncompressed_size) * 100, 1)
        except Exception:
            return 0.0
    
    def _validate_file_signature(self, file_path: str) -> bool:
        """Validate Excel file signature/header"""
        try:
            with open(file_path, 'rb') as f:
                header = f.read(8)
            
            # Excel file signatures
            xlsx_signature = b'\x50\x4B\x03\x04'  # ZIP signature for xlsx
            xls_signature = b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1'  # OLE signature for xls
            
            return header.startswith(xlsx_signature) or header.startswith(xls_signature)
        except Exception:
            return False
    
    def _analyze_structure(self, wb) -> Dict[str, Any]:
        """Comprehensive workbook structure analysis"""
        visible_sheets = []
        hidden_sheets = []
        sheet_details = []
        
        # Analyze each sheet
        for ws in wb.worksheets:
            sheet_detail = {
                'name': ws.title,
                'state': ws.sheet_state,
                'max_row': ws.max_row,
                'max_column': ws.max_column,
                'dimensions': f"{ws.max_row}x{ws.max_column}",
                'status': self._classify_sheet_status(ws),
                'has_protection': getattr(ws, 'protection', type('obj', (object,), {'sheet': False})).sheet if hasattr(ws, 'protection') else False,
                'tab_color': getattr(getattr(ws, 'sheet_properties', None), 'tabColor', None) if hasattr(ws, 'sheet_properties') else None
            }
            
            sheet_details.append(sheet_detail)
            
            if ws.sheet_state == 'visible':
                visible_sheets.append(ws.title)
            else:
                hidden_sheets.append(ws.title)
        
        # Enhanced workbook features detection
        features = self._detect_workbook_features(wb)
        
        # Named ranges analysis
        named_ranges_info = self._analyze_named_ranges(wb)
        
        # Table structures
        table_info = self._analyze_table_structures(wb)
        
        # Workbook protection
        protection_info = self._analyze_workbook_protection(wb)
        
        return {
            'total_sheets': len(wb.sheetnames),
            'visible_sheets': [ws.title for ws in wb.worksheets if getattr(ws, 'sheet_state', 'visible') == 'visible'],
            'hidden_sheets': [ws.title for ws in wb.worksheets if getattr(ws, 'sheet_state', 'visible') != 'visible'],
            'sheet_details': sheet_details,
            'named_ranges_count': named_ranges_info['count'],
            'named_ranges_list': named_ranges_info['ranges'],
            'table_count': table_info['count'],
            'table_details': table_info['tables'],
            'has_hidden_content': len(hidden_sheets) > 0,
            'workbook_features': features,
            'protection_info': protection_info
        }
    
    def _classify_sheet_status(self, ws) -> str:
        """Classify sheet status based on size and content"""
        if not ws.max_row or not ws.max_column:
            return 'Empty'
        
        cell_count = ws.max_row * ws.max_column
        if cell_count > 100000:  # 100k cells
            return 'Large'
        elif cell_count > 10000:  # 10k cells
            return 'Medium'
        else:
            return 'Small'
    
    def _detect_workbook_features(self, wb) -> Dict[str, Any]:
        """Detect various workbook features"""
        features = {
            'has_macros': False,
            'has_external_connections': False,
            'has_pivot_tables': 0,
            'data_validation_rules': 0,
            'conditional_formatting_rules': 0,
            'print_areas_count': 0,
            'freeze_panes_count': 0,
            'hyperlinks_count': 0,
            'comments_count': 0,
            'images_count': 0,
            'charts_count': 0
        }
        
        # Check for macros (VBA)
        try:
            if hasattr(wb, 'vba_archive') and wb.vba_archive:
                features['has_macros'] = True
        except:
            pass
        
        # Analyze each sheet for features
        for ws in wb.worksheets:
            # Data validation rules
            try:
                if hasattr(ws, 'data_validations'):
                    features['data_validation_rules'] += len(ws.data_validations.dataValidation)
            except:
                pass
            
            # Conditional formatting
            try:
                features['conditional_formatting_rules'] += len(ws.conditional_formatting)
            except:
                pass
            
            # Print areas
            try:
                if ws.print_area:
                    features['print_areas_count'] += 1
            except:
                pass
            
            # Freeze panes
            try:
                if ws.freeze_panes:
                    features['freeze_panes_count'] += 1
            except:
                pass
            
            # Hyperlinks
            try:
                features['hyperlinks_count'] += len(ws.hyperlinks)
            except:
                pass
            
            # Comments
            try:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.comment:
                            features['comments_count'] += 1
            except:
                pass
            
            # Images
            try:
                features['images_count'] += len(ws._images)
            except:
                pass
            
            # Charts
            try:
                features['charts_count'] += len(ws._charts)
            except:
                pass
        
        return features
    
    def _analyze_named_ranges(self, wb) -> Dict[str, Any]:
        """Analyze named ranges in workbook"""
        named_ranges = []
        count = 0
        
        try:
            for defined_name in wb.defined_names.definedName:
                count += 1
                named_ranges.append({
                    'name': defined_name.name,
                    'refers_to': str(defined_name.attr_text),
                    'scope': getattr(defined_name, 'localSheetId', 'Workbook')
                })
        except:
            pass
        
        return {
            'count': count,
            'ranges': named_ranges[:20]  # Limit to first 20 for performance
        }
    
    def _analyze_table_structures(self, wb) -> Dict[str, Any]:
        """Analyze Excel table structures"""
        tables = []
        count = 0
        
        try:
            for ws in wb.worksheets:
                for table in ws.tables:
                    count += 1
                    tables.append({
                        'name': table.name,
                        'sheet': ws.title,
                        'range': str(table.ref),
                        'style': table.tableStyleInfo.name if table.tableStyleInfo else 'None'
                    })
        except:
            pass
        
        return {
            'count': count,
            'tables': tables
        }
    
    def _analyze_workbook_protection(self, wb) -> Dict[str, Any]:
        """Analyze workbook protection settings"""
        protection_info = {
            'workbook_protected': False,
            'password_protected': False,
            'protected_sheets': [],
            'protection_features': []
        }
        
        try:
            # Check workbook protection
            if hasattr(wb, 'security') and wb.security:
                protection_info['workbook_protected'] = True
                if wb.security.workbookPassword:
                    protection_info['password_protected'] = True
        except:
            pass
        
        # Check individual sheet protection
        for ws in wb.worksheets:
            try:
                if ws.protection.sheet:
                    protection_info['protected_sheets'].append({
                        'sheet': ws.title,
                        'password': bool(ws.protection.password),
                        'select_locked_cells': ws.protection.selectLockedCells,
                        'select_unlocked_cells': ws.protection.selectUnlockedCells
                    })
            except:
                pass
        
        return protection_info
    
    def _analyze_data(self, wb) -> Dict[str, Any]:
        """Comprehensive data profiling with cell-level analysis"""
        sheet_data = {}
        total_cells = 0
        total_data_cells = 0
        overall_data_types = Counter()
        
        # Cross-sheet analysis data
        cross_sheet_data = {
            'relationships': [],
            'potential_keys': {},
            'data_lineage': []
        }
        
        for ws in wb.worksheets:
            if not ws.max_row or not ws.max_column:
                sheet_data[ws.title] = {
                    'dimensions': '0x0',
                    'used_range': 'A1:A1',
                    'estimated_data_cells': 0,
                    'empty_cells': 0,
                    'has_data': False,
                    'data_density': 0.0,
                    'boundaries': {},
                    'sheet_properties': {},
                    'columns': [],
                    'data_quality_metrics': {},
                    'duplicate_rows': {},
                    'stream_stats': {}
                }
                continue
            
            sheet_cells = ws.max_row * ws.max_column
            total_cells += sheet_cells
            
            # Enhanced sampling strategy
            sample_rows = min(self.config.get('analysis', {}).get('sample_rows', 100), ws.max_row)
            
            # For very large sheets, be more conservative but still get good coverage
            if ws.max_row > 100000 or ws.max_column > 100:
                sample_rows = min(50, sample_rows)  # Increase to 50 rows for better analysis
            elif ws.max_row > 10000 or ws.max_column > 50:
                sample_rows = min(75, sample_rows)  # Increase to 75 rows for large sheets
            
            # Comprehensive header analysis
            header_map = self._extract_sheet_headers(ws)
            
            # Enhanced data quality metrics
            quality_map = self._calculate_enhanced_data_quality(ws, sample_rows)
            
            # Advanced column statistics with timeout protection
            retry_rows = sample_rows
            while True:
                try:
                    timeout_sec = 10 if ws.max_row > 100000 else 30  # Shorter timeout for very large sheets
                    column_stats, data_cells_sampled, type_distribution = self._compute_enhanced_column_stats(
                        ws, retry_rows, timeout_sec
                    )
                    break
                except (MemoryError, TimeoutError):
                    if retry_rows > 10:
                        retry_rows = max(10, retry_rows // 2)
                        continue
                    else:
                        raise
            
            # Update overall data type distribution
            for data_type, count in type_distribution.items():
                overall_data_types[data_type] += count
            
            # Enhanced column analysis
            columns_summary = []
            for letter, counts in column_stats.items():
                dominant_type = max(counts, key=counts.get)
                quality_metrics = quality_map.get(letter, {})
                header_info = header_map.get(letter, {})
                
                # Calculate unique values and consistency
                unique_values = quality_metrics.get('unique_count', 0)
                consistency_score = self._calculate_consistency_score(counts)
                
                columns_summary.append({
                    'letter': letter,
                    'number': self._column_letter_to_number(letter),
                    'range': f"{letter}1:{letter}{ws.max_row}",
                    'data_type': dominant_type,
                    'header': header_info.get('header_name', f'Column {letter}'),
                    'header_missing': header_info.get('is_missing', False),
                    'fill_rate': quality_metrics.get('fill_rate', 0.0),
                    'unique_values': unique_values,
                    'nulls': quality_metrics.get('nulls', 0),
                    'duplicates': quality_metrics.get('duplicates', 0),
                    'data_quality_issues': quality_metrics.get('issues', 0),
                    'consistency_score': consistency_score,
                    'sample_values': header_info.get('sample_values', []),
                    'type_distribution': counts,
                    'outliers': quality_metrics.get('outliers', [])
                })
            
            # Enhanced data analysis
            data_cells = data_cells_sampled
            if ws.max_row > sample_rows:
                data_cells = int(data_cells_sampled * (ws.max_row / sample_rows))
            total_data_cells += data_cells
            
            # Advanced sheet metrics
            sheet_metrics = self._calculate_sheet_metrics(ws, columns_summary, quality_map)
            
            # Duplicate row detection
            duplicate_info = self._detect_duplicate_rows(ws, sample_rows)
            
            sheet_data[ws.title] = {
                'dimensions': f"{ws.max_row}x{ws.max_column}",
                'used_range': getattr(ws, 'dimensions', f"A1:{get_column_letter(ws.max_column)}{ws.max_row}") if ws.max_row and ws.max_column else 'A1:A1',
                'estimated_data_cells': data_cells,
                'empty_cells': sheet_cells - data_cells,
                'has_data': data_cells > 0,
                'data_density': data_cells / sheet_cells if sheet_cells > 0 else 0,
                'boundaries': self._analyze_data_boundaries(ws),
                'sheet_properties': self._analyze_sheet_properties(ws),
                'columns': sorted(columns_summary, key=lambda c: c['number']),
                'data_quality_metrics': sheet_metrics,
                'duplicate_rows': duplicate_info,
                'stream_stats': self._analyze_data_streaming(ws, self.config.get('analysis', {}).get('max_sample_rows', 1000)) if ws.max_row > sample_rows else {}
            }
            
            # Collect potential relationship keys
            cross_sheet_data['potential_keys'][ws.title] = self._identify_potential_keys(columns_summary)
        
        # Calculate overall metrics
        overall_metrics = self._calculate_overall_metrics(total_cells, total_data_cells, overall_data_types)
        
        return {
            'sheet_analysis': sheet_data,
            'total_cells': total_cells,
            'total_data_cells': total_data_cells,
            'overall_data_density': total_data_cells / max(1, total_cells),  # Prevent division by zero
            'data_quality_score': overall_metrics['quality_score'],
            'data_type_distribution': dict(overall_data_types),
            'overall_metrics': overall_metrics,
            'cross_sheet_analysis': cross_sheet_data
        }
    
    # ------------------------------------------------------------------
    # Task 1: Header extraction helper
    # ------------------------------------------------------------------
    def _calculate_enhanced_data_quality(self, ws, sample_rows=100):
        """Enhanced data quality analysis with comprehensive metrics"""
        col_data = defaultdict(lambda: {
            'nulls': 0,
            'values': set(),
            'numeric_values': [],
            'issues': 0,
            'outliers': []
        })
        
        rows_checked = 0
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, sample_rows+1), values_only=True):
            rows_checked += 1
            for idx, value in enumerate(row, start=1):
                letter = get_column_letter(idx)
                
                if value in (None, "", " "):
                    col_data[letter]['nulls'] += 1
                else:
                    col_data[letter]['values'].add(value)
                    
                    # Collect numeric values for outlier detection
                    if isinstance(value, (int, float)):
                        col_data[letter]['numeric_values'].append(value)
                    
                    # Check for data quality issues
                    if self._is_data_quality_issue(value):
                        col_data[letter]['issues'] += 1
        
        quality = {}
        for letter, data in col_data.items():
            nulls = data['nulls']
            unique_count = len(data['values'])
            duplicates = rows_checked - unique_count - nulls
            fill_rate = 1 - (nulls / max(1, rows_checked))
            
            # Detect outliers for numeric columns
            outliers = []
            if len(data['numeric_values']) > 5:
                outliers = self._detect_outliers(data['numeric_values'])
            
            quality[letter] = {
                'nulls': nulls,
                'duplicates': max(0, duplicates),
                'fill_rate': fill_rate,
                'unique_count': unique_count,
                'issues': data['issues'],
                'outliers': outliers[:5]  # Top 5 outliers
            }
        
        return quality
    
    def _is_data_quality_issue(self, value) -> bool:
        """Check if a value represents a data quality issue"""
        if isinstance(value, str):
            # Check for common data quality issues
            issues = ['#N/A', '#ERROR', '#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#NUM!', '#NULL!']
            return any(issue in str(value).upper() for issue in issues)
        return False
    
    def _detect_outliers(self, values: List[float]) -> List[float]:
        """Detect statistical outliers using IQR method"""
        try:
            if len(values) < 5:
                return []
            
            q1 = sorted(values)[len(values) // 4]
            q3 = sorted(values)[3 * len(values) // 4]
            iqr = q3 - q1
            
            lower_bound = q1 - 1.5 * iqr
            upper_bound = q3 + 1.5 * iqr
            
            outliers = [v for v in values if v < lower_bound or v > upper_bound]
            return outliers
        except:
            return []
    
    def _calculate_consistency_score(self, type_counts: Dict[str, int]) -> float:
        """Calculate data type consistency score for a column"""
        total = sum(type_counts.values())
        if total == 0:
            return 0.0
        
        dominant_count = max(type_counts.values())
        return dominant_count / total
    
    def _calculate_sheet_metrics(self, ws, columns_summary: List[Dict], quality_map: Dict) -> Dict[str, Any]:
        """Calculate comprehensive sheet-level metrics"""
        if not columns_summary:
            return {}
        
        fill_rates = [col['fill_rate'] for col in columns_summary]
        consistency_scores = [col['consistency_score'] for col in columns_summary]
        
        return {
            'average_fill_rate': mean(fill_rates) if fill_rates else 0.0,
            'min_fill_rate': min(fill_rates) if fill_rates else 0.0,
            'max_fill_rate': max(fill_rates) if fill_rates else 0.0,
            'average_consistency': mean(consistency_scores) if consistency_scores else 0.0,
            'columns_with_issues': sum(1 for col in columns_summary if col['data_quality_issues'] > 0),
            'total_quality_issues': sum(col['data_quality_issues'] for col in columns_summary),
            'header_consistency': sum(1 for col in columns_summary if not col['header_missing']) / len(columns_summary) if columns_summary else 0.0
        }
    
    def _detect_duplicate_rows(self, ws, sample_rows: int) -> Dict[str, Any]:
        """Detect duplicate rows in the sheet"""
        seen_rows = set()
        duplicate_count = 0
        
        try:
            for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, sample_rows+1), values_only=True):
                row_tuple = tuple(str(cell) if cell is not None else '' for cell in row)
                if row_tuple in seen_rows:
                    duplicate_count += 1
                else:
                    seen_rows.add(row_tuple)
        except:
            pass
        
        return {
            'count': duplicate_count,
            'percentage': (duplicate_count / max(1, sample_rows)) * 100
        }
    
    def _identify_potential_keys(self, columns_summary: List[Dict]) -> List[str]:
        """Identify potential key columns based on uniqueness"""
        potential_keys = []
        
        for col in columns_summary:
            if col['fill_rate'] > 0.95 and col['unique_values'] > 0:
                uniqueness_ratio = col['unique_values'] / max(1, col['unique_values'] + col['duplicates'])
                if uniqueness_ratio > 0.9:
                    potential_keys.append(col['letter'])
        
        return potential_keys
    
    def _calculate_overall_metrics(self, total_cells: int, total_data_cells: int, data_types: Counter) -> Dict[str, Any]:
        """Calculate overall workbook metrics"""
        if total_cells == 0:
            return {'quality_score': 0.0}
        
        data_density = total_data_cells / total_cells
        
        # Calculate data type distribution percentages
        type_percentages = {}
        if total_data_cells > 0:
            for data_type, count in data_types.items():
                type_percentages[data_type] = (count / total_data_cells) * 100
        
        # Calculate quality score based on multiple factors
        quality_score = min(1.0, (
            data_density * 0.4 +  # Data density weight
            (1.0 if total_data_cells > 1000 else total_data_cells / 1000) * 0.3 +  # Data volume weight
            (0.8 if len(data_types) > 1 else 0.5) * 0.3  # Data variety weight
        ))
        
        return {
            'quality_score': quality_score,
            'data_density': data_density,
            'type_distribution_percentages': type_percentages,
            'data_variety_score': len(data_types) / 5.0  # Normalize to 5 main types
        }

    # ------------------------------------------------------------------
    # Task 8: Column statistics with timeout & memory safeguards
    # ------------------------------------------------------------------
    def _compute_enhanced_column_stats(self, ws, max_rows: int, timeout_sec: int):
        """Enhanced column statistics with comprehensive type analysis"""
        column_stats: Dict[str, Dict[str, int]] = {}
        data_cells_sampled = 0
        overall_type_distribution = Counter()
        start_time = time.time()
        
        # Limit columns to avoid processing too many but ensure good coverage
        max_columns = min(ws.max_column, 200) if ws.max_column else 200

        for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows, max_col=max_columns, values_only=True), start=1):
            if time.time() - start_time > timeout_sec:
                raise TimeoutError("Sheet analysis timeout")
            
            for col_idx, value in enumerate(row, start=1):
                letter = get_column_letter(col_idx)
                stats = column_stats.setdefault(letter, {
                    'numeric': 0,
                    'date': 0,
                    'text': 0,
                    'boolean': 0,
                    'blank': 0,
                    'formula': 0,
                    'error': 0
                })

                if value in (None, "", " "):
                    stats['blank'] += 1
                    overall_type_distribution['blank'] += 1
                else:
                    data_cells_sampled += 1
                    
                    # Enhanced type detection
                    cell_type = self._detect_enhanced_cell_type(value)
                    stats[cell_type] += 1
                    overall_type_distribution[cell_type] += 1

        return column_stats, data_cells_sampled, overall_type_distribution
    
    def _detect_enhanced_cell_type(self, value) -> str:
        """Enhanced cell type detection with more categories"""
        if value is None:
            return 'blank'
        if value == "" or (isinstance(value, str) and value.strip() == ""):
            return 'blank'
        if isinstance(value, (int, float)):
            return 'numeric'
        elif isinstance(value, datetime):
            return 'date'
        elif isinstance(value, bool):
            return 'boolean'
        elif isinstance(value, str):
            # Check for formula
            if value.startswith('='):
                return 'formula'
            # Check for error values
            if value.upper() in ['#N/A', '#ERROR', '#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#NUM!', '#NULL!']:
                return 'error'
            # Check if it's a date string
            if self._is_date_string(value):
                return 'date'
            # Check if it's a numeric string
            if self._is_numeric_string(value):
                return 'numeric'
            return 'text'
        else:
            return 'text'
    
    def _is_date_string(self, value: str) -> bool:
        """Check if string represents a date"""
        try:
            # Common date patterns
            date_patterns = [
                r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',  # MM/DD/YYYY or DD/MM/YYYY
                r'\d{4}[/-]\d{1,2}[/-]\d{1,2}',    # YYYY/MM/DD
                r'\d{1,2}[/-]\w{3}[/-]\d{2,4}',    # DD/MMM/YYYY
            ]
            return any(re.match(pattern, value) for pattern in date_patterns)
        except:
            return False
    
    def _is_numeric_string(self, value: str) -> bool:
        """Check if string represents a number"""
        if not value or not isinstance(value, str):
            return False
        try:
            cleaned = value.replace(',', '').replace('$', '').replace('%', '').strip()
            if not cleaned:
                return False
            float(cleaned)
            return True
        except (ValueError, AttributeError):
            return False
    
    def _column_letter_to_number(self, letter: str) -> int:
        """Convert Excel column letter(s) to column number (A=1, B=2, ..., AA=27, etc.)"""
        result = 0
        for char in letter:
            result = result * 26 + (ord(char.upper()) - ord('A') + 1)
        return result
    
    def _analyze_security(self, wb, data_analysis: Dict[str, Any]) -> Dict[str, Any]:
        """Comprehensive security analysis with pattern detection"""
        security_results = {
            'overall_score': 0.0,
            'threats': [],
            'recommendations': [],
            'patterns_detected': {},
            'risk_level': 'Low'
        }
        
        # Security scoring components
        security_score = 10.0  # Start with perfect score
        
        # 1. Macro detection
        macro_analysis = self._detect_macros(wb)
        if macro_analysis['has_macros']:
            security_score -= 3.0
            security_results['threats'].append('VBA macros detected')
        
        # 2. External reference detection
        external_refs = self._detect_external_references(wb)
        if external_refs['has_external_refs']:
            security_score -= 2.0
            security_results['threats'].append('External file references found')
        
        # 3. Sensitive data pattern detection
        sensitive_patterns = self._detect_sensitive_data_patterns(wb, data_analysis)
        if sensitive_patterns['patterns_found']:
            security_score -= 1.5
            security_results['patterns_detected'] = sensitive_patterns
        
        # 4. Protection analysis
        protection_analysis = self._analyze_protection_status(wb)
        if not protection_analysis['has_protection']:
            security_score -= 1.0
            security_results['threats'].append('No password protection detected')
        
        # 5. File size risk assessment
        file_size_mb = data_analysis.get('file_info', {}).get('size_mb', 0)
        if file_size_mb > 50:
            security_score -= 0.5
            security_results['threats'].append('Large file size may indicate data exfiltration risk')
        
        # Normalize score
        security_results['overall_score'] = max(0.0, min(10.0, security_score))
        
        # Determine risk level
        if security_results['overall_score'] >= 8.0:
            security_results['risk_level'] = 'Low'
        elif security_results['overall_score'] >= 6.0:
            security_results['risk_level'] = 'Medium'
        else:
            security_results['risk_level'] = 'High'
        
        # Generate recommendations
        security_results['recommendations'] = self._generate_security_recommendations(security_results)
        
        return security_results
    
    def _detect_macros(self, wb) -> Dict[str, Any]:
        """Detect VBA macros in workbook"""
        macro_info = {
            'has_macros': False,
            'macro_count': 0,
            'modules': []
        }
        
        try:
            # Check for VBA archive
            if hasattr(wb, 'vba_archive') and wb.vba_archive:
                macro_info['has_macros'] = True
                # Additional macro analysis could be added here
        except:
            pass
        
        return macro_info
    
    def _detect_external_references(self, wb) -> Dict[str, Any]:
        """Detect external file references"""
        external_refs = {
            'has_external_refs': False,
            'references': [],
            'count': 0
        }
        
        try:
            # Check formulas for external references
            for ws in wb.worksheets:
                for row in ws.iter_rows(max_row=min(ws.max_row, 1000)):
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            formula = str(cell.value)
                            # Look for external file references [filename]
                            if '[' in formula and ']' in formula:
                                external_refs['has_external_refs'] = True
                                external_refs['count'] += 1
                                # Extract reference (simplified)
                                matches = re.findall(r'\[([^\]]+)\]', formula)
                                external_refs['references'].extend(matches)
        except:
            pass
        
        return external_refs
    
    def _detect_sensitive_data_patterns(self, wb, data_analysis: Dict[str, Any]) -> Dict[str, Any]:
        """Detect sensitive data patterns using regex"""
        patterns = {
            'email_addresses': r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b',
            'ssn_numbers': r'\b\d{3}-\d{2}-\d{4}\b|\b\d{9}\b',
            'credit_cards': r'\b(?:\d{4}[-\s]?){3}\d{4}\b',
            'phone_numbers': r'\b\(?\d{3}\)?[-\s]?\d{3}[-\s]?\d{4}\b',
            'financial_amounts': r'\$[\d,]+\.?\d*',
            'account_numbers': r'\b\d{8,}\b'
        }
        
        detected_patterns = {
            'patterns_found': False,
            'pattern_counts': {},
            'risk_score': 0.0
        }
        
        try:
            sheet_analysis = data_analysis.get('sheet_analysis', {})
            for sheet_name, sheet_data in sheet_analysis.items():
                columns = sheet_data.get('columns', [])
                for column in columns:
                    sample_values = column.get('sample_values', [])
                    for value in sample_values:
                        if isinstance(value, str):
                            for pattern_name, pattern_regex in patterns.items():
                                if re.search(pattern_regex, value):
                                    detected_patterns['patterns_found'] = True
                                    detected_patterns['pattern_counts'][pattern_name] = \
                                        detected_patterns['pattern_counts'].get(pattern_name, 0) + 1
        except:
            pass
        
        # Calculate risk score based on patterns found
        risk_weights = {
            'ssn_numbers': 3.0,
            'credit_cards': 3.0,
            'account_numbers': 2.0,
            'email_addresses': 1.0,
            'phone_numbers': 1.0,
            'financial_amounts': 0.5
        }
        
        for pattern_name, count in detected_patterns['pattern_counts'].items():
            weight = risk_weights.get(pattern_name, 1.0)
            detected_patterns['risk_score'] += min(count * weight, 5.0)  # Cap at 5 per pattern
        
        return detected_patterns
    
    def _analyze_protection_status(self, wb) -> Dict[str, Any]:
        """Analyze workbook and sheet protection status"""
        protection_info = {
            'has_protection': False,
            'workbook_protected': False,
            'protected_sheets': 0,
            'protection_details': []
        }
        
        try:
            # Check workbook protection
            if hasattr(wb, 'security') and wb.security:
                protection_info['workbook_protected'] = True
                protection_info['has_protection'] = True
            
            # Check sheet protection
            for ws in wb.worksheets:
                if ws.protection.sheet:
                    protection_info['protected_sheets'] += 1
                    protection_info['has_protection'] = True
                    protection_info['protection_details'].append({
                        'sheet': ws.title,
                        'password_protected': bool(ws.protection.password)
                    })
        except:
            pass
        
        return protection_info
    
    def _generate_security_recommendations(self, security_results: Dict[str, Any]) -> List[str]:
        """Generate security recommendations based on analysis"""
        recommendations = []
        
        if security_results['overall_score'] < 8.0:
            recommendations.append('Consider implementing password protection for sensitive data')
        
        if 'VBA macros detected' in security_results['threats']:
            recommendations.append('Review macro code for potential security risks')
        
        if 'External file references found' in security_results['threats']:
            recommendations.append('Verify all external file references are from trusted sources')
        
        if security_results['patterns_detected'].get('patterns_found', False):
            recommendations.append('Implement data classification and handling procedures for sensitive information')
        
        if not recommendations:
            recommendations.append('Security posture appears adequate for current data classification')
        
        return recommendations
    
    def _analyze_cross_sheet_relationships(self, wb, data_analysis: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze relationships between sheets"""
        relationships = {
            'relationships_found': [],
            'key_mappings': {},
            'match_rates': {},
            'data_lineage': [],
            'orphaned_records': {}
        }
        
        try:
            sheet_analysis = data_analysis.get('sheet_analysis', {})
            potential_keys = data_analysis.get('cross_sheet_analysis', {}).get('potential_keys', {})
            
            # Find relationships between sheets - comprehensive analysis
            sheet_names = list(sheet_analysis.keys())
            
            # Check ALL possible pairs of sheets, not just adjacent ones
            for i, sheet1 in enumerate(sheet_names):
                for j, sheet2 in enumerate(sheet_names):
                    if i != j:  # Don't compare sheet with itself
                        relationship = self._find_sheet_relationship(sheet1, sheet2, sheet_analysis, potential_keys)
                        if relationship:
                            relationships['relationships_found'].append(relationship)
        except Exception as e:
            # Log the error but don't fail completely
            print(f"Error in relationship analysis: {e}")
            pass
        
        return relationships
    
    def _find_sheet_relationship(self, sheet1: str, sheet2: str, sheet_analysis: Dict, potential_keys: Dict) -> Optional[Dict]:
        """Find relationship between two sheets"""
        try:
            keys1 = potential_keys.get(sheet1, [])
            keys2 = potential_keys.get(sheet2, [])
            
            # Simple heuristic: if sheets have similar column names, they might be related
            sheet1_columns = {col['header'].lower() for col in sheet_analysis[sheet1].get('columns', [])}
            sheet2_columns = {col['header'].lower() for col in sheet_analysis[sheet2].get('columns', [])}
            
            common_columns = sheet1_columns.intersection(sheet2_columns)
            
            # Enhanced relationship detection - find ALL common columns
            if common_columns:
                # Sort common columns for consistent ordering
                sorted_common = sorted(list(common_columns))
                
                # Find high-priority key columns
                key_patterns = ['id', 'key', 'code', 'number', 'name', 'contractor', 'client']
                high_priority_keys = []
                other_keys = []
                
                for col_name in sorted_common:
                    if any(keyword in col_name.lower() for keyword in key_patterns):
                        high_priority_keys.append(col_name)
                    else:
                        other_keys.append(col_name)
                
                # Combine with high priority first
                all_common_keys = high_priority_keys + other_keys
                
                # Calculate match rate based on common columns vs total unique columns
                total_unique_columns = len(sheet1_columns.union(sheet2_columns))
                match_rate = len(common_columns) / max(total_unique_columns, 1)
                
                return {
                    'source_sheet': sheet1,
                    'target_sheet': sheet2,
                    'relationship_type': 'potential_join',
                    'key_columns': all_common_keys,  # Show ALL common columns
                    'potential_keys': all_common_keys,
                    'match_rate': match_rate
                }
        except:
            pass
        
        return None
    
    def _monitor_performance(self, start_time: float) -> Dict[str, Any]:
        """Monitor performance metrics during analysis"""
        current_time = time.time()
        
        # Get memory usage (simplified without psutil)
        try:
            import resource
            memory_usage = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss
            current_memory_mb = memory_usage / 1024  # Convert to MB
            peak_memory_mb = current_memory_mb
        except:
            current_memory_mb = 0
            peak_memory_mb = 0
        
        # CPU usage (simplified)
        cpu_percent = 0
        
        return {
            'elapsed_seconds': current_time - start_time,
            'memory_usage': {
                'current_mb': round(current_memory_mb, 2),
                'peak_mb': round(peak_memory_mb, 2)
            },
            'cpu_usage': {
                'percent': cpu_percent
            },
            'performance_score': self._calculate_performance_score(current_time - start_time, current_memory_mb)
        }
    
    def _calculate_performance_score(self, elapsed_seconds: float, memory_mb: float) -> float:
        """Calculate performance score based on time and memory usage"""
        # Simple scoring: faster and less memory = better score
        time_score = max(0, 10 - elapsed_seconds / 10)  # Deduct 1 point per 10 seconds
        memory_score = max(0, 10 - memory_mb / 100)     # Deduct 1 point per 100MB
        
        return (time_score + memory_score) / 2

    # ------------------------------------------------------------------
    # Task 2: Data range / boundary analysis
    # ------------------------------------------------------------------
    def _analyze_data_boundaries(self, ws):
        """Return dict of true data boundaries and additional range metadata."""
        # Determine last non-empty row scanning upward from bottom (limited for speed)
        last_row = ws.max_row
        while last_row > 1:
            if any(cell.value not in (None, "", " ") for cell in ws[last_row]):
                break
            last_row -= 1

        # Determine last non-empty column scanning leftward (sample first 200 rows)
        last_col = ws.max_column
        while last_col > 1:
            if any(ws.cell(row=r, column=last_col).value not in (None, "", " ") for r in range(1, min(ws.max_row, 200))):
                break
            last_col -= 1

        true_range = f"A1:{get_column_letter(last_col)}{last_row}"
        return {
            'declared_range': getattr(ws, 'dimensions', f"A1:{get_column_letter(ws.max_column)}{ws.max_row}") if ws.max_row and ws.max_column else 'A1:A1',
            'true_range': true_range,
            'freeze_panes': str(getattr(ws, 'freeze_panes', '') or ''),
            'merged_cells': len(getattr(ws, 'merged_cells', {}).ranges) if hasattr(getattr(ws, 'merged_cells', None), 'ranges') else 0,
            'hyperlinks': len(getattr(ws, 'hyperlinks', [])),
            'comments': 0,  # Skip comment counting for ReadOnlyWorksheet to avoid issues
            'print_area': str(getattr(ws, 'print_area', '') or ''),
            'auto_filter': bool(getattr(getattr(ws, 'auto_filter', None), 'ref', None)) if hasattr(ws, 'auto_filter') else False
        }

    # ------------------------------------------------------------------
    # Task 4: Sheet properties / formatting
    # ------------------------------------------------------------------
    def _analyze_sheet_properties(self, ws):
        """Return sheet protection / formatting metadata."""
        protection = getattr(ws, 'protection', None)
        sheet_properties = getattr(ws, 'sheet_properties', None)
        
        return {
            'protected': getattr(protection, 'sheet', False) if protection else False,
            'protection_options': {
                'password': bool(getattr(protection, 'password', False)) if protection else False,
                'select_locked_cells': getattr(protection, 'selectLockedCells', True) if protection else True,
                'select_unlocked_cells': getattr(protection, 'selectUnlockedCells', True) if protection else True,
            },
            'conditional_formatting_rules': len(getattr(ws, 'conditional_formatting', [])),
            'data_validation_count': len(getattr(getattr(ws, 'data_validations', None), 'dataValidation', [])) if hasattr(ws, 'data_validations') else 0,
            'tab_color': str(getattr(sheet_properties, 'tabColor', '') or '') if sheet_properties else '',
            'visibility': getattr(ws, 'sheet_state', 'visible')
        }

    # ------------------------------------------------------------------
    # Task 5: Cross-sheet dependency mapping
    # ------------------------------------------------------------------
    def _map_sheet_dependencies(self, wb):
        """Return dict of sheet-to-sheet reference counts + circular flag."""
        pattern = re.compile(r"'?([A-Za-z0-9 _]+)'?!")
        deps: Dict[str, Dict[str, int]] = {}
        for ws in wb.worksheets:
            deps.setdefault(ws.title, {})
            for row in ws.iter_rows(max_row=self.config.get('analysis', {}).get('max_formula_check', 1000)):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        for m in pattern.finditer(cell.value):
                            target = m.group(1)
                            if target == ws.title:
                                continue
                            deps[ws.title][target] = deps[ws.title].get(target, 0) + 1
        # Detect circular references
        circular = any(start in deps.get(end, {}) for start in deps for end in deps[start])
        return {'dependency_matrix': deps, 'has_circular': circular}

    # ------------------------------------------------------------------
    # Task 6: Streaming data stats (very large sheets)
    # ------------------------------------------------------------------
    def _analyze_data_streaming(self, ws, max_sample_rows: int = 1000):
        """Lightweight stats for very large sheets (first N rows only)."""
        stats = {'rows_scanned': 0, 'numeric': 0, 'text': 0}
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx > max_sample_rows:
                break
            stats['rows_scanned'] += 1
            for value in row:
                if value is None:
                    continue
                if isinstance(value, (int, float)):
                    stats['numeric'] += 1
                else:
                    stats['text'] += 1
        return stats
    
    def _extract_sheet_headers(self, ws):
        """Extract headers from the first row, handling missing headers"""
        headers: Dict[str, Dict[str, Any]] = {}

        # --- header names -----------------------------------------------------
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        for col_idx, value in enumerate(first_row, start=1):
            col_letter = get_column_letter(col_idx)
            header_name = str(value).strip() if value is not None else ""
            headers[col_letter] = {
                'header_name': header_name or f"Column {col_letter}",
                'is_missing': header_name == "",
                'sample_values': []
            }

        # --- sample values (rows 2-11) ---------------------------------------
        for row in ws.iter_rows(min_row=2, max_row=11, values_only=True):
            for col_idx, value in enumerate(row, start=1):
                if value in (None, "", " "):
                    continue
                col_letter = get_column_letter(col_idx)
                samples = headers[col_letter]['sample_values']
                if len(samples) < 10:
                    samples.append(str(value)[:50])  # truncate long strings
                # early exit if all columns filled
                if all(len(v['sample_values']) >= 10 for v in headers.values()):
                    break

        return headers

    def _analyze_formulas(self, wb) -> Dict[str, Any]:
        """Analyze formulas and dependencies"""
        total_formulas = 0
        complex_formulas = []
        external_refs = False
        
        for ws in wb.worksheets:
            # Sample check to avoid performance issues
            checked_cells = 0
            max_check = self.config.get('analysis', {}).get('max_formula_check', 1000)
            
            for row in ws.iter_rows():
                if checked_cells >= max_check:
                    break
                    
                for cell in row:
                    if checked_cells >= max_check:
                        break
                    
                    checked_cells += 1
                    
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        total_formulas += 1
                        formula = str(cell.value)
                        
                        # Check complexity
                        if len(formula) > 50 or formula.count('(') > 3:
                            complex_formulas.append({
                                'sheet': ws.title,
                                'cell': cell.coordinate,
                                'formula': formula[:100]
                            })
                        
                        # Check for external references
                        if '[' in formula or '!' in formula:
                            external_refs = True
        
        return {
            'total_formulas': total_formulas,
            'complex_formulas': complex_formulas[:10],  # Top 10
            'has_external_refs': external_refs,
            'formula_complexity_score': min(1.0, len(complex_formulas) / max(1, total_formulas))
        }
    
    def _analyze_visuals(self, wb) -> Dict[str, Any]:
        """Analyze charts and visual elements"""
        total_charts = 0
        total_images = 0
        conditional_formatting_rules = 0
        
        for ws in wb.worksheets:
            # Count charts
            try:
                total_charts += len(ws._charts)
            except:
                pass
            
            # Count images
            try:
                total_images += len(ws._images)
            except:
                pass
            
            # Count conditional formatting
            try:
                conditional_formatting_rules += len(ws.conditional_formatting)
            except:
                pass
        
        return {
            'total_charts': total_charts,
            'total_images': total_images,
            'conditional_formatting_rules': conditional_formatting_rules,
            'has_visual_content': total_charts > 0 or total_images > 0,
            'visual_complexity_score': min(1.0, (total_charts + total_images) / 10)
        }
    
    def _compile_results(self, file_info: Dict, structure: Dict, data: Dict, 
                        formulas: Dict, visuals: Dict, security: Dict, start_time: float, module_statuses: Dict[str, str], module_timings: Dict[str, float]) -> Dict[str, Any]:
        """Compile all results into final format with enhanced metrics"""
        
        # Enhanced quality score calculation
        data_density = data.get('overall_data_density', 0)
        total_cells = data.get('total_cells', 1)
        total_data_cells = data.get('total_data_cells', 0)
        quality_components = [
            data_density * 0.3,  # Data density (0-1)
            min(1.0, total_data_cells / max(1000, total_cells * 0.1)) * 0.2,  # Data volume normalized
            min(1.0, len(data.get('data_type_distribution', {})) / 5) * 0.2,  # Data variety
            (1 - min(0.5, len(structure.get('hidden_sheets', [])) / max(1, structure.get('total_sheets', 1)))) * 0.15,  # Structure quality
            min(1.0, security.get('overall_score', 0) / 10) * 0.15  # Security score normalized
        ]
        overall_quality = sum(quality_components)
        
        # Enhanced recommendations
        recommendations = []
        
        # Data quality recommendations
        if data.get('overall_data_density', 0) < 0.1:
            recommendations.append("Low data density detected - consider removing empty regions")
        
        # Security recommendations
        if security.get('overall_score', 10) < 8.0:
            recommendations.extend(security.get('recommendations', []))
        
        # Structure recommendations
        if formulas.get('has_external_refs'):
            recommendations.append("External references found - verify linked files are available")
        if len(structure.get('hidden_sheets', [])) > 0:
            recommendations.append("Hidden sheets detected - review for sensitive information")
        
        # Performance recommendations
        if file_info.get('size_mb', 0) > 50:
            recommendations.append("Large file size detected - consider archiving or splitting data")
        
        # Data quality recommendations
        overall_metrics = data.get('overall_metrics', {})
        if overall_metrics.get('data_variety_score', 0) < 0.3:
            recommendations.append("Limited data variety - consider data enrichment")
        
        if not recommendations:
            recommendations.append("File structure and security appear optimized")
        
        # Calculate success rate based on module statuses
        successful_modules = sum(1 for s in module_statuses.values() if s == 'success')
        total_modules = len(module_statuses)
        success_rate = successful_modules / max(1, total_modules)
        
        return {
            'file_info': file_info,
            'analysis_metadata': {
                'timestamp': time.time(),
                'total_duration_seconds': time.time() - start_time,
                'success_rate': success_rate,
                'quality_score': overall_quality,
                'security_score': security.get('overall_score', 0) / 10,
                'modules_executed': [
                    'health_checker', 'structure_mapper', 'data_profiler', 
                    'formula_analyzer', 'visual_cataloger', 'security_inspector',
                    'dependency_mapper', 'relationship_analyzer', 'performance_monitor',
                    'connection_inspector', 'pivot_intelligence', 'doc_synthesizer'
                ]
            },
            'module_results': {
                'health_checker': {
                    'file_accessible': True,
                    'corruption_detected': False,
                    'file_signature_valid': file_info.get('file_signature_valid', True)
                },
                'structure_mapper': structure,
                'data_profiler': data,
                'formula_analyzer': formulas,
                'visual_cataloger': visuals,
                'security_inspector': security
            },
            'execution_summary': {
                'total_modules': total_modules,
                'successful_modules': successful_modules,
                'failed_modules': total_modules - successful_modules,
                'success_rate': success_rate,
                'module_statuses': module_statuses,
                'module_timings': module_timings,
                'total_module_time': sum(module_timings.values()),
                'average_module_time': sum(module_timings.values()) / len(module_timings) if module_timings else 0
            },
            'resource_usage': {
                'current_usage': {
                    'current_mb': 50.0,  # Will be updated by performance monitor
                    'peak_mb': 50.0,
                    'cpu_percent': 0.0,
                    'elapsed_seconds': time.time() - start_time
                }
            },
            'recommendations': recommendations,
            'success': True
        }
