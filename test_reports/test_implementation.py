#!/usr/bin/env python3
"""
Test the enhanced Excel Explorer implementation
"""

import os
import sys
import tempfile
import shutil
from pathlib import Path

# Add the project directory to the path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Try to import our modules with fallback handling
try:
    from analyzer import SimpleExcelAnalyzer
    from report_generator import ReportGenerator
    print("✅ Successfully imported analyzer and report generator")
except ImportError as e:
    print(f"❌ Import error: {e}")
    sys.exit(1)

def create_test_excel_file():
    """Create a simple test Excel file for validation"""
    try:
        import openpyxl
        from openpyxl import Workbook
        
        # Create a test workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Sheet"
        
        # Add some test data
        headers = ['Name', 'Age', 'City', 'Salary', 'Date']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add sample data
        test_data = [
            ['John Doe', 25, 'New York', 50000, '2023-01-15'],
            ['Jane Smith', 30, 'Los Angeles', 60000, '2023-02-20'],
            ['Bob Johnson', 35, 'Chicago', 70000, '2023-03-10'],
            ['Alice Brown', 28, 'Houston', 55000, '2023-04-05'],
            ['Charlie Davis', 32, 'Phoenix', 65000, '2023-05-12'],
        ]
        
        for row, data in enumerate(test_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Create a temporary file
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(temp_file.name)
        wb.close()
        
        return temp_file.name
        
    except Exception as e:
        print(f"❌ Failed to create test Excel file: {e}")
        return None

def test_analyzer():
    """Test the analyzer functionality"""
    print("\n🔍 Testing analyzer functionality...")
    
    # Create test file
    test_file = create_test_excel_file()
    if not test_file:
        return False
    
    try:
        # Initialize analyzer
        analyzer = SimpleExcelAnalyzer()
        print("✅ Analyzer initialized successfully")
        
        # Run analysis
        print("🔄 Running analysis...")
        results = analyzer.analyze(test_file)
        
        # Check if results have expected structure
        expected_keys = ['file_info', 'analysis_metadata', 'module_results', 'execution_summary']
        for key in expected_keys:
            if key in results:
                print(f"✅ Found {key} in results")
            else:
                print(f"❌ Missing {key} in results")
                return False
        
        # Check file info
        file_info = results.get('file_info', {})
        print(f"📊 File info keys: {list(file_info.keys())}")
        if file_info.get('name') and file_info.get('size_mb') is not None:
            print(f"✅ File info: {file_info.get('name')} ({file_info.get('size_mb'):.2f} MB)")
        else:
            print(f"❌ Missing file info - name: {file_info.get('name')}, size_mb: {file_info.get('size_mb')}")
            return False
        
        # Check analysis metadata
        metadata = results.get('analysis_metadata', {})
        if metadata.get('quality_score') is not None:
            print(f"✅ Quality score: {metadata.get('quality_score'):.2f}")
        else:
            print("❌ Missing quality score")
            return False
        
        print("✅ Analyzer test completed successfully")
        return True
        
    except Exception as e:
        print(f"❌ Analyzer test failed: {e}")
        return False
    finally:
        # Clean up
        if test_file and os.path.exists(test_file):
            os.unlink(test_file)

def test_report_generator():
    """Test the report generator functionality"""
    print("\n📊 Testing report generator functionality...")
    
    # Create test file
    test_file = create_test_excel_file()
    if not test_file:
        return False
    
    try:
        # Initialize analyzer and generate results
        analyzer = SimpleExcelAnalyzer()
        results = analyzer.analyze(test_file)
        
        # Initialize report generator
        report_gen = ReportGenerator()
        print("✅ Report generator initialized successfully")
        
        # Generate HTML report
        output_dir = Path("test_reports")
        output_dir.mkdir(exist_ok=True)
        
        html_file = output_dir / "test_report.html"
        report_gen.generate_html_report(results, str(html_file))
        
        if html_file.exists():
            print(f"✅ HTML report generated: {html_file}")
            print(f"📄 Report size: {html_file.stat().st_size / 1024:.1f} KB")
        else:
            print("❌ HTML report not generated")
            return False
        
        # Generate JSON report
        json_file = output_dir / "test_report.json"
        report_gen.generate_json_report(results, str(json_file))
        
        if json_file.exists():
            print(f"✅ JSON report generated: {json_file}")
            print(f"📄 Report size: {json_file.stat().st_size / 1024:.1f} KB")
        else:
            print("❌ JSON report not generated")
            return False
        
        print("✅ Report generator test completed successfully")
        return True
        
    except Exception as e:
        print(f"❌ Report generator test failed: {e}")
        return False
    finally:
        # Clean up
        if test_file and os.path.exists(test_file):
            os.unlink(test_file)

def main():
    """Run all tests"""
    print("🚀 Starting Excel Explorer Enhanced Test Suite")
    print("=" * 50)
    
    # Test analyzer
    analyzer_success = test_analyzer()
    
    # Test report generator
    report_success = test_report_generator()
    
    print("\n" + "=" * 50)
    print("📋 Test Results Summary:")
    print(f"Analyzer: {'✅ PASSED' if analyzer_success else '❌ FAILED'}")
    print(f"Report Generator: {'✅ PASSED' if report_success else '❌ FAILED'}")
    
    if analyzer_success and report_success:
        print("\n🎉 All tests passed! The enhanced Excel Explorer is ready.")
        return 0
    else:
        print("\n⚠️  Some tests failed. Please check the implementation.")
        return 1

if __name__ == "__main__":
    sys.exit(main())