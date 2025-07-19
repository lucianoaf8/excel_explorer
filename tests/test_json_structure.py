#!/usr/bin/env python3
"""
Quick test to see JSON report structure
"""

import tempfile
import openpyxl
from pathlib import Path
import json
import sys
import os

# Add src to path so we can import modules
src_path = Path(__file__).parent.parent / "src"
sys.path.insert(0, str(src_path))

from core import SimpleExcelAnalyzer
from reports import ReportGenerator

def create_test_file():
    """Create a simple test Excel file"""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Data"
        
        # Add some test data
        ws['A1'] = "Name"
        ws['B1'] = "Age" 
        ws['A2'] = "Alice"
        ws['B2'] = 25
        
        wb.save(tmp.name)
        return tmp.name

def main():
    """Test JSON report structure"""
    test_file = create_test_file()
    
    try:
        print(f"Testing with file: {test_file}")
        
        # Run analysis
        analyzer = SimpleExcelAnalyzer('config.yaml')
        results = analyzer.analyze(test_file)
        
        # Generate JSON report
        output_dir = Path(tempfile.gettempdir()) / "test_reports"
        output_dir.mkdir(exist_ok=True)
        
        json_file = output_dir / "test_structure.json"
        generator = ReportGenerator()
        generator.generate_json_report(results, str(json_file))
        
        # Read and examine the JSON structure
        with open(json_file, 'r') as f:
            json_data = json.load(f)
        
        print("\nJSON Report Structure:")
        print("=" * 40)
        
        def print_structure(data, indent=0):
            for key, value in data.items():
                if isinstance(value, dict):
                    print(f"{'  ' * indent}{key}: (dict with {len(value)} keys)")
                    if indent < 2:  # Only show 2 levels deep
                        print_structure(value, indent + 1)
                elif isinstance(value, list):
                    print(f"{'  ' * indent}{key}: (list with {len(value)} items)")
                else:
                    print(f"{'  ' * indent}{key}: {type(value).__name__}")
        
        print_structure(json_data)
        
        print(f"\nJSON file created at: {json_file}")
        print(f"File size: {json_file.stat().st_size} bytes")
        
    finally:
        # Cleanup
        try:
            os.unlink(test_file)
        except:
            pass

if __name__ == "__main__":
    main()
