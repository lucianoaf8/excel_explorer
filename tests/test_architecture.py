#!/usr/bin/env python3
"""
Integration Test for Excel Explorer v2.0 Architecture
Tests the consolidated entry points and CLI functionality
"""

import sys
import subprocess
import tempfile
from pathlib import Path
import openpyxl
import json


def create_test_excel_file(file_path: str) -> str:
    """Create a simple test Excel file for testing"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Data"
    
    # Add some test data
    ws['A1'] = "Name"
    ws['B1'] = "Age" 
    ws['C1'] = "Score"
    ws['A2'] = "Alice"
    ws['B2'] = 25
    ws['C2'] = 95.5
    ws['A3'] = "Bob"
    ws['B3'] = 30
    ws['C3'] = 87.2
    
    # Add a second sheet
    ws2 = wb.create_sheet("Summary")
    ws2['A1'] = "Total Records"
    ws2['B1'] = 2
    
    wb.save(file_path)
    return file_path


def test_entry_point_consolidation():
    """Test that main.py is the unified entry point"""
    print("Testing entry point consolidation...")
    
    # Test GUI mode help
    result = subprocess.run([
        sys.executable, "main.py", "--help"
    ], capture_output=True, text=True, cwd="C:\\Projects\\excel_explorer")
    
    if result.returncode == 0:
        print("PASS: Main.py help command works")
        # Check that it mentions both CLI and GUI modes
        if "cli" in result.stdout.lower() and "gui" in result.stdout.lower():
            print("PASS: Both CLI and GUI modes mentioned in help")
        else:
            print("FAIL: Missing mode options in help output")
    else:
        print(f"FAIL: Main.py help failed: {result.stderr}")
        return False
    
    return True


def test_cli_functionality():
    """Test CLI mode with a temporary Excel file"""
    print("Testing CLI functionality...")
    
    with tempfile.TemporaryDirectory() as temp_dir:
        # Create test Excel file
        test_file = Path(temp_dir) / "test.xlsx"
        create_test_excel_file(str(test_file))
        
        # Test CLI analysis with JSON output
        output_dir = Path(temp_dir) / "reports"
        
        result = subprocess.run([
            sys.executable, "main.py",
            "--mode", "cli",
            "--file", str(test_file),
            "--format", "json",
            "--output", str(output_dir),
            "--verbose"
        ], capture_output=True, text=True, cwd=str(Path(__file__).parent.parent))
        
        if result.returncode == 0:
            print("PASS: CLI analysis completed successfully")
            
            # Check if report was generated
            json_files = list(output_dir.glob("*.json"))
            if json_files:
                json_file = json_files[0]
                print(f"PASS: JSON report generated: {json_file.name}")
                
                # Validate JSON content
                try:
                    with open(json_file, 'r') as f:
                        report_data = json.load(f)
                    
                    # Check for required sections (new standardized structure)
                    required_sections = ['file_summary', 'quality_metrics', 'module_execution']
                    missing_sections = [s for s in required_sections if s not in report_data]
                    
                    if not missing_sections:
                        print("PASS: JSON report contains all required sections")
                        
                        # Check specific metrics
                        file_summary = report_data.get('file_summary', {})
                        if file_summary.get('sheet_count') == 2:
                            print("PASS: Correct sheet count detected")
                        else:
                            print(f"FAIL: Expected 2 sheets, got {file_summary.get('sheet_count')}")
                        
                    else:
                        print(f"FAIL: Missing sections in JSON report: {missing_sections}")
                        return False
                        
                except json.JSONDecodeError as e:
                    print(f"FAIL: Invalid JSON format: {e}")
                    return False
            else:
                print("FAIL: No JSON report generated")
                return False
        else:
            print(f"FAIL: CLI analysis failed: {result.stderr}")
            return False
    
    return True


def test_configuration_loading():
    """Test configuration management"""
    print("Testing configuration loading...")
    
    # Test config manager directly
    try:
        # Add src to path for imports
        src_path = Path(__file__).parent.parent / "src"
        sys.path.insert(0, str(src_path))
        
        from core import ConfigManager
        
        config = ConfigManager()
        cfg = config.load_config()
        
        # Check default values
        sample_rows = config.get('analysis.sample_rows')
        if sample_rows is not None:
            print(f"PASS: Configuration loaded: sample_rows = {sample_rows}")
        else:
            print("FAIL: Configuration loading failed")
            return False
            
        # Test environment override
        import os
        os.environ['EXCEL_EXPLORER_SAMPLE_ROWS'] = '150'
        config.reload_config()
        
        updated_sample_rows = config.get('analysis.sample_rows')
        if updated_sample_rows == 150:
            print("PASS: Environment variable override works")
        else:
            print(f"FAIL: Environment override failed: got {updated_sample_rows}, expected 150")
            return False
            
    except Exception as e:
        print(f"FAIL: Configuration test failed: {e}")
        return False
    
    return True


def test_report_consistency():
    """Test report consistency validation"""
    print("Testing report consistency...")
    
    # Check if validate_reports.py exists and can be imported
    try:
        from utils import validate_reports
        print("PASS: Report validation module imported successfully")
        
        # Test ReportDataModel
        from reports import ReportDataModel
        
        sample_results = {
            'file_info': {
                'name': 'test.xlsx',
                'size_mb': 1.5,
                'sheet_count': 2
            },
            'analysis_metadata': {
                'quality_score': 0.85,
                'security_score': 9.2
            }
        }
        
        data_model = ReportDataModel(sample_results)
        standardized = data_model.get_standardized_data()
        
        if 'file_summary' in standardized and 'quality_metrics' in standardized:
            print("PASS: Report data model standardization works")
        else:
            print("FAIL: Report data model standardization failed")
            return False
            
    except Exception as e:
        print(f"FAIL: Report consistency test failed: {e}")
        return False
    
    return True


def main():
    """Run all integration tests"""
    print("Starting Excel Explorer v2.0 Integration Tests")
    print("="*60)
    
    tests = [
        ("Entry Point Consolidation", test_entry_point_consolidation),
        ("CLI Functionality", test_cli_functionality),
        ("Configuration Loading", test_configuration_loading),
        ("Report Consistency", test_report_consistency),
    ]
    
    passed = 0
    total = len(tests)
    
    for test_name, test_func in tests:
        print(f"\n{test_name}")
        print("-" * 40)
        
        try:
            if test_func():
                passed += 1
                print(f"PASS: {test_name}")
            else:
                print(f"FAIL: {test_name}")
        except Exception as e:
            print(f"ERROR: {test_name} - {e}")
    
    # Summary
    print("\n" + "="*60)
    print(f"TEST RESULTS: {passed}/{total} PASSED")
    
    if passed == total:
        print("All tests passed! Architecture consolidation successful.")
        return 0
    else:
        print("Some tests failed. Check the output above for details.")
        return 1


if __name__ == "__main__":
    sys.exit(main())
