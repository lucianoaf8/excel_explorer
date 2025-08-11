#!/usr/bin/env python3
"""
Tests for Excel Anonymizer functionality
"""

import unittest
import tempfile
import json
from pathlib import Path
import sys
import os

# Add src to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

try:
    import openpyxl
    from utils.anonymizer import ExcelAnonymizer, anonymize_file
except ImportError as e:
    print(f"Import error: {e}")
    print("Make sure to install dependencies: pip install faker openpyxl")
    sys.exit(1)


class TestExcelAnonymizer(unittest.TestCase):
    """Test cases for Excel Anonymizer"""
    
    def setUp(self):
        """Create a test Excel file for each test"""
        self.temp_dir = tempfile.mkdtemp()
        self.test_file = Path(self.temp_dir) / "test_data.xlsx"
        
        # Create test Excel file with sample data
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Headers
        headers = ["Name", "Company", "Email", "Phone", "Address", "Amount"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Sample data
        test_data = [
            ["John Doe", "Acme Corp", "john@acme.com", "555-1234", "123 Main St", 1000],
            ["Jane Smith", "Beta Inc", "jane@beta.com", "555-5678", "456 Oak Ave", 2000],
            ["Bob Johnson", "Gamma LLC", "bob@gamma.com", "555-9012", "789 Pine Rd", 1500],
            ["Alice Brown", "Delta Co", "alice@delta.com", "555-3456", "321 Elm St", 3000],
            ["", "", "", "", "", 0],  # Empty row
            ["Mike Wilson", "Echo Systems", "mike@echo.com", "555-7890", "654 Maple Dr", 2500]
        ]
        
        for row, data in enumerate(test_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Add another sheet
        ws2 = wb.create_sheet("Contractors")
        headers2 = ["Contractor Name", "Client Organization", "Project", "Status"]
        for col, header in enumerate(headers2, 1):
            ws2.cell(row=1, column=col, value=header)
        
        contractor_data = [
            ["David Lee", "TechCorp", "Web Development", "Active"],
            ["Sarah Wilson", "InnovateLab", "Data Analysis", "Complete"],
            ["Tom Anderson", "StartupXYZ", "Mobile App", "In Progress"]
        ]
        
        for row, data in enumerate(contractor_data, 2):
            for col, value in enumerate(data, 1):
                ws2.cell(row=row, column=col, value=value)
        
        wb.save(self.test_file)
        self.wb = wb
    
    def tearDown(self):
        """Clean up test files"""
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)
    
    def test_file_loading(self):
        """Test that the anonymizer can load an Excel file"""
        anonymizer = ExcelAnonymizer(str(self.test_file))
        self.assertIsNotNone(anonymizer.workbook)
        self.assertEqual(len(anonymizer.workbook.sheetnames), 2)
    
    def test_sensitive_column_detection(self):
        """Test automatic detection of sensitive columns"""
        anonymizer = ExcelAnonymizer(str(self.test_file))
        sensitive = anonymizer.find_sensitive_columns()
        
        # Should find sensitive columns in both sheets
        self.assertIn("Sheet1", sensitive)
        self.assertIn("Contractors", sensitive)
        
        # Check Sheet1 columns
        sheet1_cols = dict(sensitive["Sheet1"])
        self.assertIn("A", sheet1_cols)  # Name column
        self.assertEqual(sheet1_cols["A"], "name")
        self.assertIn("B", sheet1_cols)  # Company column
        self.assertEqual(sheet1_cols["B"], "company")
        self.assertIn("C", sheet1_cols)  # Email column
        self.assertEqual(sheet1_cols["C"], "email")
        
        # Check Contractors sheet columns
        contractor_cols = dict(sensitive["Contractors"])
        self.assertIn("A", contractor_cols)  # Contractor Name
        self.assertEqual(contractor_cols["A"], "name")
        self.assertIn("B", contractor_cols)  # Client Organization
        self.assertEqual(contractor_cols["B"], "company")
    
    def test_anonymization_consistency(self):
        """Test that same values get same fake replacements"""
        anonymizer = ExcelAnonymizer(str(self.test_file))
        
        # Get fake value twice for same original
        fake1 = anonymizer._get_fake_value("John Doe", "name", "test_col")
        fake2 = anonymizer._get_fake_value("John Doe", "name", "test_col")
        
        # Should be identical
        self.assertEqual(fake1, fake2)
        
        # Different originals should get different fakes
        fake3 = anonymizer._get_fake_value("Jane Smith", "name", "test_col")
        self.assertNotEqual(fake1, fake3)
    
    def test_full_anonymization(self):
        """Test complete anonymization process"""
        anonymizer = ExcelAnonymizer(str(self.test_file))
        
        # Anonymize with auto-detection
        stats = anonymizer.anonymize_columns(auto_detect=True)
        
        # Should have processed multiple columns
        self.assertGreater(len(stats), 0)
        
        # Check that values were actually changed
        ws = anonymizer.workbook["Sheet1"]
        
        # Original names should be replaced
        name_cell = ws.cell(row=2, column=1)  # First data row, Name column
        self.assertNotEqual(name_cell.value, "John Doe")
        self.assertIsNotNone(name_cell.value)
        
        # Company should be replaced
        company_cell = ws.cell(row=2, column=2)  # First data row, Company column
        self.assertNotEqual(company_cell.value, "Acme Corp")
        self.assertIsNotNone(company_cell.value)
    
    def test_mapping_generation(self):
        """Test that mappings are properly generated"""
        anonymizer = ExcelAnonymizer(str(self.test_file))
        anonymizer.anonymize_columns(auto_detect=True)
        
        # Should have mappings
        self.assertGreater(len(anonymizer.mappings), 0)
        
        # Mappings should contain original->fake pairs
        for column_key, mapping in anonymizer.mappings.items():
            self.assertIsInstance(mapping, dict)
            if mapping:  # If there are mappings
                # Should have string keys and values
                for original, fake in mapping.items():
                    self.assertIsInstance(original, str)
                    self.assertIsInstance(fake, str)
                    self.assertNotEqual(original, fake)
    
    def test_save_anonymized_file(self):
        """Test saving anonymized file"""
        anonymizer = ExcelAnonymizer(str(self.test_file))
        anonymizer.anonymize_columns(auto_detect=True)
        
        # Save anonymized file
        output_path = anonymizer.save_anonymized_file()
        
        # File should exist
        self.assertTrue(Path(output_path).exists())
        
        # Should be able to load the saved file
        wb = openpyxl.load_workbook(output_path)
        self.assertEqual(len(wb.sheetnames), 2)
    
    def test_mapping_save_json(self):
        """Test saving mapping to JSON"""
        anonymizer = ExcelAnonymizer(str(self.test_file))
        anonymizer.anonymize_columns(auto_detect=True)
        
        # Save mappings as JSON
        mapping_path = anonymizer.save_mappings(format='json')
        
        # File should exist
        self.assertTrue(Path(mapping_path).exists())
        
        # Should be valid JSON
        with open(mapping_path, 'r') as f:
            data = json.load(f)
        
        # Should have expected structure
        self.assertIn('metadata', data)
        self.assertIn('mappings', data)
        self.assertIn('created', data['metadata'])
        self.assertIn('source_file', data['metadata'])
    
    def test_mapping_save_excel(self):
        """Test saving mapping to Excel"""
        anonymizer = ExcelAnonymizer(str(self.test_file))
        anonymizer.anonymize_columns(auto_detect=True)
        
        # Save mappings as Excel
        mapping_path = anonymizer.save_mappings(format='excel')
        
        # File should exist
        self.assertTrue(Path(mapping_path).exists())
        
        # Should be valid Excel file
        wb = openpyxl.load_workbook(mapping_path)
        
        # Should have Summary sheet
        self.assertIn('Summary', wb.sheetnames)
        
        # Should have at least one mapping sheet
        self.assertGreater(len(wb.sheetnames), 1)
    
    def test_reverse_anonymization(self):
        """Test reversing anonymization using mappings"""
        # First, anonymize the file
        anonymizer = ExcelAnonymizer(str(self.test_file))
        anonymizer.anonymize_columns(auto_detect=True)
        
        # Save anonymized file and mappings
        anon_path = anonymizer.save_anonymized_file()
        mapping_path = anonymizer.save_mappings(format='json')
        
        # Now reverse the anonymization
        restored_path = anonymizer.reverse_anonymization(anon_path, mapping_path)
        
        # Restored file should exist
        self.assertTrue(Path(restored_path).exists())
        
        # Load original and restored files
        original_wb = openpyxl.load_workbook(str(self.test_file))
        restored_wb = openpyxl.load_workbook(restored_path)
        
        # Compare key values (should be identical)
        original_ws = original_wb["Sheet1"]
        restored_ws = restored_wb["Sheet1"]
        
        # Check that John Doe is restored in first data row
        original_name = original_ws.cell(row=2, column=1).value
        restored_name = restored_ws.cell(row=2, column=1).value
        self.assertEqual(original_name, restored_name)
    
    def test_convenience_function(self):
        """Test the convenience anonymize_file function"""
        result = anonymize_file(
            file_path=str(self.test_file),
            auto_detect=True,
            mapping_format='json'
        )
        
        anon_path, mapping_path = result
        
        # Both files should be created
        self.assertTrue(Path(anon_path).exists())
        self.assertTrue(Path(mapping_path).exists())
        
        # Should be able to load both files
        wb = openpyxl.load_workbook(anon_path)
        self.assertIsNotNone(wb)
        
        with open(mapping_path, 'r') as f:
            mapping_data = json.load(f)
        self.assertIn('mappings', mapping_data)
    
    def test_empty_cells_handling(self):
        """Test that empty cells are handled properly"""
        anonymizer = ExcelAnonymizer(str(self.test_file))
        anonymizer.anonymize_columns(auto_detect=True)
        
        # Check that empty cells remain empty
        ws = anonymizer.workbook["Sheet1"]
        empty_cell = ws.cell(row=6, column=1)  # Row with empty data
        self.assertIsNone(empty_cell.value)
    
    def test_specific_columns(self):
        """Test anonymizing specific columns only"""
        anonymizer = ExcelAnonymizer(str(self.test_file))
        
        # Anonymize only Name column in Sheet1
        columns_to_anonymize = {
            "Sheet1": [("A", "name")]
        }
        
        stats = anonymizer.anonymize_columns(columns_to_anonymize, auto_detect=False)
        
        # Should have processed only one column
        self.assertEqual(len(stats), 1)
        
        # Name should be changed, Company should not
        ws = anonymizer.workbook["Sheet1"]
        name_cell = ws.cell(row=2, column=1)
        company_cell = ws.cell(row=2, column=2)
        
        self.assertNotEqual(name_cell.value, "John Doe")  # Changed
        self.assertEqual(company_cell.value, "Acme Corp")  # Unchanged


if __name__ == '__main__':
    # Run the tests
    unittest.main(verbosity=2)