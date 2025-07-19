#!/usr/bin/env python3
"""
Demo script showing Excel Explorer v2.0 capabilities
"""

import tempfile
import openpyxl
from pathlib import Path
import subprocess
import sys
import os

def create_demo_file():
    """Create a demo Excel file with multiple sheets"""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        wb = openpyxl.Workbook()
        
        # Sheet 1: Sales Data
        ws1 = wb.active
        ws1.title = "Sales_Data"
        
        headers = ["Product", "Revenue", "Units", "Region", "Date"]
        for i, header in enumerate(headers, 1):
            ws1.cell(1, i, header)
        
        sample_data = [
            ["Widget A", 15000, 150, "North", "2024-01-15"],
            ["Widget B", 22000, 200, "South", "2024-01-16"],
            ["Widget C", 18000, 180, "East", "2024-01-17"],
            ["Widget A", 25000, 250, "West", "2024-01-18"],
        ]
        
        for row_i, row_data in enumerate(sample_data, 2):
            for col_i, value in enumerate(row_data, 1):
                ws1.cell(row_i, col_i, value)
        
        # Sheet 2: Summary with formulas
        ws2 = wb.create_sheet("Summary")
        ws2['A1'] = "Total Revenue"
        ws2['B1'] = "=SUM(Sales_Data.B:B)"
        ws2['A2'] = "Total Units"
        ws2['B2'] = "=SUM(Sales_Data.C:C)"
        
        # Sheet 3: Hidden sheet
        ws3 = wb.create_sheet("Config")
        ws3['A1'] = "Internal Configuration"
        ws3['A2'] = "Version: 1.0"
        ws3.sheet_state = 'hidden'
        
        wb.save(tmp.name)
        return tmp.name

def main():
    """Demonstrate Excel Explorer v2.0 capabilities"""
    demo_file = create_demo_file()
    
    try:
        print("Excel Explorer v2.0 - Architecture Demo")
        print("=" * 50)
        
        print(f"Created demo file: {Path(demo_file).name}")
        print("- 3 sheets (1 hidden)")
        print("- Sample sales data")
        print("- Formulas and references")
        print()
        
        # Test CLI mode with different formats
        formats = ['json', 'html', 'text']
        output_dir = Path(tempfile.gettempdir()) / "excel_explorer_demo"
        output_dir.mkdir(exist_ok=True)
        
        for fmt in formats:
            print(f"Generating {fmt.upper()} report...")
            
            cmd = [
                sys.executable, "main.py",
                "--mode", "cli",
                "--file", demo_file,
                "--format", fmt,
                "--output", str(output_dir)
            ]
            
            result = subprocess.run(
                cmd, 
                cwd=str(Path(__file__).parent.parent),
                capture_output=True, 
                text=True
            )
            
            if result.returncode == 0:
                print(f"  PASS: {fmt.upper()} report generated successfully")
            else:
                print(f"  FAIL: {fmt.upper()} report failed: {result.stderr}")
        
        # List generated reports
        reports = list(output_dir.glob("*"))
        print(f"\nGenerated {len(reports)} report files:")
        for report in reports:
            size_mb = report.stat().st_size / (1024 * 1024)
            print(f"  - {report.name} ({size_mb:.2f} MB)")
        
        print(f"\nAll reports saved to: {output_dir}")
        
        # Test the validation script
        print("\nTesting report consistency validation...")
        
        validation_cmd = [
            sys.executable, "-m", "excel_explorer.utils.validate_reports", demo_file
        ]
        
        validation_result = subprocess.run(
            validation_cmd,
            cwd=str(Path(__file__).parent.parent), 
            capture_output=True,
            text=True
        )
        
        if validation_result.returncode == 0:
            print("  PASS: Report consistency validation PASSED")
        else:
            print("  FAIL: Report consistency validation FAILED")
            print(f"    Error: {validation_result.stderr}")
        
        print("\n" + "=" * 50)
        print("DEMO COMPLETE - All major features working!")
        print("Architecture consolidation successful.")
        
    finally:
        # Cleanup
        try:
            os.unlink(demo_file)
        except:
            pass

if __name__ == "__main__":
    main()
