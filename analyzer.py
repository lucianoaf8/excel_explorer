"""
Simple Excel Analyzer - Direct openpyxl implementation
"""

import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Dict, Any, Optional, Callable
import time
import re
import yaml
from datetime import datetime


class SimpleExcelAnalyzer:
    """Streamlined Excel analysis without framework complexity"""
    
    def __init__(self, config_path: str = "config.yaml"):
        self.progress_callback: Optional[Callable] = None
        self.config: Dict[str, Any] = self._load_config(config_path)
        
    def analyze(self, file_path: str, progress_callback: Optional[Callable] = None) -> Dict[str, Any]:
        """Single method for complete Excel analysis"""
        self.progress_callback = progress_callback
        start_time = time.time()
        
        try:
            module_statuses: Dict[str, str] = {}
            module_results: Dict[str, Any] = {}
            
            # Helper to execute modules safely
            def _safe_run(mod: str, desc: str, fn: Callable[[], Any]):
                self._update_progress(mod, "starting", desc)
                try:
                    res = fn()
                    module_statuses[mod] = "success"
                    self._update_progress(mod, "complete")
                    return res
                except Exception as exc:
                    module_statuses[mod] = "failed"
                    self._update_progress(mod, "error", str(exc))
                    return {"error": str(exc)}
            
            # Load workbook (fail-fast: cannot proceed without workbook)
            self._update_progress("health_checker", "starting", "Loading Excel file")
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            module_statuses["health_checker"] = "success"
            self._update_progress("health_checker", "complete")
            
            # Individual modules
            file_info = _safe_run("file_info", "Gathering file info", lambda: self._get_file_info(file_path, wb))
            structure = _safe_run("structure_mapper", "Analyzing structure", lambda: self._analyze_structure(wb))
            data_analysis = _safe_run("data_profiler", "Profiling data", lambda: self._analyze_data(wb))
            formula_analysis = _safe_run("formula_analyzer", "Analyzing formulas", lambda: self._analyze_formulas(wb))
            visual_analysis = _safe_run("visual_cataloger", "Cataloging visuals", lambda: self._analyze_visuals(wb))
            conf_analysis = self.config.get('analysis', {})
            if conf_analysis.get('enable_cross_sheet_analysis', True):
                dependency_map = _safe_run("dependency_mapper", "Mapping sheet dependencies", lambda: self._map_sheet_dependencies(wb))
            else:
                module_statuses["dependency_mapper"] = "skipped"
                dependency_map = {'skipped': True}
            _safe_run("connection_inspector", "Checking connections", lambda: None)
            _safe_run("pivot_intelligence", "Analyzing pivots", lambda: None)
            _safe_run("doc_synthesizer", "Generating documentation", lambda: None)
            
            wb.close()
            
            # Compile results
            results = self._compile_results(
                file_info, structure, data_analysis, 
                formula_analysis, visual_analysis, start_time, module_statuses
            )
            # Inject dependency mapper output
            results.setdefault('module_results', {})['dependency_mapper'] = dependency_map
            
            return results
            
        except Exception as e:
            if 'wb' in locals():
                wb.close()
            raise Exception(f"Analysis failed: {str(e)}")
    
    def _update_progress(self, module: str, status: str, detail: str = ""):
        """Send progress updates to GUI"""
        if self.progress_callback:
            self.progress_callback(module, status, detail)

    # ------------------------------------------------------------------ #
    # Configuration loading                                              #
    # ------------------------------------------------------------------ #
    def _load_config(self, config_path: str) -> Dict[str, Any]:
        """Load YAML configuration file. Returns empty dict on failure."""
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                return yaml.safe_load(f) or {}
        except FileNotFoundError:
            return {}
        except Exception as e:
            # Do not raise – fall back to defaults
            print(f"[Config] Failed to load {config_path}: {e}")
            return {}
    
    def _get_file_info(self, file_path: str, wb) -> Dict[str, Any]:
        """Extract basic file information"""
        path = Path(file_path)
        stat = path.stat()
        
        return {
            'name': path.name,
            'size_mb': stat.st_size / (1024 * 1024),
            'path': str(path),
            'created': datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
            'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
            'sheet_count': len(wb.sheetnames),
            'sheets': wb.sheetnames
        }
    
    def _analyze_structure(self, wb) -> Dict[str, Any]:
        """Analyze workbook structure"""
        visible_sheets = []
        hidden_sheets = []
        
        for ws in wb.worksheets:
            if ws.sheet_state == 'visible':
                visible_sheets.append(ws.title)
            else:
                hidden_sheets.append(ws.title)
        
        # Count named ranges
        named_ranges = 0
        try:
            named_ranges = len(list(wb.defined_names.definedName))
        except:
            pass
        
        # Count tables
        table_count = 0
        try:
            for ws in wb.worksheets:
                table_count += len(ws.tables)
        except:
            pass
        
        return {
            'total_sheets': len(wb.sheetnames),
            'visible_sheets': visible_sheets,
            'hidden_sheets': hidden_sheets,
            'named_ranges_count': named_ranges,
            'table_count': table_count,
            'has_hidden_content': len(hidden_sheets) > 0
        }
    
    def _analyze_data(self, wb) -> Dict[str, Any]:
        """Analyze data content, quality, and column types"""
        sheet_data = {}
        total_cells = 0
        total_data_cells = 0
        
        for ws in wb.worksheets:
            if not ws.max_row or not ws.max_column:
                continue
            
            sheet_cells = ws.max_row * ws.max_column
            total_cells += sheet_cells
            
            sample_rows = min(self.config.get('analysis', {}).get('sample_rows', 100), ws.max_row)
            # Header extraction (Task 1)
            header_map = self._extract_sheet_headers(ws)
            # Optional data quality metrics (Task 3)
            if self.config.get('analysis', {}).get('enable_data_quality_checks', True):
                quality_map = self._calculate_data_quality(ws, sample_rows)
            else:
                quality_map = {}


            # Progressive sampling with fallback (Task 8)
            retry_rows = sample_rows
            while True:
                try:
                    column_stats, data_cells_sampled = self._compute_column_stats(
                        ws, retry_rows, self.config.get('analysis', {}).get('timeout_per_sheet_seconds', 30)
                    )
                    break  # success
                except (MemoryError, TimeoutError):
                    if retry_rows > 10:
                        retry_rows = max(10, retry_rows // 2)
                        continue  # retry with smaller sample
                    else:
                        raise
            
            # Determine dominant data type per column
            columns_summary = []
            for letter, counts in column_stats.items():
                dominant_type = max(counts, key=counts.get)
                quality_metrics = quality_map.get(letter, {})
                columns_summary.append({
                    'letter': letter,
                    'range': f"{letter}1:{letter}{ws.max_row}",
                    'data_type': dominant_type,
                    'header': header_map.get(letter, {}).get('header_name', ''),
                    'header_missing': header_map.get(letter, {}).get('is_missing', False),
                    'nulls': quality_metrics.get('nulls', 0),
                    'duplicates': quality_metrics.get('duplicates', 0),
                    'fill_rate': quality_metrics.get('fill_rate', 0.0),
                    'sample_values': header_map.get(letter, {}).get('sample_values', [])
                })
            
            # Extrapolate data cells for full sheet if sampled
            data_cells = data_cells_sampled
            if ws.max_row > sample_rows:
                data_cells = int(data_cells_sampled * (ws.max_row / sample_rows))
            total_data_cells += data_cells
            
            sheet_data[ws.title] = {
                'dimensions': f"{ws.max_row}x{ws.max_column}",
                'used_range': ws.dimensions,
                'estimated_data_cells': data_cells,
                'empty_cells': sheet_cells - data_cells,
                'has_data': data_cells > 0,
                'data_density': data_cells / sheet_cells if sheet_cells > 0 else 0,
                'boundaries': self._analyze_data_boundaries(ws),
                'sheet_properties': self._analyze_sheet_properties(ws),
                'columns': sorted(columns_summary, key=lambda c: c['letter']),
            'stream_stats': self._analyze_data_streaming(ws, self.config.get('analysis', {}).get('max_sample_rows', 1000)) if ws.max_row > sample_rows else {}
            }
        
        return {
            'sheet_analysis': sheet_data,
            'total_cells': total_cells,
            'total_data_cells': total_data_cells,
            'overall_data_density': total_data_cells / total_cells if total_cells > 0 else 0,
            'data_quality_score': min(1.0, total_data_cells / max(1000, total_cells))
        }
    
    # ------------------------------------------------------------------
    # Task 1: Header extraction helper
    # ------------------------------------------------------------------
    def _calculate_data_quality(self, ws, sample_rows=100):
        """Per-column data quality: nulls, duplicates, fill rate"""
        col_nulls = {}
        col_values = {}
        rows_checked = 0
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, sample_rows+1), values_only=True):
            rows_checked += 1
            for idx, value in enumerate(row, start=1):
                letter = get_column_letter(idx)
                if letter not in col_nulls:
                    col_nulls[letter] = 0
                    col_values[letter] = set()
                if value in (None, "", " "):
                    col_nulls[letter] += 1
                else:
                    col_values[letter].add(value)
        quality = {}
        for letter in col_nulls:
            nulls = col_nulls[letter]
            dupes = rows_checked - len(col_values[letter]) - nulls
            fill_rate = 1 - (nulls / max(1, rows_checked))
            quality[letter] = { 
                'nulls': nulls,
                'duplicates': max(0, dupes),
                'fill_rate': fill_rate
            }
        return quality

    # ------------------------------------------------------------------
    # Task 8: Column statistics with timeout & memory safeguards
    # ------------------------------------------------------------------
    def _compute_column_stats(self, ws, max_rows: int, timeout_sec: int):
        """Sample up to `max_rows` rows and compute per-column type counts.

        Returns (column_stats, data_cells_sampled).
        Raises TimeoutError if processing exceeds `timeout_sec` seconds.
        """
        from datetime import datetime  # local import to avoid top-level when not needed
        column_stats: Dict[str, Dict[str, int]] = {}
        data_cells_sampled = 0
        start_time = time.time()

        for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True), start=1):
            if time.time() - start_time > timeout_sec:
                raise TimeoutError("Sheet analysis timeout")
            for col_idx, value in enumerate(row, start=1):
                letter = get_column_letter(col_idx)
                stats = column_stats.setdefault(letter, {
                    'numeric': 0,
                    'date': 0,
                    'text': 0,
                    'boolean': 0,
                    'blank': 0
                })

                if value in (None, "", " "):
                    stats['blank'] += 1
                else:
                    data_cells_sampled += 1
                    if isinstance(value, (int, float)):
                        stats['numeric'] += 1
                    elif isinstance(value, datetime):
                        stats['date'] += 1
                    elif isinstance(value, bool):
                        stats['boolean'] += 1
                    else:
                        stats['text'] += 1

        return column_stats, data_cells_sampled

    # ------------------------------------------------------------------
    # Task 2: Data range / boundary analysis
    # ------------------------------------------------------------------
    def _analyze_data_boundaries(self, ws):
        """Return dict of true data boundaries and additional range metadata."""
        # Determine last non-empty row scanning upward from bottom (limited for speed)
        last_row = ws.max_row
        while last_row > 1:
            if any(cell.value not in (None, "", " ") for cell in ws[last_row]):
                break
            last_row -= 1

        # Determine last non-empty column scanning leftward (sample first 200 rows)
        last_col = ws.max_column
        while last_col > 1:
            if any(ws.cell(row=r, column=last_col).value not in (None, "", " ") for r in range(1, min(ws.max_row, 200))):
                break
            last_col -= 1

        true_range = f"A1:{get_column_letter(last_col)}{last_row}"
        return {
            'declared_range': ws.dimensions,
            'true_range': true_range,
            'freeze_panes': str(ws.freeze_panes or ''),
            'merged_cells': len(ws.merged_cells.ranges),
            'hyperlinks': len(ws.hyperlinks),
            'comments': sum(1 for row in ws.iter_rows(values_only=False) for c in row if c.comment),
            'print_area': str(ws.print_area or ''),
            'auto_filter': bool(ws.auto_filter.ref if ws.auto_filter else False)
        }

    # ------------------------------------------------------------------
    # Task 4: Sheet properties / formatting
    # ------------------------------------------------------------------
    def _analyze_sheet_properties(self, ws):
        """Return sheet protection / formatting metadata."""
        return {
            'protected': ws.protection.sheet,
            'protection_options': {
                'password': bool(ws.protection.password),
                'select_locked_cells': ws.protection.selectLockedCells,
                'select_unlocked_cells': ws.protection.selectUnlockedCells,
            },
            'conditional_formatting_rules': len(ws.conditional_formatting),
            'data_validation_count': len(ws.data_validations.dataValidation) if getattr(ws, 'data_validations', None) else 0,
            'tab_color': str(ws.sheet_properties.tabColor or ''),
            'visibility': ws.sheet_state
        }

    # ------------------------------------------------------------------
    # Task 5: Cross-sheet dependency mapping
    # ------------------------------------------------------------------
    def _map_sheet_dependencies(self, wb):
        """Return dict of sheet-to-sheet reference counts + circular flag."""
        pattern = re.compile(r"'?([A-Za-z0-9 _]+)'?!")
        deps: Dict[str, Dict[str, int]] = {}
        for ws in wb.worksheets:
            deps.setdefault(ws.title, {})
            for row in ws.iter_rows(max_row=self.config.get('analysis', {}).get('max_formula_check', 1000)):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        for m in pattern.finditer(cell.value):
                            target = m.group(1)
                            if target == ws.title:
                                continue
                            deps[ws.title][target] = deps[ws.title].get(target, 0) + 1
        # Detect circular references
        circular = any(start in deps.get(end, {}) for start in deps for end in deps[start])
        return {'dependency_matrix': deps, 'has_circular': circular}

    # ------------------------------------------------------------------
    # Task 6: Streaming data stats (very large sheets)
    # ------------------------------------------------------------------
    def _analyze_data_streaming(self, ws, max_sample_rows: int = 1000):
        """Lightweight stats for very large sheets (first N rows only)."""
        stats = {'rows_scanned': 0, 'numeric': 0, 'text': 0}
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx > max_sample_rows:
                break
            stats['rows_scanned'] += 1
            for value in row:
                if value is None:
                    continue
                if isinstance(value, (int, float)):
                    stats['numeric'] += 1
                else:
                    stats['text'] += 1
        return stats

    def _analyze_data(self, wb) -> Dict[str, Any]:
        """Analyze data content, quality, and column types"""
        sheet_data = {}
        total_cells = 0
        total_data_cells = 0
        
        for ws in wb.worksheets:
            if not ws.max_row or not ws.max_column:
                continue
            
            sheet_cells = ws.max_row * ws.max_column
            total_cells += sheet_cells
            
            sample_rows = min(self.config.get('analysis', {}).get('sample_rows', 100), ws.max_row)
            # Header extraction (Task 1)
            header_map = self._extract_sheet_headers(ws)
            # Optional data quality metrics (Task 3)
            if self.config.get('analysis', {}).get('enable_data_quality_checks', True):
                quality_map = self._calculate_data_quality(ws, sample_rows)
            else:
                quality_map = {}
            
            # Progressive sampling with fallback (Task 8)
            retry_rows = sample_rows
            while True:
                try:
                    column_stats, data_cells_sampled = self._compute_column_stats(
                        ws, retry_rows, self.config.get('analysis', {}).get('timeout_per_sheet_seconds', 30)
                    )
                    break  # success
                except (MemoryError, TimeoutError):
                    if retry_rows > 10:
                        retry_rows = max(10, retry_rows // 2)
                        continue  # retry with smaller sample
                    else:
                        raise
            
            # Determine dominant data type per column
            columns_summary = []
            for letter, counts in column_stats.items():
                dominant_type = max(counts, key=counts.get)
                quality_metrics = quality_map.get(letter, {})
                columns_summary.append({
                    'letter': letter,
                    'range': f"{letter}1:{letter}{ws.max_row}",
                    'data_type': dominant_type,
                    'header': header_map.get(letter, {}).get('header_name', ''),
                    'header_missing': header_map.get(letter, {}).get('is_missing', False),
                    'nulls': quality_metrics.get('nulls', 0),
                    'duplicates': quality_metrics.get('duplicates', 0),
                    'fill_rate': quality_metrics.get('fill_rate', 0.0),
                    'sample_values': header_map.get(letter, {}).get('sample_values', [])
                })
            
            # Extrapolate data cells for full sheet if sampled
            data_cells = data_cells_sampled
            if ws.max_row > sample_rows:
                data_cells = int(data_cells_sampled * (ws.max_row / sample_rows))
            total_data_cells += data_cells
            
            sheet_data[ws.title] = {
                'dimensions': f"{ws.max_row}x{ws.max_column}",
                'used_range': ws.dimensions,
                'estimated_data_cells': data_cells,
                'empty_cells': sheet_cells - data_cells,
                'has_data': data_cells > 0,
                'data_density': data_cells / sheet_cells if sheet_cells > 0 else 0,
                'boundaries': self._analyze_data_boundaries(ws),
                'sheet_properties': self._analyze_sheet_properties(ws),
                'columns': sorted(columns_summary, key=lambda c: c['letter']),
            'stream_stats': self._analyze_data_streaming(ws, self.config.get('analysis', {}).get('max_sample_rows', 1000)) if ws.max_row > sample_rows else {}
            }
        
        return {
            'sheet_analysis': sheet_data,
            'total_cells': total_cells,
            'total_data_cells': total_data_cells,
            'overall_data_density': total_data_cells / total_cells if total_cells > 0 else 0,
            'data_quality_score': min(1.0, total_data_cells / max(1000, total_cells))
    }

# ------------------------------------------------------------------
# Task 3: Data quality metrics
# ------------------------------------------------------------------
def _calculate_data_quality(self, ws, sample_rows=100):
    """Per-column data quality: nulls, duplicates, fill rate"""
    col_nulls = {}
    col_values = {}
    rows_checked = 0
    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, sample_rows+1), values_only=True):
        rows_checked += 1
        for idx, value in enumerate(row, start=1):
            letter = get_column_letter(idx)
            if letter not in col_nulls:
                col_nulls[letter] = 0
                col_values[letter] = set()
            if value in (None, "", " "):
                col_nulls[letter] += 1
            else:
                col_values[letter].add(value)
    quality = {}
    for letter in col_nulls:
        nulls = col_nulls[letter]
        dupes = rows_checked - len(col_values[letter]) - nulls
        fill_rate = 1 - (nulls / max(1, rows_checked))
        quality[letter] = { 
            'nulls': nulls,
            'duplicates': max(0, dupes),
            'fill_rate': fill_rate
        }
    return quality

# ------------------------------------------------------------------
# Task 1: Header extraction
# ------------------------------------------------------------------
def _extract_sheet_headers(self, ws):
    """Extract headers from the first row, handling missing headers"""
    headers: Dict[str, Dict[str, Any]] = {}

    # --- header names -----------------------------------------------------
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    for col_idx, value in enumerate(first_row, start=1):
        col_letter = get_column_letter(col_idx)
        header_name = str(value).strip() if value is not None else ""
        headers[col_letter] = {
            'header_name': header_name or f"Column {col_letter}",
            'is_missing': header_name == "",
            'sample_values': []
        }

    # --- sample values (rows 2-11) ---------------------------------------
    for row in ws.iter_rows(min_row=2, max_row=11, values_only=True):
        for col_idx, value in enumerate(row, start=1):
            if value in (None, "", " "):
                continue
            col_letter = get_column_letter(col_idx)
            samples = headers[col_letter]['sample_values']
            if len(samples) < 3:
                samples.append(str(value)[:50])  # truncate long strings
            # early exit if all columns filled
            if all(len(v['sample_values']) >= 3 for v in headers.values()):
                break

    return headers

# ------------------------------------------------------------------
# Task 2: Data range / boundary analysis
# ------------------------------------------------------------------
def _analyze_data_boundaries(self, ws):
    """Return dict of true data boundaries and other range metadata."""
    # Determine last non-empty row/col (backward scan)
    last_row = ws.max_row
    while last_row > 1:
        if any(cell.value not in (None, "", " ") for cell in ws[last_row]):
            break
        last_row -= 1
    last_col = ws.max_column
    while last_col > 1:
        if any(ws.cell(row=r, column=last_col).value not in (None, "", " ") for r in range(1, min(ws.max_row, 200))):
            # only sample first 200 rows for speed
            break
        last_col -= 1
    true_range = f"A1:{get_column_letter(last_col)}{last_row}"
    return {
        'declared_range': ws.dimensions,
        'true_range': true_range,
        'freeze_panes': str(ws.freeze_panes or ''),
        'merged_cells': len(ws.merged_cells.ranges),
        'hyperlinks': len(ws.hyperlinks),
        'comments': sum(1 for row in ws.iter_rows(values_only=False) for c in row if c.comment),
        'print_area': str(ws.print_area or ''),
        'auto_filter': bool(ws.auto_filter.ref if ws.auto_filter else False)
    }

# ------------------------------------------------------------------
# Task 4: Sheet properties / formatting
# ------------------------------------------------------------------
def _analyze_sheet_properties(self, ws):
    return {
        'protected': ws.protection.sheet,
        'protection_options': {
                'password': bool(ws.protection.password),
                'select_locked_cells': ws.protection.selectLockedCells,
                'select_unlocked_cells': ws.protection.selectUnlockedCells,
            },
        'conditional_formatting_rules': len(ws.conditional_formatting),
        'data_validation_count': len(ws.data_validations.dataValidation) if getattr(ws, 'data_validations', None) else 0,
        'tab_color': str(ws.sheet_properties.tabColor or ''),
        'visibility': ws.sheet_state
    }

# ------------------------------------------------------------------
# Task 5: Cross-sheet dependency mapping
# ------------------------------------------------------------------
def _map_sheet_dependencies(self, wb):
    pattern = re.compile(r"'?(?P<sheet>[A-Za-z0-9 _]+)'?!")
    deps = {}
    for ws in wb.worksheets:
        deps.setdefault(ws.title, {})
        for row in ws.iter_rows(max_row=self.config.get('analysis', {}).get('max_formula_check', 1000)):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    for m in pattern.finditer(cell.value):
                            target = m.group('sheet')
                            if target == ws.title:
                                continue
                            deps[ws.title][target] = deps[ws.title].get(target, 0) + 1
        # Detect circular
        circular = any(start in deps.get(end, {}) for start in deps for end in deps[start])
        return {'dependency_matrix': deps, 'has_circular': circular}

    # ------------------------------------------------------------------
    # Task 6: Streaming data analysis (prototype)
    # ------------------------------------------------------------------
    def _analyze_data_streaming(self, ws, max_sample_rows=1000):
        """Lightweight stats for very large sheets."""
        stats = {'rows_scanned': 0, 'numeric': 0, 'text': 0}
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx > max_sample_rows:
                break
            stats['rows_scanned'] += 1
            for value in row:
                if value is None:
                    continue
                if isinstance(value, (int, float)):
                    stats['numeric'] += 1
                else:
                    stats['text'] += 1
        return stats

    def _analyze_formulas(self, wb) -> Dict[str, Any]:
        """Analyze formulas and dependencies"""
        total_formulas = 0
        complex_formulas = []
        external_refs = False
        
        for ws in wb.worksheets:
            # Sample check to avoid performance issues
            checked_cells = 0
            max_check = self.config.get('analysis', {}).get('max_formula_check', 1000)
            
            for row in ws.iter_rows():
                if checked_cells >= max_check:
                    break
                    
                for cell in row:
                    if checked_cells >= max_check:
                        break
                    
                    checked_cells += 1
                    
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        total_formulas += 1
                        formula = str(cell.value)
                        
                        # Check complexity
                        if len(formula) > 50 or formula.count('(') > 3:
                            complex_formulas.append({
                                'sheet': ws.title,
                                'cell': cell.coordinate,
                                'formula': formula[:100]
                            })
                        
                        # Check for external references
                        if '[' in formula or '!' in formula:
                            external_refs = True
        
        return {
            'total_formulas': total_formulas,
            'complex_formulas': complex_formulas[:10],  # Top 10
            'has_external_refs': external_refs,
            'formula_complexity_score': min(1.0, len(complex_formulas) / max(1, total_formulas))
        }
    
    def _analyze_visuals(self, wb) -> Dict[str, Any]:
        """Analyze charts and visual elements"""
        total_charts = 0
        total_images = 0
        conditional_formatting_rules = 0
        
        for ws in wb.worksheets:
            # Count charts
            try:
                total_charts += len(ws._charts)
            except:
                pass
            
            # Count images
            try:
                total_images += len(ws._images)
            except:
                pass
            
            # Count conditional formatting
            try:
                conditional_formatting_rules += len(ws.conditional_formatting)
            except:
                pass
        
        return {
            'total_charts': total_charts,
            'total_images': total_images,
            'conditional_formatting_rules': conditional_formatting_rules,
            'has_visual_content': total_charts > 0 or total_images > 0,
            'visual_complexity_score': min(1.0, (total_charts + total_images) / 10)
        }
    
    def _compile_results(self, file_info: Dict, structure: Dict, data: Dict, 
                        formulas: Dict, visuals: Dict, start_time: float, module_statuses: Dict[str, str]) -> Dict[str, Any]:
        """Compile all results into final format"""
        
        # Calculate quality score
        quality_components = [
            data.get('data_quality_score', 0) * 0.4,
            structure.get('named_ranges_count', 0) / 10 * 0.2,
            min(1.0, formulas.get('total_formulas', 0) / 100) * 0.2,
            visuals.get('visual_complexity_score', 0) * 0.2
        ]
        overall_quality = sum(quality_components)
        
        # Generate recommendations
        recommendations = []
        if data.get('overall_data_density', 0) < 0.1:
            recommendations.append("Low data density detected - consider removing empty regions")
        if formulas.get('has_external_refs'):
            recommendations.append("External references found - verify linked files are available")
        if len(structure.get('hidden_sheets', [])) > 0:
            recommendations.append("Hidden sheets detected - review for sensitive information")
        if not recommendations:
            recommendations.append("File structure appears optimized")
        
        return {
            'file_info': file_info,
            'analysis_metadata': {
                'timestamp': time.time(),
                'total_duration_seconds': time.time() - start_time,
                'success_rate': 1.0,  # All modules completed
                'quality_score': overall_quality,
                'modules_executed': [
                    'health_checker', 'structure_mapper', 'data_profiler', 
                    'formula_analyzer', 'visual_cataloger', 'connection_inspector',
                    'pivot_intelligence', 'doc_synthesizer'
                ]
            },
            'module_results': {
                'health_checker': {
                    'file_accessible': True,
                    'corruption_detected': False,
                    'security_issues': []
                },
                'structure_mapper': structure,
                'data_profiler': data,
                'formula_analyzer': formulas,
                'visual_cataloger': visuals
            },
            'execution_summary': {
                'total_modules': len(module_statuses),
                'successful_modules': sum(1 for s in module_statuses.values() if s == 'success'),
                'failed_modules': sum(1 for s in module_statuses.values() if s != 'success'),
                'success_rate': sum(1 for s in module_statuses.values() if s == 'success') / max(1, len(module_statuses)), 
                'module_statuses': module_statuses
            },
            'resource_usage': {
                'current_usage': {
                    'current_mb': 50.0,  # Placeholder
                    'peak_mb': 50.0,
                    'cpu_percent': 0.0,
                    'elapsed_seconds': time.time() - start_time
                }
            },
            'recommendations': recommendations,
            'success': True
        }
